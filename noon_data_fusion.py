"""
noon_data_fusion.py — Multi-source data fusion tool

Merges data from multiple sources into a unified product list:
  - Enhanced report ("Cleaned Products") or raw report ("All Products") as base
  - Price tracking data (by product_id): latest_price, price_change_pct, sales_velocity
  - Seller analysis data (by product_id or subcategory): seller_name, seller_competition_level
Recomputes composite_score with time-series acceleration bonus.

Usage:
  python noon_data_fusion.py \
    --report report_automotive.xlsx \
    --enhanced report_automotive_enhanced.xlsx \
    --tracking tracking_dir/ \
    --sellers seller_analysis.xlsx \
    --output unified_automotive.xlsx
"""

import argparse
import logging
import os
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import CONFIG and calc_full_profit from noon_config
sys.path.insert(0, 'D:/claude noon')
from noon_config import CONFIG, calc_full_profit

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("data_fusion")

# ============================================================
# Excel styling constants
# ============================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


# ============================================================
# Data loading functions
# ============================================================

def load_base_data(report_path: str | None, enhanced_path: str | None) -> tuple[pd.DataFrame, str]:
    """
    Load base product data.
    Priority: enhanced "Cleaned Products" > raw "All Products".
    Returns (DataFrame, source_name).
    """
    # Try enhanced report first
    if enhanced_path and Path(enhanced_path).exists():
        try:
            df = pd.read_excel(enhanced_path, sheet_name="Cleaned Products")
            logger.info(f"Loaded enhanced 'Cleaned Products': {len(df)} rows, {len(df.columns)} cols")
            return df, "Enhanced Report (Cleaned Products)"
        except Exception as e:
            logger.warning(f"Failed to read 'Cleaned Products' from enhanced report: {e}")

    # Fall back to raw report
    if report_path and Path(report_path).exists():
        try:
            df = pd.read_excel(report_path, sheet_name="All Products")
            logger.info(f"Loaded raw 'All Products': {len(df)} rows, {len(df.columns)} cols")
            return df, "Raw Report (All Products)"
        except Exception as e:
            logger.warning(f"Failed to read 'All Products' from raw report: {e}")

    # Nothing available
    return pd.DataFrame(), "None"


def load_tracking_data(tracking_path: str | None) -> pd.DataFrame:
    """
    Load price tracking data from a directory of snapshot Excel/CSV files
    or a single file. Expected columns: product_id, price, timestamp/date.
    Computes: latest_price, price_change_pct, sales_velocity.
    """
    if not tracking_path:
        return pd.DataFrame()

    path = Path(tracking_path)
    if not path.exists():
        logger.warning(f"Tracking path does not exist: {tracking_path}")
        return pd.DataFrame()

    frames = []

    if path.is_dir():
        # Load all xlsx/csv files from directory
        for f in sorted(path.glob("*.xlsx")) + sorted(path.glob("*.csv")):
            try:
                if f.suffix == ".csv":
                    tmp = pd.read_csv(f)
                else:
                    tmp = pd.read_excel(f)
                # Try to extract date from filename
                fname = f.stem
                tmp["_source_file"] = fname
                frames.append(tmp)
            except Exception as e:
                logger.warning(f"Failed to load tracking file {f.name}: {e}")
    elif path.is_file():
        try:
            if path.suffix == ".csv":
                frames.append(pd.read_csv(path))
            else:
                frames.append(pd.read_excel(path))
        except Exception as e:
            logger.warning(f"Failed to load tracking file {path.name}: {e}")

    if not frames:
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)
    logger.info(f"Loaded tracking data: {len(df)} rows from {len(frames)} file(s)")

    # Require product_id
    if "product_id" not in df.columns:
        logger.warning("Tracking data missing 'product_id' column, skipping")
        return pd.DataFrame()

    # Detect price column
    price_col = None
    for candidate in ["price", "latest_price", "current_price", "selling_price"]:
        if candidate in df.columns:
            price_col = candidate
            break
    if not price_col:
        logger.warning("Tracking data has no recognizable price column, skipping")
        return pd.DataFrame()

    # Detect date/timestamp column
    date_col = None
    for candidate in ["timestamp", "date", "crawl_date", "snapshot_date", "_source_file"]:
        if candidate in df.columns:
            date_col = candidate
            break

    # Compute per-product tracking summary
    df[price_col] = pd.to_numeric(df[price_col], errors="coerce")
    df = df.dropna(subset=[price_col])

    # Detect sold_recently / sales column
    sales_col = None
    for candidate in ["sold_recently", "sales", "units_sold", "sold_count"]:
        if candidate in df.columns:
            sales_col = candidate
            break

    summary_rows = []
    for pid, group in df.groupby("product_id"):
        group_sorted = group.sort_index()  # assume chronological order
        prices = group_sorted[price_col].tolist()
        latest = prices[-1]
        earliest = prices[0]
        change_pct = ((latest - earliest) / earliest * 100) if earliest > 0 else 0

        # Sales velocity: average sold_recently across snapshots
        velocity = None
        if sales_col and sales_col in group.columns:
            sales_vals = pd.to_numeric(group[sales_col], errors="coerce").dropna()
            if len(sales_vals) > 0:
                velocity = sales_vals.mean()

        # Sales acceleration: is recent velocity > earlier velocity?
        acceleration = 0
        if sales_col and sales_col in group.columns and len(group) >= 2:
            sales_vals = pd.to_numeric(group_sorted[sales_col], errors="coerce").dropna()
            if len(sales_vals) >= 2:
                mid = len(sales_vals) // 2
                earlier_avg = sales_vals.iloc[:mid].mean()
                recent_avg = sales_vals.iloc[mid:].mean()
                if earlier_avg > 0:
                    acceleration = (recent_avg - earlier_avg) / earlier_avg

        summary_rows.append({
            "product_id": pid,
            "latest_price": round(latest, 2),
            "price_change_pct": round(change_pct, 2),
            "price_snapshots": len(prices),
            "sales_velocity": round(velocity, 1) if velocity is not None else None,
            "sales_acceleration": round(acceleration, 3),
        })

    result = pd.DataFrame(summary_rows)
    logger.info(f"Tracking summary: {len(result)} unique products")
    return result


def load_seller_data(sellers_path: str | None) -> pd.DataFrame:
    """
    Load seller analysis data.
    Expected columns: product_id or subcategory, seller_name, seller_competition_level, etc.
    """
    if not sellers_path:
        return pd.DataFrame()

    path = Path(sellers_path)
    if not path.exists():
        logger.warning(f"Seller file does not exist: {sellers_path}")
        return pd.DataFrame()

    try:
        if path.suffix == ".csv":
            df = pd.read_csv(path)
        else:
            # Try common sheet names
            for sheet in [None, "Sellers", "sellers", "Sheet1", "Seller Analysis"]:
                try:
                    df = pd.read_excel(path, sheet_name=sheet)
                    break
                except Exception:
                    continue
            else:
                df = pd.read_excel(path)

        logger.info(f"Loaded seller data: {len(df)} rows, columns: {list(df.columns)}")
        return df
    except Exception as e:
        logger.warning(f"Failed to load seller data: {e}")
        return pd.DataFrame()


# ============================================================
# Fusion logic
# ============================================================

def merge_tracking(base_df: pd.DataFrame, tracking_df: pd.DataFrame) -> pd.DataFrame:
    """LEFT JOIN tracking data by product_id."""
    if tracking_df.empty or "product_id" not in base_df.columns:
        return base_df

    tracking_cols = ["product_id", "latest_price", "price_change_pct",
                     "price_snapshots", "sales_velocity", "sales_acceleration"]
    available_cols = [c for c in tracking_cols if c in tracking_df.columns]
    if len(available_cols) <= 1:
        return base_df

    merged = base_df.merge(
        tracking_df[available_cols],
        on="product_id",
        how="left",
        suffixes=("", "_tracking"),
    )
    joined = merged["latest_price"].notna().sum() if "latest_price" in merged.columns else 0
    logger.info(f"Tracking merge: {joined}/{len(base_df)} products matched")
    return merged


def merge_sellers(base_df: pd.DataFrame, seller_df: pd.DataFrame) -> pd.DataFrame:
    """LEFT JOIN seller data by product_id, falling back to subcategory."""
    if seller_df.empty:
        return base_df

    # Strategy 1: merge by product_id
    if "product_id" in seller_df.columns and "product_id" in base_df.columns:
        seller_cols = [c for c in seller_df.columns
                       if c == "product_id" or c.startswith("seller")]
        if len(seller_cols) > 1:
            merged = base_df.merge(
                seller_df[seller_cols].drop_duplicates("product_id"),
                on="product_id",
                how="left",
                suffixes=("", "_seller"),
            )
            matched = sum(merged[[c for c in seller_cols if c != "product_id"][0]].notna())
            logger.info(f"Seller merge (by product_id): {matched}/{len(base_df)} matched")
            return merged

    # Strategy 2: merge by subcategory
    sub_col_base = "_subcategory" if "_subcategory" in base_df.columns else "subcategory"
    sub_col_seller = None
    for candidate in ["_subcategory", "subcategory", "Subcategory", "category"]:
        if candidate in seller_df.columns:
            sub_col_seller = candidate
            break

    if sub_col_base in base_df.columns and sub_col_seller:
        seller_sub = seller_df.rename(columns={sub_col_seller: sub_col_base})
        seller_cols = [c for c in seller_sub.columns
                       if c == sub_col_base or c.startswith("seller")]
        if len(seller_cols) > 1:
            # Aggregate to subcategory level (take first)
            seller_agg = seller_sub[seller_cols].drop_duplicates(sub_col_base)
            merged = base_df.merge(
                seller_agg,
                on=sub_col_base,
                how="left",
                suffixes=("", "_seller"),
            )
            matched_col = [c for c in seller_cols if c != sub_col_base][0]
            matched = merged[matched_col].notna().sum()
            logger.info(f"Seller merge (by subcategory): {matched}/{len(base_df)} matched")
            return merged

    logger.warning("Could not find join key for seller data (need product_id or subcategory)")
    return base_df


def compute_composite_score(df: pd.DataFrame) -> pd.DataFrame:
    """
    Recompute composite_score incorporating all available dimensions.
    Products with sales acceleration get a bonus.
    """
    df = df.copy()

    # --- Profit component (0-40 points) ---
    margin_col = "margin_war_pct"
    if margin_col in df.columns:
        margin = pd.to_numeric(df[margin_col], errors="coerce").fillna(0)
        # Scale: -30% -> 0, 0% -> 15, 30%+ -> 40
        df["_score_profit"] = ((margin + 30) / 60 * 40).clip(0, 40)
    else:
        df["_score_profit"] = 20  # neutral

    # --- Demand component (0-25 points) ---
    # Use sold_recently, review_count, sales_velocity
    demand_signals = pd.Series(0.0, index=df.index)
    if "sold_recently" in df.columns:
        sr = pd.to_numeric(df["sold_recently"], errors="coerce").fillna(0)
        # Normalize: top 1% cap
        cap = sr.quantile(0.99) if sr.max() > 0 else 1
        demand_signals += (sr / max(cap, 1)).clip(0, 1) * 12
    if "review_count" in df.columns:
        rc = pd.to_numeric(df["review_count"], errors="coerce").fillna(0)
        cap = rc.quantile(0.99) if rc.max() > 0 else 1
        demand_signals += (rc / max(cap, 1)).clip(0, 1) * 8
    if "sales_velocity" in df.columns:
        sv = pd.to_numeric(df["sales_velocity"], errors="coerce").fillna(0)
        cap = sv.quantile(0.99) if sv.max() > 0 else 1
        demand_signals += (sv / max(cap, 1)).clip(0, 1) * 5
    df["_score_demand"] = demand_signals.clip(0, 25)

    # --- Competition component (0-20 points, lower competition = higher score) ---
    if "competition_score" in df.columns:
        # competition_score from enhanced report (higher = more competitive)
        cs = pd.to_numeric(df["competition_score"], errors="coerce").fillna(5)
        # Invert: 10 -> 0 points, 0 -> 20 points
        df["_score_competition"] = ((10 - cs) / 10 * 20).clip(0, 20)
    elif "is_express" in df.columns:
        # Proxy: less express = less competition
        exp = df["is_express"].apply(lambda x: 1 if x in [True, "True", "Yes", 1] else 0)
        df["_score_competition"] = (1 - exp) * 12 + 4  # non-express gets 16, express gets 4
    else:
        df["_score_competition"] = 10

    # --- Price stability component (0-15 points) ---
    if "price_change_pct" in df.columns:
        pc = pd.to_numeric(df["price_change_pct"], errors="coerce").fillna(0).abs()
        # Stable prices score higher: 0% change -> 15, 50%+ change -> 0
        df["_score_price_stability"] = ((50 - pc) / 50 * 15).clip(0, 15)
    else:
        df["_score_price_stability"] = 7.5

    # --- Acceleration bonus (0-10 points) ---
    if "sales_acceleration" in df.columns:
        acc = pd.to_numeric(df["sales_acceleration"], errors="coerce").fillna(0)
        # Positive acceleration -> bonus; >50% growth -> max bonus
        df["_score_acceleration"] = (acc / 0.5 * 10).clip(0, 10)
    else:
        df["_score_acceleration"] = 0

    # --- Composite score (0-100 + up to 10 bonus) ---
    df["composite_score"] = (
        df["_score_profit"]
        + df["_score_demand"]
        + df["_score_competition"]
        + df["_score_price_stability"]
        + df["_score_acceleration"]
    ).round(2)

    # Clean up intermediate columns
    score_cols = [c for c in df.columns if c.startswith("_score_")]
    df.drop(columns=score_cols, inplace=True)

    logger.info(
        f"Composite score computed: "
        f"mean={df['composite_score'].mean():.1f}, "
        f"median={df['composite_score'].median():.1f}, "
        f"max={df['composite_score'].max():.1f}"
    )
    return df


# ============================================================
# Data coverage summary
# ============================================================

def build_coverage_summary(
    base_df: pd.DataFrame,
    base_source: str,
    tracking_df: pd.DataFrame,
    seller_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build a summary of which data sources contributed what."""
    total = len(base_df)
    rows = []

    rows.append({
        "Data Source": "Base Products",
        "Source File": base_source,
        "Records Available": total,
        "Coverage %": 100.0,
        "Key Columns Added": "product_id, title, price, margin_war_pct, ...",
    })

    if not tracking_df.empty and "latest_price" in base_df.columns:
        matched = base_df["latest_price"].notna().sum()
        rows.append({
            "Data Source": "Price Tracking",
            "Source File": "tracking directory",
            "Records Available": len(tracking_df) if not tracking_df.empty else 0,
            "Coverage %": round(matched / total * 100, 1) if total > 0 else 0,
            "Key Columns Added": "latest_price, price_change_pct, sales_velocity, sales_acceleration",
        })
    else:
        rows.append({
            "Data Source": "Price Tracking",
            "Source File": "(not provided)",
            "Records Available": 0,
            "Coverage %": 0,
            "Key Columns Added": "N/A",
        })

    seller_cols_in_base = [c for c in base_df.columns if c.startswith("seller")]
    if seller_cols_in_base:
        first_sc = seller_cols_in_base[0]
        matched = base_df[first_sc].notna().sum()
        rows.append({
            "Data Source": "Seller Analysis",
            "Source File": "seller file",
            "Records Available": len(seller_df) if not seller_df.empty else 0,
            "Coverage %": round(matched / total * 100, 1) if total > 0 else 0,
            "Key Columns Added": ", ".join(seller_cols_in_base),
        })
    else:
        rows.append({
            "Data Source": "Seller Analysis",
            "Source File": "(not provided)",
            "Records Available": 0,
            "Coverage %": 0,
            "Key Columns Added": "N/A",
        })

    # Enhanced report score columns
    enhanced_cols = ["competition_score", "demand_score", "profit_score",
                     "opportunity_score", "operability_score"]
    present = [c for c in enhanced_cols if c in base_df.columns]
    if present:
        matched = base_df[present[0]].notna().sum()
        rows.append({
            "Data Source": "Enhanced Scores",
            "Source File": "enhanced report",
            "Records Available": matched,
            "Coverage %": round(matched / total * 100, 1) if total > 0 else 0,
            "Key Columns Added": ", ".join(present),
        })

    # Composite score
    if "composite_score" in base_df.columns:
        matched = base_df["composite_score"].notna().sum()
        rows.append({
            "Data Source": "Composite Score (Fused)",
            "Source File": "computed from all sources",
            "Records Available": matched,
            "Coverage %": round(matched / total * 100, 1) if total > 0 else 0,
            "Key Columns Added": "composite_score",
        })

    return pd.DataFrame(rows)


# ============================================================
# Excel output
# ============================================================

def _style_header(ws, col_count: int):
    """Apply dark blue header style."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, max_width: int = 28):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _write_cell_safe(ws, row, col, value):
    """Write a value to a cell, handling type issues."""
    try:
        if pd.isna(value):
            return
    except (ValueError, TypeError):
        pass

    try:
        if isinstance(value, (list, dict)):
            ws.cell(row=row, column=col, value=str(value))
        elif isinstance(value, bool):
            ws.cell(row=row, column=col, value="Yes" if value else "")
        else:
            ws.cell(row=row, column=col, value=value)
    except (ValueError, TypeError):
        ws.cell(row=row, column=col, value=str(value))


def write_unified_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 1: Unified Products — all products with all dimensions."""
    ws = wb.create_sheet("Unified Products", 0)

    # Define column order — put most useful columns first
    priority_cols = [
        "product_id", "title", "_subcategory", "price", "latest_price",
        "price_change_pct", "margin_war_pct", "margin_peace_pct",
        "net_profit_war_cny", "composite_score",
        "sold_recently", "sales_velocity", "sales_acceleration",
        "review_count", "rating", "is_express", "is_bestseller",
        "competition_score", "demand_score", "profit_score",
        "opportunity_score", "operability_score",
        "v6_category", "size_tier", "purchase_est_cny",
        "fbn_fee_sar", "commission_pct", "price_band",
        "seller_name", "seller_competition_level",
        "price_snapshots", "image_count", "original_price",
        "is_ad", "has_sales", "product_url",
    ]

    # Build final column list: priority first, then remaining
    all_cols = list(df.columns)
    ordered_cols = [c for c in priority_cols if c in all_cols]
    remaining = [c for c in all_cols if c not in ordered_cols]
    final_cols = ordered_cols + remaining

    # Write header
    for c, col_name in enumerate(final_cols, 1):
        ws.cell(row=1, column=c, value=col_name)

    # Write data (cap at 200k rows for performance)
    max_rows = min(len(df), 200000)
    if len(df) > max_rows:
        logger.warning(f"Truncating Unified Products to {max_rows} rows (total: {len(df)})")

    for r_idx, (_, row) in enumerate(df.head(max_rows).iterrows(), 2):
        for c_idx, col_name in enumerate(final_cols, 1):
            _write_cell_safe(ws, r_idx, c_idx, row.get(col_name))

        # Color-code margin
        if "margin_war_pct" in final_cols:
            margin_c = final_cols.index("margin_war_pct") + 1
            val = row.get("margin_war_pct")
            try:
                if val is not None and not pd.isna(val):
                    if float(val) >= 30:
                        ws.cell(row=r_idx, column=margin_c).fill = GREEN_FILL
                    elif float(val) < 0:
                        ws.cell(row=r_idx, column=margin_c).fill = RED_FILL
            except (ValueError, TypeError):
                pass

    _style_header(ws, len(final_cols))
    _auto_width(ws)
    ws.freeze_panes = "A2"
    logger.info(f"Sheet 'Unified Products': {max_rows} rows, {len(final_cols)} cols")


def write_coverage_sheet(wb: Workbook, coverage_df: pd.DataFrame):
    """Sheet 2: Data Coverage — summary of data sources."""
    ws = wb.create_sheet("Data Coverage")

    cols = list(coverage_df.columns)
    for c, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=col_name)

    for r_idx, (_, row) in enumerate(coverage_df.iterrows(), 2):
        for c_idx, col_name in enumerate(cols, 1):
            _write_cell_safe(ws, r_idx, c_idx, row.get(col_name))

        # Color coverage %
        cov_col = cols.index("Coverage %") + 1 if "Coverage %" in cols else None
        if cov_col:
            val = row.get("Coverage %", 0)
            try:
                if val >= 80:
                    ws.cell(row=r_idx, column=cov_col).fill = GREEN_FILL
                elif val >= 30:
                    ws.cell(row=r_idx, column=cov_col).fill = YELLOW_FILL
                elif val > 0:
                    ws.cell(row=r_idx, column=cov_col).fill = RED_FILL
            except (ValueError, TypeError):
                pass

    _style_header(ws, len(cols))
    _auto_width(ws, max_width=50)
    ws.freeze_panes = "A2"


def write_top_opportunities_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 3: Top Opportunities — re-ranked top 100."""
    ws = wb.create_sheet("Top Opportunities")

    # Sort by composite_score descending
    if "composite_score" not in df.columns:
        logger.warning("No composite_score column, skipping Top Opportunities")
        return

    top = df.nlargest(100, "composite_score").copy()
    top.insert(0, "rank", range(1, len(top) + 1))

    cols_to_show = [
        ("rank", "#"),
        ("product_id", "Product ID"),
        ("title", "Title"),
        ("_subcategory", "Subcategory"),
        ("price", "Price (SAR)"),
        ("latest_price", "Latest Price"),
        ("price_change_pct", "Price Change %"),
        ("margin_war_pct", "Margin War %"),
        ("margin_peace_pct", "Margin Peace %"),
        ("net_profit_war_cny", "Net Profit CNY"),
        ("composite_score", "Composite Score"),
        ("sold_recently", "Sold Recently"),
        ("sales_velocity", "Sales Velocity"),
        ("sales_acceleration", "Acceleration"),
        ("review_count", "Reviews"),
        ("rating", "Rating"),
        ("is_express", "Express"),
        ("is_bestseller", "Bestseller"),
        ("competition_score", "Competition"),
        ("demand_score", "Demand"),
        ("profit_score", "Profit Score"),
        ("opportunity_score", "Opportunity"),
        ("seller_name", "Seller"),
        ("seller_competition_level", "Seller Competition"),
        ("v6_category", "V6 Category"),
        ("product_url", "URL"),
    ]

    # Filter to columns that actually exist
    available = [(k, h) for k, h in cols_to_show if k in top.columns]

    # Write header
    for c, (_, header) in enumerate(available, 1):
        ws.cell(row=1, column=c, value=header)

    # Write data
    for r_idx, (_, row) in enumerate(top.iterrows(), 2):
        for c_idx, (col_key, _) in enumerate(available, 1):
            _write_cell_safe(ws, r_idx, c_idx, row.get(col_key))

        # Color composite score
        cs_idx = next((i for i, (k, _) in enumerate(available, 1) if k == "composite_score"), None)
        if cs_idx:
            val = row.get("composite_score")
            try:
                if val is not None and not pd.isna(val):
                    if float(val) >= 70:
                        ws.cell(row=r_idx, column=cs_idx).fill = GREEN_FILL
                    elif float(val) >= 50:
                        ws.cell(row=r_idx, column=cs_idx).fill = BLUE_FILL
                    elif float(val) >= 30:
                        ws.cell(row=r_idx, column=cs_idx).fill = YELLOW_FILL
            except (ValueError, TypeError):
                pass

        # Color margin
        m_idx = next((i for i, (k, _) in enumerate(available, 1) if k == "margin_war_pct"), None)
        if m_idx:
            val = row.get("margin_war_pct")
            try:
                if val is not None and not pd.isna(val):
                    if float(val) >= 30:
                        ws.cell(row=r_idx, column=m_idx).fill = GREEN_FILL
                    elif float(val) < 0:
                        ws.cell(row=r_idx, column=m_idx).fill = RED_FILL
            except (ValueError, TypeError):
                pass

    _style_header(ws, len(available))
    _auto_width(ws, max_width=40)
    ws.freeze_panes = "A2"
    logger.info(f"Sheet 'Top Opportunities': {len(top)} rows")


def generate_output(
    df: pd.DataFrame,
    coverage_df: pd.DataFrame,
    output_path: str,
):
    """Generate the unified Excel output."""
    wb = Workbook()

    # Sheet 1: Unified Products
    write_unified_sheet(wb, df)

    # Sheet 2: Data Coverage
    write_coverage_sheet(wb, coverage_df)

    # Sheet 3: Top Opportunities
    write_top_opportunities_sheet(wb, df)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Ensure output directory exists
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(out_path)
    logger.info(f"Output saved: {out_path} ({out_path.stat().st_size / 1024 / 1024:.1f} MB)")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Noon multi-source data fusion tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Only raw report (most common scenario)
  python noon_data_fusion.py --report report_automotive.xlsx --output unified.xlsx

  # Enhanced + raw report
  python noon_data_fusion.py --report report.xlsx --enhanced enhanced.xlsx --output unified.xlsx

  # All sources
  python noon_data_fusion.py --report report.xlsx --enhanced enhanced.xlsx \\
    --tracking tracking_dir/ --sellers sellers.xlsx --output unified.xlsx
""",
    )
    parser.add_argument("--report", help="Raw category report (.xlsx)")
    parser.add_argument("--enhanced", help="Enhanced category report (.xlsx)")
    parser.add_argument("--tracking", help="Price tracking directory or file")
    parser.add_argument("--sellers", help="Seller analysis file (.xlsx or .csv)")
    parser.add_argument("--output", default="unified_output.xlsx",
                        help="Output file path (default: unified_output.xlsx)")

    args = parser.parse_args()

    # Validate: at least one input required
    if not args.report and not args.enhanced:
        parser.error("At least one of --report or --enhanced is required")

    print("=" * 60)
    print("  Noon Data Fusion Tool")
    print("=" * 60)
    print(f"  Report:   {args.report or '(not provided)'}")
    print(f"  Enhanced: {args.enhanced or '(not provided)'}")
    print(f"  Tracking: {args.tracking or '(not provided)'}")
    print(f"  Sellers:  {args.sellers or '(not provided)'}")
    print(f"  Output:   {args.output}")
    print("=" * 60)

    # Step 1: Load base data
    logger.info("Step 1: Loading base product data...")
    base_df, base_source = load_base_data(args.report, args.enhanced)
    if base_df.empty:
        logger.error("No base data loaded. Cannot proceed.")
        sys.exit(1)
    logger.info(f"Base data: {len(base_df)} products from '{base_source}'")

    # Step 2: Load supplementary sources
    logger.info("Step 2: Loading supplementary data sources...")
    tracking_df = load_tracking_data(args.tracking)
    seller_df = load_seller_data(args.sellers)

    # Step 3: Merge tracking data
    logger.info("Step 3: Merging tracking data...")
    merged_df = merge_tracking(base_df, tracking_df)

    # Step 4: Merge seller data
    logger.info("Step 4: Merging seller data...")
    merged_df = merge_sellers(merged_df, seller_df)

    # Step 5: Recompute composite score
    logger.info("Step 5: Computing composite score...")
    merged_df = compute_composite_score(merged_df)

    # Step 6: Build coverage summary
    logger.info("Step 6: Building data coverage summary...")
    coverage_df = build_coverage_summary(merged_df, base_source, tracking_df, seller_df)

    # Step 7: Generate output
    logger.info("Step 7: Generating Excel output...")
    generate_output(merged_df, coverage_df, args.output)

    # Print summary
    print()
    print("=" * 60)
    print("  Fusion Complete")
    print("=" * 60)
    print(f"  Total products:     {len(merged_df):,}")
    print(f"  Base source:        {base_source}")
    print(f"  Tracking coverage:  {tracking_df['product_id'].nunique() if not tracking_df.empty and 'product_id' in tracking_df.columns else 0:,} products")
    print(f"  Seller coverage:    {len(seller_df):,} records")
    if "composite_score" in merged_df.columns:
        print(f"  Composite score:    mean={merged_df['composite_score'].mean():.1f}, "
              f"median={merged_df['composite_score'].median():.1f}")
    if "margin_war_pct" in merged_df.columns:
        valid_margins = merged_df["margin_war_pct"].dropna()
        if len(valid_margins) > 0:
            profitable = (valid_margins > 0).sum()
            print(f"  Profitable products: {profitable:,} / {len(valid_margins):,} "
                  f"({profitable / len(valid_margins) * 100:.1f}%)")
    print(f"  Output:             {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()
