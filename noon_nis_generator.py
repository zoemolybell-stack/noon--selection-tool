"""
noon_nis_generator.py — NIS (Noon Item Setup) upload file generator for Noon Seller Lab.

Reads product data from CSV or selection report, generates compliant NIS .xlsx
files by filling the official NIS template while preserving metadata rows and
validation sheets.

Usage:
    # Batch mode from CSV
    python noon_nis_generator.py --input products.csv --nis-template NIS.xlsx --output upload_ready.xlsx

    # From selection report
    python noon_nis_generator.py --from-selection selection_report_automotive.xlsx --nis-template NIS.xlsx --output upload_batch1.xlsx

    # Interactive single product
    python noon_nis_generator.py --interactive --nis-template NIS.xlsx
"""

import sys
import os
import csv
import argparse
import re
from copy import copy

sys.path.insert(0, 'D:/claude noon')
from noon_config import CONFIG

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================
# Constants
# ============================================================

# NIS template_data: Rows 1-9 are metadata (openpyxl 1-indexed).
# Data starts at row 10.
DATA_START_ROW = 10

# Column mapping (1-indexed) matching NIS template
COL = {
    'family': 1,
    'product_type': 2,
    'product_subtype': 3,
    'seller_sku': 4,
    'brand': 5,
    'parent_group_key': 6,
    'parent_child_variation': 7,
    'size_variation': 8,
    'long_description_en': 9,
    'long_description_ar': 10,
    'product_title_en': 11,
    'product_title_ar': 12,
    'item_condition': 13,
    'image_url_1': 14,
    'image_url_2': 15,
    'image_url_3': 16,
    'image_url_4': 17,
    'image_url_5': 18,
    'image_url_6': 19,
    'image_url_7': 20,
    'feature_bullet_1_en': 21,
    'feature_bullet_2_en': 22,
    'feature_bullet_3_en': 23,
    'feature_bullet_4_en': 24,
    'feature_bullet_5_en': 25,
    'feature_bullet_1_ar': 26,
    'feature_bullet_2_ar': 27,
    'feature_bullet_3_ar': 28,
    'feature_bullet_4_ar': 29,
    'feature_bullet_5_ar': 30,
    'whats_in_the_box_en': 31,
    'whats_in_the_box_ar': 32,
    'vat_rate_ae': 33,
    'vat_rate_sa': 34,
    'vat_rate_eg': 35,
}

# Automotive subcategory → (family, product_type) mapping for --from-selection
AUTOMOTIVE_MAPPING = {
    # Engine
    'engine parts': ('fmautomotive', 'Engine Parts'),
    'engine oil': ('fmautomotive', 'Batteries & Lubricants'),
    'engine cooling': ('fmautomotive', 'Engine/Interior Cooling'),
    'engine/interior cooling': ('fmautomotive', 'Engine/Interior Cooling'),
    # Brake
    'brake parts': ('fmautomotive', 'Brake Parts'),
    'brake pads': ('fmautomotive', 'Brake Parts'),
    'brake discs': ('fmautomotive', 'Brake Parts'),
    # Electronics
    'car electronics': ('fmautomotive', 'Car Audio/Video'),
    'car audio/video': ('fmautomotive', 'Car Audio/Video'),
    'dash cameras': ('fmautomotive', 'Car Audio/Video'),
    'car audio': ('fmautomotive', 'Car Audio/Video'),
    # Interior
    'interior accessories': ('fmautomotive', 'Interior Accessories'),
    'car interior': ('fmautomotive', 'Interior Accessories'),
    'seat covers': ('fmautomotive', 'Interior Accessories'),
    'floor mats': ('fmautomotive', 'Interior Accessories'),
    # Exterior
    'exterior accessories': ('fmautomotive', 'Exterior Accessories'),
    'car exterior': ('fmautomotive', 'Exterior Accessories'),
    'car covers': ('fmautomotive', 'Exterior Accessories'),
    # Suspension
    'suspension parts': ('fmautomotive', 'Suspension Parts'),
    'leaf springs & parts': ('fmautomotive', 'Suspension Parts'),
    'shock absorbers': ('fmautomotive', 'Suspension Parts'),
    # Electrical & Lighting
    'electrical & lighting': ('fmautomotive', 'Electrical & Lighting'),
    'car lighting': ('fmautomotive', 'Electrical & Lighting'),
    'headlights': ('fmautomotive', 'Electrical & Lighting'),
    # Transmission
    'transmission parts': ('fmautomotive', 'Transmission Parts'),
    # Batteries
    'batteries & lubricants': ('fmautomotive', 'Batteries & Lubricants'),
    'car batteries': ('fmautomotive', 'Batteries & Lubricants'),
    # Body
    'body parts': ('fmautomotive', 'Body Parts'),
    # Wheels / Hubs
    'wheel bearings hubs': ('fmautomotive', 'Suspension Parts'),
    'wheel bearings & hubs': ('fmautomotive', 'Suspension Parts'),
    # Cleaning
    'cleaning accessories': ('fmautomotive', 'Cleaning Accessories'),
    'car cleaning': ('fmautomotive', 'Cleaning Accessories'),
    # Garage
    'garage equipment': ('fmautomotive', 'Garage Equipment'),
    # Motorbike
    'motorbike parts & accessories': ('fmautomotive', 'Motorbike Parts & Accessories'),
    'motorbike protective gear': ('fmautomotive', 'Motorbike Protective Gear'),
    # Gifts
    'gifts & merchandise': ('fmautomotive', 'Gifts & Merchandise'),
}

# Default fallback for unmapped automotive subcategories
AUTOMOTIVE_DEFAULT = ('fmautomotive', 'Cleaning Accessories')


# ============================================================
# Validation
# ============================================================

def load_pft_data(nis_template_path):
    """Load pft_data sheet: returns {family: [product_types...]}."""
    wb = openpyxl.load_workbook(nis_template_path, read_only=True)
    ws = wb['pft_data']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    pft = {}
    if not rows:
        return pft
    headers = rows[0]
    for col_idx, header in enumerate(headers):
        if header is None:
            continue
        family = str(header).strip()
        types = []
        for r in range(1, len(rows)):
            if col_idx < len(rows[r]) and rows[r][col_idx] is not None:
                types.append(str(rows[r][col_idx]).strip())
        pft[family] = types
    return pft


def validate_nis_entry(family, product_type, nis_template_path):
    """
    Validate family + product_type against pft_data.
    Returns (is_valid, suggestion_or_error).
    """
    wb = openpyxl.load_workbook(nis_template_path, read_only=True)
    ws = wb['pft_data']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    for col_idx, header in enumerate(rows[0]):
        if header is not None and str(header).strip() == family:
            valid_types = [
                str(rows[r][col_idx]).strip()
                for r in range(1, len(rows))
                if col_idx < len(rows[r]) and rows[r][col_idx] is not None
            ]
            if product_type in valid_types:
                return True, None
            from difflib import get_close_matches
            suggestions = get_close_matches(product_type, valid_types, n=3, cutoff=0.4)
            return False, suggestions
    return False, f"Family '{family}' not found in pft_data"


# ============================================================
# Content Generation
# ============================================================

def generate_title(product_name, brand, key_features=None, max_len=160):
    """
    Generate NIS-compliant product title (<=160 chars).
    Format: Brand + Core Product + Key Attributes + Use Case
    Title Case. No all-caps, no special symbols, no promo words.
    """
    promo_words = {'best', 'cheap', 'free', 'sale', 'discount', 'offer',
                   'deal', 'limited', 'exclusive', '#1', 'number one',
                   'top selling', 'bestseller', 'guaranteed'}

    # Clean product name: remove promo words
    name_clean = product_name
    for pw in promo_words:
        name_clean = re.sub(re.escape(pw), '', name_clean, flags=re.IGNORECASE)
    name_clean = re.sub(r'\s+', ' ', name_clean).strip()

    # Build title: Brand + Product Name
    title = f"{brand} {name_clean}"

    # Append key features if there's room
    if key_features:
        feats = key_features if isinstance(key_features, list) else key_features.split(';')
        feats = [f.strip() for f in feats if f.strip()]
        for feat in feats[:3]:
            candidate = f"{title} {feat}"
            if len(candidate) <= max_len:
                title = candidate
            else:
                break

    # Enforce Title Case, remove consecutive spaces
    title = title.title()
    title = re.sub(r'\s+', ' ', title).strip()

    # Remove special symbols except standard punctuation
    title = re.sub(r'[^\w\s\-/.,()&+]', '', title)

    # Truncate
    if len(title) > max_len:
        title = title[:max_len].rsplit(' ', 1)[0]

    return title


def generate_bullets(key_features, material=None, dimensions=None, weight=None,
                     color=None, target_market=None):
    """
    Generate 5 feature bullets (50-150 chars each).
    Start with function keyword, include specs.
    Last bullet may include GCC localization.
    """
    feats = key_features if isinstance(key_features, list) else key_features.split(';')
    feats = [f.strip() for f in feats if f.strip()]

    bullets = []

    # Generate bullets from key features
    for i, feat in enumerate(feats[:5]):
        bullet = feat
        # Ensure it starts with a capitalized word
        bullet = bullet[0].upper() + bullet[1:] if bullet else bullet

        # Pad short bullets with specs
        if len(bullet) < 50:
            extras = []
            if material and i == 0:
                extras.append(f"made from {material}")
            if dimensions and i == 1:
                extras.append(f"compact size {dimensions}")
            if color and i == 2:
                extras.append(f"available in {color}")
            if extras:
                bullet = f"{bullet} - {', '.join(extras)}"

        # Ensure bullet is within range
        if len(bullet) > 150:
            bullet = bullet[:150].rsplit(' ', 1)[0]

        bullets.append(bullet)

    # Fill remaining bullets
    fill_parts = []
    if material:
        fill_parts.append(f"Premium quality {material} construction for long-lasting durability and reliable performance")
    if dimensions and weight:
        fill_parts.append(f"Compact dimensions of {dimensions} cm and lightweight at {weight} kg for easy handling and installation")
    if color:
        fill_parts.append(f"Sleek {color} finish that complements any vehicle interior or exterior style")

    while len(bullets) < 4 and fill_parts:
        bullets.append(fill_parts.pop(0))

    # Last bullet: GCC localization
    if len(bullets) < 5:
        gcc_bullet = "Engineered to withstand extreme heat and harsh desert environments across the GCC region"
        if target_market == 'KSA':
            gcc_bullet = "Designed and tested for Saudi Arabian road conditions and extreme desert temperatures"
        bullets.append(gcc_bullet)

    # Fill any remaining with generic
    while len(bullets) < 5:
        bullets.append("Professional grade quality backed by manufacturer warranty for peace of mind")

    # Final trim
    for i in range(len(bullets)):
        if len(bullets[i]) > 150:
            bullets[i] = bullets[i][:150].rsplit(' ', 1)[0]
        if len(bullets[i]) < 50:
            bullets[i] = bullets[i] + " - built for quality, performance, and lasting reliability"
            if len(bullets[i]) > 150:
                bullets[i] = bullets[i][:150].rsplit(' ', 1)[0]

    return bullets[:5]


def generate_long_description(product_name, brand, key_features, material=None,
                              dimensions=None, weight=None, color=None,
                              package_contents=None):
    """
    Generate long description (200-500 words).
    Uses allowed HTML tags: <b> <br> <ul> <li> <p>
    Structure: Opening + Features + Specs + Package contents
    """
    feats = key_features if isinstance(key_features, list) else key_features.split(';')
    feats = [f.strip() for f in feats if f.strip()]

    parts = []

    # Opening paragraph (~50 words)
    top_feats = ', '.join(feats[:3]) if feats else 'advanced features'
    parts.append(
        f"<p><b>Introducing the {brand} {product_name}</b> - a premium automotive "
        f"product featuring {top_feats}. Designed for drivers who demand quality and "
        f"reliability, this product delivers exceptional performance in all conditions "
        f"including the demanding GCC climate.</p>"
    )

    # Features paragraph
    if feats:
        parts.append("<p><b>Key Features:</b></p>")
        parts.append("<ul>")
        for feat in feats:
            parts.append(f"<li>{feat}</li>")
        parts.append("</ul>")

    # Specifications
    specs = []
    if material:
        specs.append(f"<li><b>Material:</b> {material}</li>")
    if dimensions:
        specs.append(f"<li><b>Dimensions:</b> {dimensions} cm</li>")
    if weight:
        specs.append(f"<li><b>Weight:</b> {weight} kg</li>")
    if color:
        specs.append(f"<li><b>Color:</b> {color}</li>")

    if specs:
        parts.append("<p><b>Specifications:</b></p>")
        parts.append("<ul>")
        parts.extend(specs)
        parts.append("</ul>")

    # Package contents
    if package_contents:
        parts.append(f"<p><b>Package Contents:</b><br>{package_contents}</p>")

    # Quality assurance closing
    parts.append(
        "<p>Every unit undergoes strict quality control to ensure it meets international "
        "standards. Built to perform in the extreme temperatures and dusty conditions "
        "common across the Middle East region.</p>"
    )

    return '\n'.join(parts)


# ============================================================
# Automotive subcategory mapping (for --from-selection)
# ============================================================

def map_subcategory_to_family(subcategory):
    """Map a selection report subcategory to (family, product_type)."""
    if not subcategory:
        return AUTOMOTIVE_DEFAULT

    key = subcategory.strip().lower()

    # Direct match
    if key in AUTOMOTIVE_MAPPING:
        return AUTOMOTIVE_MAPPING[key]

    # Partial match
    for pattern, mapping in AUTOMOTIVE_MAPPING.items():
        if pattern in key or key in pattern:
            return mapping

    # Keyword fallback
    key_lower = key.lower()
    if any(w in key_lower for w in ['engine', 'motor', 'piston', 'cylinder']):
        return ('fmautomotive', 'Engine Parts')
    if any(w in key_lower for w in ['brake', 'caliper', 'rotor']):
        return ('fmautomotive', 'Brake Parts')
    if any(w in key_lower for w in ['audio', 'video', 'camera', 'dash', 'gps', 'stereo']):
        return ('fmautomotive', 'Car Audio/Video')
    if any(w in key_lower for w in ['interior', 'seat', 'mat', 'cover', 'cushion']):
        return ('fmautomotive', 'Interior Accessories')
    if any(w in key_lower for w in ['exterior', 'bumper', 'mirror', 'rack']):
        return ('fmautomotive', 'Exterior Accessories')
    if any(w in key_lower for w in ['suspension', 'spring', 'shock', 'strut', 'hub', 'bearing', 'wheel bearing']):
        return ('fmautomotive', 'Suspension Parts')
    if any(w in key_lower for w in ['light', 'lamp', 'led', 'bulb', 'electrical', 'wiring']):
        return ('fmautomotive', 'Electrical & Lighting')
    if any(w in key_lower for w in ['transmission', 'clutch', 'gear']):
        return ('fmautomotive', 'Transmission Parts')
    if any(w in key_lower for w in ['battery', 'oil', 'lubricant', 'fluid']):
        return ('fmautomotive', 'Batteries & Lubricants')
    if any(w in key_lower for w in ['body', 'fender', 'hood', 'door']):
        return ('fmautomotive', 'Body Parts')
    if any(w in key_lower for w in ['clean', 'wash', 'polish', 'wax']):
        return ('fmautomotive', 'Cleaning Accessories')

    return AUTOMOTIVE_DEFAULT


# ============================================================
# Input Readers
# ============================================================

def read_csv_input(csv_path):
    """Read products from CSV file."""
    products = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            key_features = row.get('key_features', '')
            image_urls = [u.strip() for u in row.get('image_urls', '').split(';') if u.strip()]

            product = {
                'product_name': row.get('product_name', ''),
                'family': row.get('family', ''),
                'product_type': row.get('product_type', ''),
                'brand': row.get('brand', ''),
                'sku': row.get('sku', ''),
                'key_features': key_features,
                'material': row.get('material', ''),
                'dimensions': row.get('dimensions_cm', ''),
                'weight': row.get('weight_kg', ''),
                'color': row.get('color', ''),
                'package_contents': row.get('package_contents', ''),
                'image_urls': image_urls,
                'target_market': row.get('target_market', 'UAE'),
            }
            products.append(product)
    return products


def read_selection_report(report_path):
    """
    Read products from a selection report .xlsx (Selection Dashboard sheet).
    Maps subcategory to (family, product_type).
    """
    wb = openpyxl.load_workbook(report_path, read_only=True)
    ws = wb['Selection Dashboard']

    # Read headers
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]

    # Build column index
    col_map = {}
    for i, h in enumerate(headers):
        col_map[h.lower()] = i

    products = []
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue

        def get(name, default=''):
            idx = col_map.get(name.lower())
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return default

        title = get('Title')
        subcategory = get('Subcategory')
        product_id = get('Product ID')
        price = get('Price (SAR)')
        url = get('URL')

        if not title:
            continue

        family, product_type = map_subcategory_to_family(subcategory)

        # Extract brand from title (first word or first two words)
        title_parts = title.split()
        brand = title_parts[0] if title_parts else 'Generic'
        # If brand looks like a short abbreviation or common brand pattern, keep it
        # Otherwise take first word
        if len(brand) <= 2 and len(title_parts) > 1:
            brand = f"{title_parts[0]} {title_parts[1]}"

        # Generate a SKU from product_id
        sku = product_id if product_id else f"AUTO-{len(products)+1:04d}"

        # Extract key features from title
        # Remove brand from title for feature extraction
        product_desc = title
        for bp in brand.split():
            product_desc = product_desc.replace(bp, '', 1).strip()

        key_features = [f.strip() for f in product_desc.split(',') if f.strip()]
        if not key_features:
            key_features = [product_desc]

        product = {
            'product_name': title,
            'family': family,
            'product_type': product_type,
            'brand': brand,
            'sku': sku,
            'key_features': ';'.join(key_features[:6]),
            'material': '',
            'dimensions': '',
            'weight': '',
            'color': '',
            'package_contents': '',
            'image_urls': [],
            'target_market': 'KSA',  # selection reports are typically KSA
            'source_url': url,
            'source_subcategory': subcategory,
            'source_price': price,
        }
        products.append(product)

    return products


def interactive_input():
    """Interactively collect product data from the user."""
    print("\n=== NIS Product Entry (Interactive Mode) ===\n")

    product = {}
    product['product_name'] = input("Product Name: ").strip()
    product['family'] = input("Family (e.g., fmautomotive): ").strip()
    product['product_type'] = input("Product Type (e.g., Car Audio/Video): ").strip()
    product['brand'] = input("Brand: ").strip() or 'Null'
    product['sku'] = input("Partner SKU: ").strip()
    product['key_features'] = input("Key Features (semicolon-separated): ").strip()
    product['material'] = input("Material (optional): ").strip()
    product['dimensions'] = input("Dimensions L x W x H cm (e.g., 10x6x4, optional): ").strip()
    product['weight'] = input("Weight kg (optional): ").strip()
    product['color'] = input("Color (optional): ").strip()
    product['package_contents'] = input("Package Contents (optional): ").strip()

    img_input = input("Image URLs (semicolon-separated, optional): ").strip()
    product['image_urls'] = [u.strip() for u in img_input.split(';') if u.strip()] if img_input else []

    product['target_market'] = input("Target Market (UAE/KSA, default=UAE): ").strip().upper() or 'UAE'

    return [product]


# ============================================================
# Product Processing
# ============================================================

def process_product(raw_product):
    """
    Transform a raw product dict into NIS-ready fields with generated content.
    Returns a dict ready for write_to_nis.
    """
    name = raw_product.get('product_name', '')
    brand = raw_product.get('brand', 'Null')
    key_features = raw_product.get('key_features', '')
    material = raw_product.get('material', '')
    dimensions = raw_product.get('dimensions', '')
    weight = raw_product.get('weight', '')
    color = raw_product.get('color', '')
    package_contents = raw_product.get('package_contents', '')
    target_market = raw_product.get('target_market', 'UAE')

    # Generate title
    title_en = generate_title(name, brand, key_features)

    # Generate bullets
    bullets_en = generate_bullets(
        key_features, material=material, dimensions=dimensions,
        weight=weight, color=color, target_market=target_market
    )

    # Generate long description
    long_desc_en = generate_long_description(
        name, brand, key_features, material=material,
        dimensions=dimensions, weight=weight, color=color,
        package_contents=package_contents
    )

    # What's in the box
    whats_in_box = package_contents if package_contents else name

    return {
        'family': raw_product.get('family', ''),
        'product_type': raw_product.get('product_type', ''),
        'product_subtype': raw_product.get('product_subtype', ''),
        'sku': raw_product.get('sku', ''),
        'brand': brand,
        'parent_group_key': raw_product.get('parent_group_key', ''),
        'title_en': title_en,
        'long_description_en': long_desc_en,
        'bullets_en': bullets_en,
        'whats_in_box_en': whats_in_box,
        'image_urls': raw_product.get('image_urls', []),
    }


# ============================================================
# NIS File Generation
# ============================================================

def generate_nis_file(products, output_path, nis_template_path):
    """
    Write processed products into a copy of the NIS template.
    Preserves first 9 rows (metadata/headers), valid values sheet, and pft_data sheet.
    """
    wb = openpyxl.load_workbook(nis_template_path)
    ws = wb['template_data']

    for i, product in enumerate(products):
        r = DATA_START_ROW + i  # openpyxl 1-indexed, Row 10 onwards

        ws.cell(row=r, column=1, value=product['family'])
        ws.cell(row=r, column=2, value=product['product_type'])
        ws.cell(row=r, column=3, value=product.get('product_subtype', ''))
        ws.cell(row=r, column=4, value=product['sku'])
        ws.cell(row=r, column=5, value=product.get('brand', 'Null'))
        ws.cell(row=r, column=6, value=product.get('parent_group_key', ''))
        ws.cell(row=r, column=7, value='Parent')
        ws.cell(row=r, column=8, value='')  # size_variation
        ws.cell(row=r, column=9, value=product['long_description_en'])
        ws.cell(row=r, column=10, value='')  # AR留空
        ws.cell(row=r, column=11, value=product['title_en'])
        ws.cell(row=r, column=12, value='')  # AR留空
        ws.cell(row=r, column=13, value='')  # item_condition
        for img_idx, url in enumerate(product.get('image_urls', [])[:7]):
            ws.cell(row=r, column=14 + img_idx, value=url)
        for b_idx, bullet in enumerate(product.get('bullets_en', [])[:5]):
            ws.cell(row=r, column=21 + b_idx, value=bullet)
        # AR bullets留空 (col 26-30)
        ws.cell(row=r, column=31, value=product.get('whats_in_box_en', ''))
        ws.cell(row=r, column=32, value='')  # AR
        ws.cell(row=r, column=33, value='Std')
        ws.cell(row=r, column=34, value='Std')
        ws.cell(row=r, column=35, value='Std')

    wb.save(output_path)
    return len(products)


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='NIS upload file generator for Noon Seller Lab',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python noon_nis_generator.py --input products.csv --nis-template NIS.xlsx --output upload.xlsx
  python noon_nis_generator.py --from-selection selection_report.xlsx --nis-template NIS.xlsx --output batch.xlsx
  python noon_nis_generator.py --interactive --nis-template NIS.xlsx
        """
    )
    parser.add_argument('--input', '-i', help='Input CSV file with product data')
    parser.add_argument('--from-selection', help='Selection report .xlsx (automotive)')
    parser.add_argument('--interactive', action='store_true', help='Interactive single product entry')
    parser.add_argument('--nis-template', required=True, help='Path to NIS.xlsx template')
    parser.add_argument('--output', '-o', help='Output .xlsx path (default: nis_output.xlsx)',
                        default='nis_output.xlsx')
    parser.add_argument('--skip-validation', action='store_true',
                        help='Skip family/product_type validation against pft_data')
    parser.add_argument('--limit', type=int, default=0,
                        help='Limit number of products to process (0 = all)')

    args = parser.parse_args()

    # Validate arguments: must specify exactly one input mode
    input_modes = sum([
        args.input is not None,
        args.from_selection is not None,
        args.interactive,
    ])
    if input_modes != 1:
        parser.error("Specify exactly one input mode: --input, --from-selection, or --interactive")

    # Verify template exists
    if not os.path.exists(args.nis_template):
        print(f"ERROR: NIS template not found: {args.nis_template}")
        sys.exit(1)

    # ── Read raw products ──
    try:
        if args.input:
            if not os.path.exists(args.input):
                print(f"ERROR: Input CSV not found: {args.input}")
                sys.exit(1)
            print(f"[1/4] Reading CSV: {args.input}")
            raw_products = read_csv_input(args.input)
        elif args.from_selection:
            if not os.path.exists(args.from_selection):
                print(f"ERROR: Selection report not found: {args.from_selection}")
                sys.exit(1)
            print(f"[1/4] Reading selection report: {args.from_selection}")
            raw_products = read_selection_report(args.from_selection)
        else:
            print("[1/4] Interactive mode")
            raw_products = interactive_input()
    except Exception as e:
        print(f"ERROR reading input: {e}")
        sys.exit(1)

    if not raw_products:
        print("ERROR: No products found in input.")
        sys.exit(1)

    # Apply limit
    if args.limit > 0:
        raw_products = raw_products[:args.limit]

    print(f"       Found {len(raw_products)} product(s)")

    # ── Validate family/product_type ──
    print(f"[2/4] Validating family/product_type against pft_data...")
    pft = load_pft_data(args.nis_template)
    valid_products = []
    skipped = 0

    for idx, prod in enumerate(raw_products):
        family = prod.get('family', '')
        ptype = prod.get('product_type', '')

        if args.skip_validation:
            valid_products.append(prod)
            continue

        if not family or not ptype:
            print(f"  WARNING: Product #{idx+1} '{prod.get('sku', '?')}' missing family or product_type, skipping.")
            skipped += 1
            continue

        is_valid, suggestion = validate_nis_entry(family, ptype, args.nis_template)
        if is_valid:
            valid_products.append(prod)
        else:
            if isinstance(suggestion, list):
                suggest_str = ', '.join(suggestion) if suggestion else 'none'
                print(f"  WARNING: Product #{idx+1} '{prod.get('sku', '?')}': "
                      f"'{ptype}' not valid for family '{family}'. "
                      f"Suggestions: {suggest_str}. Skipping.")
            else:
                print(f"  WARNING: Product #{idx+1} '{prod.get('sku', '?')}': {suggestion}. Skipping.")
            skipped += 1

    if skipped:
        print(f"       Validated: {len(valid_products)} passed, {skipped} skipped")
    else:
        print(f"       All {len(valid_products)} product(s) validated OK")

    if not valid_products:
        print("ERROR: No valid products to write.")
        sys.exit(1)

    # ── Process products (generate content) ──
    print(f"[3/4] Generating titles, descriptions, and bullets...")
    processed = []
    for idx, prod in enumerate(valid_products):
        try:
            p = process_product(prod)
            processed.append(p)
            if (idx + 1) % 10 == 0 or idx == 0:
                print(f"       Processed {idx+1}/{len(valid_products)}: {p['sku']}")
        except Exception as e:
            print(f"  WARNING: Failed to process product #{idx+1} '{prod.get('sku', '?')}': {e}")

    if not processed:
        print("ERROR: No products processed successfully.")
        sys.exit(1)

    # ── Write NIS file ──
    print(f"[4/4] Writing NIS file: {args.output}")
    try:
        count = generate_nis_file(processed, args.output, args.nis_template)
        print(f"\n  SUCCESS: {count} product(s) written to {args.output}")
        print(f"  Template metadata rows (1-9) preserved.")
        print(f"  'valid values' and 'pft_data' sheets preserved.")
    except Exception as e:
        print(f"ERROR writing output: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
