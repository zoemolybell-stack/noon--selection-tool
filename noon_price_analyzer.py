#!/usr/bin/env python3
"""
noon_price_analyzer.py — Noon 价格追踪分析工具

用法:
  python noon_price_analyzer.py --baseline report_automotive.xlsx --tracking tracking_dir/ --output price_analysis.xlsx

分析维度:
  1. Price Movement: 价格变动趋势
  2. Sales Velocity: 销售速度排名
  3. Competition Changes: Express竞争对手变化 & 评论增长
  4. Anomaly Alerts: 异常预警(价格暴跌/暴涨, 销量激增, 新Express竞争者)
"""

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, 'D:/claude noon')
from noon_config import CONFIG, calc_full_profit


# ============================================================
# Styling constants
# ============================================================
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
GREEN_FONT = Font(color="006100")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FONT = Font(color="9C0006")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FONT = Font(color="9C5700")


# ============================================================
# Helper functions
# ============================================================

def safe_float(val, default=0.0):
    """Safely convert a value to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Safely convert a value to int."""
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_bool(val, default=False):
    """Safely convert a value to bool."""
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.lower() in ('true', '1', 'yes')
    return bool(val)


def style_header_row(ws, num_cols):
    """Apply dark blue header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_column_width(ws, min_width=10, max_width=40, sample_rows=100):
    """Auto-fit column widths by sampling first N rows (avoids full-sheet scan)."""
    # Determine number of columns from row 1
    max_col = ws.max_column or 1
    col_widths = [min_width] * (max_col + 1)
    for row in ws.iter_rows(min_row=1, max_row=min(sample_rows, ws.max_row or 1),
                            min_col=1, max_col=max_col):
        for cell in row:
            try:
                val = str(cell.value) if cell.value else ""
                col_widths[cell.column] = max(col_widths[cell.column], len(val) + 2)
            except Exception:
                pass
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(col_widths[col], max_width)


def apply_change_color(cell, value):
    """Apply green/red color based on positive/negative value."""
    if value is None:
        return
    if value > 0:
        cell.font = GREEN_FONT
        cell.fill = GREEN_FILL
    elif value < 0:
        cell.font = RED_FONT
        cell.fill = RED_FILL


# ============================================================
# Data loading
# ============================================================

def load_baseline(filepath):
    """Load baseline data from the 'All Products' sheet of the report xlsx."""
    print(f"[1/4] Loading baseline: {filepath}")
    if not os.path.exists(filepath):
        print(f"  ERROR: Baseline file not found: {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if 'All Products' not in wb.sheetnames:
        print(f"  ERROR: 'All Products' sheet not found. Available: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    ws = wb['All Products']

    # Stream rows instead of loading all into memory at once
    headers = None
    products = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(row)]
            continue
        try:
            record = dict(zip(headers, row))
            products.append(record)
        except Exception as e:
            print(f"  WARNING: Skipping row {row_idx}: {e}")
        if row_idx % 20000 == 0:
            print(f"  ... loaded {row_idx - 1} rows")

    if not headers:
        print("  ERROR: Empty sheet.")
        wb.close()
        sys.exit(1)

    wb.close()
    print(f"  Loaded {len(products)} products from baseline.")
    return products


def load_tracking_data(tracking_dir):
    """
    Load tracking data from directory containing JSON or CSV files.
    Returns dict: product_id -> list of snapshots [{date, price, sold_recently, review_count, is_express, ...}]
    Graceful: returns empty dict if dir doesn't exist or has no data.
    """
    print(f"[2/4] Loading tracking data: {tracking_dir}")
    tracking = defaultdict(list)

    if not tracking_dir or not os.path.exists(tracking_dir):
        print("  No tracking directory found. Will produce baseline-only analysis.")
        return {}

    files_loaded = 0
    for fname in sorted(os.listdir(tracking_dir)):
        fpath = os.path.join(tracking_dir, fname)
        if not os.path.isfile(fpath):
            continue

        try:
            if fname.endswith('.json'):
                with open(fpath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # Support both list-of-records and {product_id: record} formats
                records = data if isinstance(data, list) else list(data.values()) if isinstance(data, dict) else []
                date_str = fname.replace('.json', '').split('_')[-1] if '_' in fname else fname.replace('.json', '')
                for rec in records:
                    if isinstance(rec, dict) and 'product_id' in rec:
                        rec.setdefault('date', date_str)
                        tracking[rec['product_id']].append(rec)
                files_loaded += 1

            elif fname.endswith('.csv'):
                with open(fpath, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    date_str = fname.replace('.csv', '').split('_')[-1] if '_' in fname else fname.replace('.csv', '')
                    for rec in reader:
                        if 'product_id' in rec:
                            rec.setdefault('date', date_str)
                            tracking[rec['product_id']].append(rec)
                files_loaded += 1

        except Exception as e:
            print(f"  WARNING: Failed to load {fname}: {e}")

    print(f"  Loaded {files_loaded} tracking files, {len(tracking)} tracked products.")
    return dict(tracking)


# ============================================================
# Analysis engine
# ============================================================

def analyze(products, tracking):
    """
    Run all analysis dimensions. Returns dict with results for each sheet.
    """
    print("[3/4] Analyzing data...")
    has_tracking = bool(tracking)

    # Build product lookup by ID
    by_id = {}
    for p in products:
        pid = p.get('product_id', '')
        if pid:
            by_id[pid] = p

    # ── Sheet 1: Price Trends ──
    price_trends = []
    for p in products:
        try:
            pid = p.get('product_id', '')
            title = str(p.get('title', ''))[:80]
            subcat = str(p.get('_subcategory', ''))
            baseline_price = safe_float(p.get('price'))

            latest_price = baseline_price
            change_pct = 0.0
            trend_label = "Stable"

            if has_tracking and pid in tracking:
                snapshots = sorted(tracking[pid], key=lambda s: str(s.get('date', '')))
                if snapshots:
                    latest_price = safe_float(snapshots[-1].get('price', baseline_price), baseline_price)
                    if baseline_price > 0:
                        change_pct = (latest_price - baseline_price) / baseline_price * 100
                    if change_pct > 5:
                        trend_label = "Up"
                    elif change_pct < -5:
                        trend_label = "Down"

            price_trends.append({
                'product_id': pid,
                'title': title,
                'subcategory': subcat,
                'baseline_price': round(baseline_price, 2),
                'latest_price': round(latest_price, 2),
                'change_pct': round(change_pct, 2),
                'trend_label': trend_label,
            })
        except Exception as e:
            print(f"  WARNING: Price trend error for {p.get('product_id', '?')}: {e}")

    # Sort by absolute change descending (volatility ranking)
    price_trends.sort(key=lambda x: abs(x['change_pct']), reverse=True)

    # ── Sheet 2: Sales Velocity ──
    sales_velocity = []
    for p in products:
        try:
            pid = p.get('product_id', '')
            title = str(p.get('title', ''))[:80]
            subcat = str(p.get('_subcategory', ''))
            price = safe_float(p.get('price'))
            sold = p.get('sold_recently')
            review_count = safe_int(p.get('review_count'))
            is_express = safe_bool(p.get('is_express'))
            is_bestseller = safe_bool(p.get('is_bestseller'))
            margin_war = safe_float(p.get('margin_war_pct'))

            # Interpret sold_recently: could be int, string like "100+", or None
            sold_val = 0
            if sold is not None:
                sold_str = str(sold).replace('+', '').replace(',', '').strip()
                try:
                    sold_val = int(sold_str)
                except ValueError:
                    sold_val = 0

            # Demand score: composite of sold, reviews, bestseller, express
            demand_score = sold_val * 10 + review_count + (50 if is_bestseller else 0) + (20 if is_express else 0)

            # Tracking-based velocity change
            velocity_change = "N/A"
            if has_tracking and pid in tracking:
                snapshots = sorted(tracking[pid], key=lambda s: str(s.get('date', '')))
                if len(snapshots) >= 2:
                    old_sold = safe_int(snapshots[0].get('sold_recently'))
                    new_sold = safe_int(snapshots[-1].get('sold_recently'))
                    if old_sold > 0:
                        vel_pct = (new_sold - old_sold) / old_sold * 100
                        velocity_change = f"{vel_pct:+.0f}%"
                    elif new_sold > 0:
                        velocity_change = "New Sales"

            sales_velocity.append({
                'product_id': pid,
                'title': title,
                'subcategory': subcat,
                'price': round(price, 2),
                'sold_recently': sold_val,
                'review_count': review_count,
                'is_express': is_express,
                'is_bestseller': is_bestseller,
                'demand_score': demand_score,
                'margin_war_pct': round(margin_war, 2),
                'velocity_change': velocity_change,
            })
        except Exception as e:
            print(f"  WARNING: Sales velocity error for {p.get('product_id', '?')}: {e}")

    # Rank by demand score descending
    sales_velocity.sort(key=lambda x: x['demand_score'], reverse=True)

    # ── Sheet 3: Competition Changes ──
    competition = []
    for p in products:
        try:
            pid = p.get('product_id', '')
            title = str(p.get('title', ''))[:80]
            subcat = str(p.get('_subcategory', ''))
            baseline_express = safe_bool(p.get('is_express'))
            baseline_reviews = safe_int(p.get('review_count'))
            price = safe_float(p.get('price'))

            latest_express = baseline_express
            latest_reviews = baseline_reviews
            review_growth = 0
            express_change = "No Change"

            if has_tracking and pid in tracking:
                snapshots = sorted(tracking[pid], key=lambda s: str(s.get('date', '')))
                if snapshots:
                    latest_express = safe_bool(snapshots[-1].get('is_express', baseline_express))
                    latest_reviews = safe_int(snapshots[-1].get('review_count', baseline_reviews))
                    review_growth = latest_reviews - baseline_reviews
                    if not baseline_express and latest_express:
                        express_change = "NEW Express"
                    elif baseline_express and not latest_express:
                        express_change = "Lost Express"

            competition.append({
                'product_id': pid,
                'title': title,
                'subcategory': subcat,
                'price': round(price, 2),
                'baseline_express': baseline_express,
                'latest_express': latest_express,
                'express_change': express_change,
                'baseline_reviews': baseline_reviews,
                'latest_reviews': latest_reviews,
                'review_growth': review_growth,
            })
        except Exception as e:
            print(f"  WARNING: Competition error for {p.get('product_id', '?')}: {e}")

    # Sort: new express first, then by review growth descending
    competition.sort(key=lambda x: (0 if x['express_change'] == 'NEW Express' else 1, -x['review_growth']))

    # ── Sheet 4: Alerts ──
    alerts = []
    for p in products:
        try:
            pid = p.get('product_id', '')
            title = str(p.get('title', ''))[:80]
            subcat = str(p.get('_subcategory', ''))
            baseline_price = safe_float(p.get('price'))
            baseline_express = safe_bool(p.get('is_express'))
            baseline_sold = 0
            sold_raw = p.get('sold_recently')
            if sold_raw is not None:
                try:
                    baseline_sold = int(str(sold_raw).replace('+', '').replace(',', '').strip())
                except ValueError:
                    pass

            if not has_tracking or pid not in tracking:
                continue

            snapshots = sorted(tracking[pid], key=lambda s: str(s.get('date', '')))
            if not snapshots:
                continue

            latest = snapshots[-1]
            latest_price = safe_float(latest.get('price', baseline_price), baseline_price)
            latest_express = safe_bool(latest.get('is_express', baseline_express))
            latest_sold = 0
            latest_sold_raw = latest.get('sold_recently')
            if latest_sold_raw is not None:
                try:
                    latest_sold = int(str(latest_sold_raw).replace('+', '').replace(',', '').strip())
                except ValueError:
                    pass

            # Price crash > 20%
            if baseline_price > 0:
                price_change_pct = (latest_price - baseline_price) / baseline_price * 100
                if price_change_pct < -20:
                    alerts.append({
                        'product_id': pid, 'title': title, 'subcategory': subcat,
                        'alert_type': 'Price Crash',
                        'severity': 'HIGH',
                        'detail': f"Price dropped {price_change_pct:.1f}% ({baseline_price:.0f} -> {latest_price:.0f})",
                        'baseline_value': f"{baseline_price:.0f}",
                        'latest_value': f"{latest_price:.0f}",
                    })

                # Price spike > 30%
                if price_change_pct > 30:
                    alerts.append({
                        'product_id': pid, 'title': title, 'subcategory': subcat,
                        'alert_type': 'Price Spike',
                        'severity': 'MEDIUM',
                        'detail': f"Price surged {price_change_pct:.1f}% ({baseline_price:.0f} -> {latest_price:.0f})",
                        'baseline_value': f"{baseline_price:.0f}",
                        'latest_value': f"{latest_price:.0f}",
                    })

            # Sales surge > 100%
            if baseline_sold > 0 and latest_sold > 0:
                sales_change_pct = (latest_sold - baseline_sold) / baseline_sold * 100
                if sales_change_pct > 100:
                    alerts.append({
                        'product_id': pid, 'title': title, 'subcategory': subcat,
                        'alert_type': 'Sales Surge',
                        'severity': 'MEDIUM',
                        'detail': f"Sales surged {sales_change_pct:.0f}% ({baseline_sold} -> {latest_sold})",
                        'baseline_value': str(baseline_sold),
                        'latest_value': str(latest_sold),
                    })

            # New Express competitor
            if not baseline_express and latest_express:
                alerts.append({
                    'product_id': pid, 'title': title, 'subcategory': subcat,
                    'alert_type': 'New Express',
                    'severity': 'HIGH',
                    'detail': "Product gained Express badge",
                    'baseline_value': 'No',
                    'latest_value': 'Yes',
                })

        except Exception as e:
            print(f"  WARNING: Alert check error for {p.get('product_id', '?')}: {e}")

    # Sort alerts by severity
    severity_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
    alerts.sort(key=lambda x: severity_order.get(x.get('severity', 'LOW'), 9))

    # ── Subcategory summary for baseline-only mode ──
    subcat_summary = defaultdict(lambda: {'count': 0, 'total_price': 0, 'min_price': float('inf'),
                                           'max_price': 0, 'express_count': 0, 'total_reviews': 0,
                                           'total_sold': 0})
    for p in products:
        try:
            subcat = str(p.get('_subcategory', 'Unknown'))
            price = safe_float(p.get('price'))
            s = subcat_summary[subcat]
            s['count'] += 1
            s['total_price'] += price
            s['min_price'] = min(s['min_price'], price) if price > 0 else s['min_price']
            s['max_price'] = max(s['max_price'], price)
            if safe_bool(p.get('is_express')):
                s['express_count'] += 1
            s['total_reviews'] += safe_int(p.get('review_count'))
            sold_raw = p.get('sold_recently')
            if sold_raw is not None:
                try:
                    s['total_sold'] += int(str(sold_raw).replace('+', '').replace(',', '').strip())
                except ValueError:
                    pass
        except Exception:
            pass

    print(f"  Analysis complete: {len(price_trends)} price records, {len(alerts)} alerts.")

    return {
        'price_trends': price_trends,
        'sales_velocity': sales_velocity,
        'competition': competition,
        'alerts': alerts,
        'has_tracking': has_tracking,
        'subcat_summary': dict(subcat_summary),
        'total_products': len(products),
    }


# ============================================================
# Output generation
# ============================================================


def write_excel(results, output_path):
    """Write analysis results to formatted Excel workbook."""
    print(f"[4/4] Writing output: {output_path}")
    total = results['total_products']
    large_mode = total > 5000
    if large_mode:
        print(f"  Large dataset ({total} rows) — using fast write mode (no per-cell borders)")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Price Trends ──
    ws1 = wb.active
    ws1.title = "Price Trends"
    headers1 = ["product_id", "title", "subcategory", "baseline_price", "latest_price", "change_pct", "trend_label"]
    ws1.append(headers1)
    style_header_row(ws1, len(headers1))

    print("  Writing Price Trends...")
    r = 1
    for i, row_data in enumerate(results['price_trends']):
        r = i + 2
        if large_mode and i % 20000 == 0 and i > 0:
            print(f"    ... {i} rows written")
        ws1.cell(row=r, column=1, value=row_data['product_id'])
        ws1.cell(row=r, column=2, value=row_data['title'])
        ws1.cell(row=r, column=3, value=row_data['subcategory'])
        ws1.cell(row=r, column=4, value=row_data['baseline_price']).number_format = '#,##0.00'
        ws1.cell(row=r, column=5, value=row_data['latest_price']).number_format = '#,##0.00'
        pct_cell = ws1.cell(row=r, column=6, value=row_data['change_pct'])
        pct_cell.number_format = '0.00"%"'
        apply_change_color(pct_cell, row_data['change_pct'])
        trend_cell = ws1.cell(row=r, column=7, value=row_data['trend_label'])
        if row_data['trend_label'] == 'Up':
            trend_cell.font = GREEN_FONT
        elif row_data['trend_label'] == 'Down':
            trend_cell.font = RED_FONT

        if not large_mode:
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=r, column=c).border = THIN_BORDER

    if not results['has_tracking']:
        r = r + 2
        note_cell = ws1.cell(row=r, column=1, value="Note: Awaiting tracking data for trend analysis. Showing baseline prices only.")
        note_cell.font = Font(italic=True, color="666666")

    ws1.freeze_panes = "A2"
    auto_column_width(ws1)

    # ── Sheet 2: Sales Velocity ──
    ws2 = wb.create_sheet("Sales Velocity")
    headers2 = ["product_id", "title", "subcategory", "price", "sold_recently", "review_count",
                 "is_express", "is_bestseller", "demand_score", "margin_war_pct", "velocity_change"]
    ws2.append(headers2)
    style_header_row(ws2, len(headers2))

    for i, row_data in enumerate(results['sales_velocity']):
        r = i + 2
        ws2.cell(row=r, column=1, value=row_data['product_id'])
        ws2.cell(row=r, column=2, value=row_data['title'])
        ws2.cell(row=r, column=3, value=row_data['subcategory'])
        ws2.cell(row=r, column=4, value=row_data['price']).number_format = '#,##0.00'
        ws2.cell(row=r, column=5, value=row_data['sold_recently'])
        ws2.cell(row=r, column=6, value=row_data['review_count'])
        express_cell = ws2.cell(row=r, column=7, value="Yes" if row_data['is_express'] else "No")
        if row_data['is_express']:
            express_cell.font = Font(bold=True, color="006100")
        ws2.cell(row=r, column=8, value="Yes" if row_data['is_bestseller'] else "No")
        ws2.cell(row=r, column=9, value=row_data['demand_score'])
        margin_cell = ws2.cell(row=r, column=10, value=row_data['margin_war_pct'])
        margin_cell.number_format = '0.00"%"'
        apply_change_color(margin_cell, row_data['margin_war_pct'])
        ws2.cell(row=r, column=11, value=row_data['velocity_change'])

        if not large_mode:
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r, column=c).border = THIN_BORDER

    ws2.freeze_panes = "A2"
    auto_column_width(ws2)

    # ── Sheet 3: Competition Changes ──
    ws3 = wb.create_sheet("Competition Changes")
    headers3 = ["product_id", "title", "subcategory", "price", "baseline_express", "latest_express",
                 "express_change", "baseline_reviews", "latest_reviews", "review_growth"]
    ws3.append(headers3)
    style_header_row(ws3, len(headers3))

    for i, row_data in enumerate(results['competition']):
        r = i + 2
        ws3.cell(row=r, column=1, value=row_data['product_id'])
        ws3.cell(row=r, column=2, value=row_data['title'])
        ws3.cell(row=r, column=3, value=row_data['subcategory'])
        ws3.cell(row=r, column=4, value=row_data['price']).number_format = '#,##0.00'
        ws3.cell(row=r, column=5, value="Yes" if row_data['baseline_express'] else "No")
        ws3.cell(row=r, column=6, value="Yes" if row_data['latest_express'] else "No")
        change_cell = ws3.cell(row=r, column=7, value=row_data['express_change'])
        if row_data['express_change'] == 'NEW Express':
            change_cell.font = Font(bold=True, color="9C0006")
            change_cell.fill = RED_FILL
        elif row_data['express_change'] == 'Lost Express':
            change_cell.font = GREEN_FONT
            change_cell.fill = GREEN_FILL
        ws3.cell(row=r, column=8, value=row_data['baseline_reviews'])
        ws3.cell(row=r, column=9, value=row_data['latest_reviews'])
        growth_cell = ws3.cell(row=r, column=10, value=row_data['review_growth'])
        apply_change_color(growth_cell, row_data['review_growth'])

        if not large_mode:
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=r, column=c).border = THIN_BORDER

    ws3.freeze_panes = "A2"
    auto_column_width(ws3)

    # ── Sheet 4: Alerts ──
    ws4 = wb.create_sheet("Alerts")
    headers4 = ["product_id", "title", "subcategory", "alert_type", "severity", "detail",
                 "baseline_value", "latest_value"]
    ws4.append(headers4)
    style_header_row(ws4, len(headers4))

    if results['alerts']:
        for i, row_data in enumerate(results['alerts']):
            r = i + 2
            ws4.cell(row=r, column=1, value=row_data['product_id'])
            ws4.cell(row=r, column=2, value=row_data['title'])
            ws4.cell(row=r, column=3, value=row_data['subcategory'])
            type_cell = ws4.cell(row=r, column=4, value=row_data['alert_type'])
            sev_cell = ws4.cell(row=r, column=5, value=row_data['severity'])
            if row_data['severity'] == 'HIGH':
                sev_cell.font = RED_FONT
                sev_cell.fill = RED_FILL
            elif row_data['severity'] == 'MEDIUM':
                sev_cell.font = ORANGE_FONT
                sev_cell.fill = YELLOW_FILL
            ws4.cell(row=r, column=6, value=row_data['detail'])
            ws4.cell(row=r, column=7, value=row_data['baseline_value'])
            ws4.cell(row=r, column=8, value=row_data['latest_value'])

            # Color-code alert types
            if row_data['alert_type'] == 'Price Crash':
                type_cell.font = RED_FONT
            elif row_data['alert_type'] == 'Price Spike':
                type_cell.font = ORANGE_FONT
            elif row_data['alert_type'] == 'New Express':
                type_cell.font = Font(bold=True, color="9C0006")

            if not large_mode:
                for c in range(1, len(headers4) + 1):
                    ws4.cell(row=r, column=c).border = THIN_BORDER
    else:
        r = 2
        if results['has_tracking']:
            ws4.cell(row=r, column=1, value="No anomalies detected.").font = Font(italic=True, color="006100")
        else:
            ws4.cell(row=r, column=1,
                     value="Awaiting tracking data for anomaly detection. Baseline snapshot captured.").font = Font(italic=True, color="666666")

    ws4.freeze_panes = "A2"
    auto_column_width(ws4)

    # Save
    wb.save(output_path)
    print(f"  Saved: {output_path}")


def write_alerts_md(results, output_path):
    """Write human-readable alert summary as Markdown."""
    md_path = output_path.replace('.xlsx', '').rsplit('.', 1)[0] if not output_path.endswith('.xlsx') else output_path
    md_path = os.path.join(os.path.dirname(output_path) or '.', 'price_alerts.md')

    lines = []
    lines.append("# Noon Price Tracker - Alert Summary")
    lines.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Total products analyzed: {results['total_products']}")
    lines.append(f"Tracking data available: {'Yes' if results['has_tracking'] else 'No'}")

    # Subcategory overview
    lines.append("\n## Subcategory Overview\n")
    lines.append("| Subcategory | Count | Avg Price | Price Range | Express | Reviews |")
    lines.append("|---|---|---|---|---|---|")
    for subcat, stats in sorted(results['subcat_summary'].items(), key=lambda x: x[1]['count'], reverse=True):
        avg_price = stats['total_price'] / stats['count'] if stats['count'] > 0 else 0
        min_p = stats['min_price'] if stats['min_price'] != float('inf') else 0
        lines.append(f"| {subcat} | {stats['count']} | {avg_price:.0f} | {min_p:.0f}-{stats['max_price']:.0f} | "
                     f"{stats['express_count']} | {stats['total_reviews']} |")

    # Alerts
    if results['alerts']:
        lines.append(f"\n## Alerts ({len(results['alerts'])} total)\n")

        # Group by type
        by_type = defaultdict(list)
        for a in results['alerts']:
            by_type[a['alert_type']].append(a)

        for alert_type, items in by_type.items():
            lines.append(f"\n### {alert_type} ({len(items)})\n")
            for a in items[:20]:  # Cap at 20 per type
                severity_icon = "[!]" if a['severity'] == 'HIGH' else "[~]"
                lines.append(f"- {severity_icon} **{a['product_id']}** — {a['detail']}")
                lines.append(f"  - {a['title'][:60]}")
            if len(items) > 20:
                lines.append(f"- ... and {len(items) - 20} more")
    else:
        if results['has_tracking']:
            lines.append("\n## Alerts\n")
            lines.append("No anomalies detected. All metrics within normal ranges.")
        else:
            lines.append("\n## Alerts\n")
            lines.append("Awaiting tracking data for trend analysis.")
            lines.append("Baseline snapshot has been captured. Run tracking collection and re-analyze to detect:")
            lines.append("- Price crashes (>20% drop)")
            lines.append("- Price spikes (>30% increase)")
            lines.append("- Sales surges (>100% increase)")
            lines.append("- New Express competitors")

    # Price distribution snapshot
    lines.append("\n## Price Distribution (Baseline)\n")
    prices = [safe_float(p.get('price')) for p in results.get('_products_ref', []) if safe_float(p.get('price')) > 0]
    if prices:
        prices.sort()
        n = len(prices)
        lines.append(f"- Products with price > 0: {n}")
        lines.append(f"- Min: {prices[0]:.0f}")
        lines.append(f"- P25: {prices[n // 4]:.0f}")
        lines.append(f"- Median: {prices[n // 2]:.0f}")
        lines.append(f"- P75: {prices[3 * n // 4]:.0f}")
        lines.append(f"- Max: {prices[-1]:.0f}")
        lines.append(f"- Average: {sum(prices) / n:.0f}")

    lines.append("\n---\n*Generated by noon_price_analyzer.py*\n")

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"  Saved: {md_path}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Noon Price Tracking Analyzer - 价格追踪分析工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python noon_price_analyzer.py --baseline report_automotive.xlsx
  python noon_price_analyzer.py --baseline report.xlsx --tracking tracking_dir/ --output analysis.xlsx
        """
    )
    parser.add_argument('--baseline', required=True, help='Baseline report xlsx (must have "All Products" sheet)')
    parser.add_argument('--tracking', default=None, help='Directory with tracking JSON/CSV files (optional)')
    parser.add_argument('--output', default='price_analysis.xlsx', help='Output xlsx path (default: price_analysis.xlsx)')
    args = parser.parse_args()

    print("=" * 60)
    print("  Noon Price Tracker Analyzer")
    print("=" * 60)

    # Load data
    products = load_baseline(args.baseline)
    tracking = load_tracking_data(args.tracking)

    # Analyze
    results = analyze(products, tracking)
    # Attach raw products for price distribution in MD report
    results['_products_ref'] = products

    # Output
    write_excel(results, args.output)
    write_alerts_md(results, args.output)

    # Summary
    print("\n" + "=" * 60)
    print("  Summary")
    print("=" * 60)
    print(f"  Products analyzed:  {results['total_products']}")
    print(f"  Tracking data:      {'Yes' if results['has_tracking'] else 'No (baseline-only mode)'}")
    print(f"  Alerts generated:   {len(results['alerts'])}")

    # Quick stats
    up = sum(1 for t in results['price_trends'] if t['trend_label'] == 'Up')
    down = sum(1 for t in results['price_trends'] if t['trend_label'] == 'Down')
    stable = sum(1 for t in results['price_trends'] if t['trend_label'] == 'Stable')
    print(f"  Price trends:       {up} up / {down} down / {stable} stable")

    express_count = sum(1 for s in results['sales_velocity'] if s['is_express'])
    print(f"  Express products:   {express_count}")
    print(f"\n  Output files:")
    print(f"    {args.output}")
    print(f"    price_alerts.md")
    print("=" * 60)


if __name__ == '__main__':
    main()
