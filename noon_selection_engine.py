#!/usr/bin/env python3
"""
Noon 选品决策引擎 — 深度整合爬虫数据与 V6 定价逻辑

功能:
  1. NoonPricingEngine: 自包含定价计算器 (FBN + FBP 双模式)
  2. analyze_products(): 读取爬虫报告, 对 Top 100 产品做全维度分析
  3. 输出 selection_report_automotive.xlsx (3个Sheet)

依赖: pandas, openpyxl
"""

import sys
import math
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── 导入现有定价引擎模块 ──
sys.path.insert(0, "D:/claude noon/noon-selection-tool")
from analysis.pricing_engine import (
    calculate_ksa_fbn_profit,
    determine_size_tier,
    calculate_commission,
    lookup_fbn_fee,
)
from config.v6_data import (
    CATEGORIES,
    CATEGORY_INDEX,
    FBN_SIZE_TIERS,
    FBN_KSA_FEES,
    KSA_VAT,
)
from config.cost_defaults import get_defaults
from config.category_mapping import map_category

# ── FBP 配送费表 (从 HTML JS 迁移) ──
# [max_weight, uae_dropoff, uae_pickup, ksa_dropoff, ksa_pickup]
FBP_FEES = [
    [0.25, 12, 14, 17, 19],
    [0.50, 12.5, 14.5, 17.5, 19.5],
    [1, 13, 15, 18.5, 20.5],
    [1.5, 13.5, 15.5, 19.5, 21.5],
    [2, 14, 16, 20.5, 22.5],
    [3, 15.5, 17.5, 27, 29],
    [4, 16.5, 18.5, 28, 30],
    [5, 17.5, 19.5, 29, 31],
    [6, 18.5, 20.5, 30, 32],
    [7, 19.5, 21.5, 31, 33],
    [8, 20.5, 22.5, 32, 34],
    [9, 21.5, 23.5, 33, 35],
    [10, 22.5, 24.5, 34, 36],
    [15, 30, 32, 39, 41],
    [20, 34, 36, 43, 45],
    [25, 38, 40, 47, 49],
    [30, 42, 44, 51, 53],
]


# ═══════════════════════════════════════════════════════════
#  2.1  NoonPricingEngine — FBN + FBP 双模式定价计算器
# ═══════════════════════════════════════════════════════════

class NoonPricingEngine:
    """
    自包含的 KSA 定价计算器, 同时输出 FBN 和 FBP 两种模式的完整利润。
    复用 analysis.pricing_engine 中的核心函数, 新增 FBP 模式计算。
    """

    KSA_RESTOCK_FEE = 3.0  # SAR

    def __init__(
        self,
        exchange_rate: float = 1.90,
        ad_rate: float = 0.05,
        return_rate: float = 0.03,
        shipping_rate_cbm: float = 2000,
        shipping_rate_cbm_peace: float = 950,
    ):
        self.exchange_rate = exchange_rate
        self.ad_rate = ad_rate
        self.return_rate = return_rate
        self.shipping_rate_cbm = shipping_rate_cbm
        self.shipping_rate_cbm_peace = shipping_rate_cbm_peace

    # ── FBP 配送费查表 ──
    @staticmethod
    def lookup_fbp_fee(
        chargeable_weight: float,
        country: str = "KSA",
        mode: str = "dropoff",
    ) -> float:
        """FBP 配送费查表 (KSA Drop-off 为默认)"""
        for row in FBP_FEES:
            if chargeable_weight <= row[0]:
                if country == "UAE":
                    return row[1] if mode == "dropoff" else row[2]
                return row[3] if mode == "dropoff" else row[4]
        last = FBP_FEES[-1]
        if country == "UAE":
            return last[1] if mode == "dropoff" else last[2]
        return last[3] if mode == "dropoff" else last[4]

    # ── FBP Reimbursement (补贴) ──
    @staticmethod
    def get_reimbursement(price_sar: float, country: str = "KSA") -> float:
        """
        FBP 运费补贴 — 作为正向收入计入利润端。
        KSA: <100 SAR → 12, 100-500 → 6, >500 → 0
        UAE: <100 → 10, >=100 → 0
        """
        if country == "UAE":
            return 10.0 if price_sar < 100 else 0.0
        # KSA
        if price_sar < 100:
            return 12.0
        elif price_sar <= 500:
            return 6.0
        else:
            return 0.0

    def calculate_fbn(
        self,
        purchase_price_cny: float,
        length_cm: float,
        width_cm: float,
        height_cm: float,
        weight_kg: float,
        target_price_sar: float,
        category_name: str,
    ) -> dict:
        """KSA FBN 模式完整利润计算 (委托给已验证的引擎)"""
        return calculate_ksa_fbn_profit(
            purchase_price_cny=purchase_price_cny,
            length_cm=length_cm,
            width_cm=width_cm,
            height_cm=height_cm,
            weight_kg=weight_kg,
            target_price_sar=target_price_sar,
            category_name=category_name,
            shipping_rate_cbm=self.shipping_rate_cbm,
            shipping_rate_cbm_peacetime=self.shipping_rate_cbm_peace,
            exchange_rate=self.exchange_rate,
            ad_rate=self.ad_rate,
            return_rate=self.return_rate,
        )

    def calculate_fbp(
        self,
        purchase_price_cny: float,
        length_cm: float,
        width_cm: float,
        height_cm: float,
        weight_kg: float,
        target_price_sar: float,
        category_name: str,
    ) -> dict:
        """
        KSA FBP (海外仓) 模式完整利润计算。

        与 FBN 的关键差异:
        - 佣金使用 ksa_fbp 费率
        - 配送费查 FBP_FEES 表 (KSA Drop-off)
        - 计费重 = MAX(实重, 体积重), 无包装附加, 无尺寸档上限
        - 有 Reimbursement 补贴 (正向收入)
        """
        price = target_price_sar
        exch = self.exchange_rate

        # 1. 体积重
        volume_weight = length_cm * width_cm * height_cm / 5000

        # 2. FBP 计费重 = MAX(实重, 体积重) — 无包装, 无上限
        fbp_cw = max(weight_kg, volume_weight)

        # 3. 尺寸分档 (仅用于展示/风险评估)
        tier = determine_size_tier(length_cm, width_cm, height_cm)
        tier_name = tier[0]

        # 4. VAT 倒算
        vat_amount = price / (1 + KSA_VAT) * KSA_VAT

        # 5. 佣金 (ksa_fbp 模式)
        commission = calculate_commission(price, category_name, "ksa_fbp")

        # 6. 佣金 VAT
        commission_vat = commission * KSA_VAT

        # 7. FBP 配送费 (KSA Drop-off)
        fbp_fee = self.lookup_fbp_fee(fbp_cw, "KSA", "dropoff")

        # 8. 退货管理费 (单件)
        return_fee = min(15, commission * 0.2)

        # 9. 广告费
        ads = price * self.ad_rate

        # 10. 退货管理费摊销
        return_admin = return_fee * self.return_rate

        # 11. 物流退货损耗 (FBP 使用配送费 + restock)
        logistics_loss = (fbp_fee + self.KSA_RESTOCK_FEE) * self.return_rate

        # 12. Reimbursement (补贴, 正向收入)
        reimbursement = self.get_reimbursement(price, "KSA")

        # 13. 平台端费用合计 (SAR) — 补贴作为负成本
        platform_costs_sar = (
            vat_amount + commission + commission_vat + fbp_fee
            + ads + return_admin + logistics_loss
            - reimbursement
        )

        # 14. 收入 (CNY)
        revenue_cny = price * exch

        # ── 战时费率 ──
        vol_cbm = length_cm * width_cm * height_cm / 1_000_000
        head_cost_war = vol_cbm * self.shipping_rate_cbm
        total_cost_war = purchase_price_cny + head_cost_war + platform_costs_sar * exch
        net_profit_war = revenue_cny - total_cost_war
        margin_war = net_profit_war / revenue_cny if revenue_cny > 0 else 0
        roi_war = net_profit_war / total_cost_war if total_cost_war > 0 else 0

        # ── 和平时费率 ──
        head_cost_peace = vol_cbm * self.shipping_rate_cbm_peace
        total_cost_peace = purchase_price_cny + head_cost_peace + platform_costs_sar * exch
        net_profit_peace = revenue_cny - total_cost_peace
        margin_peace = net_profit_peace / revenue_cny if revenue_cny > 0 else 0
        roi_peace = net_profit_peace / total_cost_peace if total_cost_peace > 0 else 0

        return {
            "category": category_name,
            "target_price_sar": price,
            "size_tier": tier_name,
            "volume_weight_kg": round(volume_weight, 4),
            "fbp_chargeable_weight_kg": round(fbp_cw, 3),
            # 费用明细 (SAR)
            "vat_sar": round(vat_amount, 2),
            "commission_sar": round(commission, 2),
            "commission_rate_pct": round(commission / price * 100, 2) if price > 0 else 0,
            "commission_vat_sar": round(commission_vat, 2),
            "fbp_fee_sar": round(fbp_fee, 2),
            "return_fee_sar": round(return_fee, 2),
            "ads_sar": round(ads, 2),
            "return_admin_sar": round(return_admin, 4),
            "logistics_loss_sar": round(logistics_loss, 4),
            "reimbursement_sar": round(reimbursement, 2),
            "platform_costs_sar": round(platform_costs_sar, 2),
            # 战时
            "head_cost_war_cny": round(head_cost_war, 2),
            "total_cost_war_cny": round(total_cost_war, 2),
            "net_profit_war_cny": round(net_profit_war, 2),
            "margin_war_pct": round(margin_war * 100, 2),
            "roi_war_pct": round(roi_war * 100, 2),
            # 和平时
            "head_cost_peace_cny": round(head_cost_peace, 2),
            "total_cost_peace_cny": round(total_cost_peace, 2),
            "net_profit_peace_cny": round(net_profit_peace, 2),
            "margin_peace_pct": round(margin_peace * 100, 2),
            "roi_peace_pct": round(roi_peace * 100, 2),
            # 通用
            "revenue_cny": round(revenue_cny, 2),
        }

    def calculate_both(
        self,
        purchase_price_cny: float,
        length_cm: float,
        width_cm: float,
        height_cm: float,
        weight_kg: float,
        target_price_sar: float,
        category_name: str,
    ) -> dict:
        """同时计算 FBN 和 FBP, 返回合并结果 dict"""
        fbn = self.calculate_fbn(
            purchase_price_cny, length_cm, width_cm, height_cm,
            weight_kg, target_price_sar, category_name,
        )
        fbp = self.calculate_fbp(
            purchase_price_cny, length_cm, width_cm, height_cm,
            weight_kg, target_price_sar, category_name,
        )
        return {"fbn": fbn, "fbp": fbp}


# ═══════════════════════════════════════════════════════════
#  2.2  analyze_products — 综合选品分析
# ═══════════════════════════════════════════════════════════

# ── 评级权重 ──
WEIGHT_MARGIN = 0.30       # 利润率 (取 FBN/FBP 较优)
WEIGHT_COMPETITION = 0.25  # 竞争强度 (Express% 低 + Ad 密度低 = 好)
WEIGHT_REVIEW = 0.20       # 评论壁垒 (评论少 = 易进入)
WEIGHT_RISK = 0.15         # 风险 (退货/尺寸/佣金综合)
WEIGHT_ROI = 0.10          # ROI


def _score_margin(margin_pct: float) -> float:
    """利润率 → 0-100 分"""
    if margin_pct >= 40:
        return 100
    elif margin_pct >= 30:
        return 80 + (margin_pct - 30) * 2
    elif margin_pct >= 20:
        return 60 + (margin_pct - 20) * 2
    elif margin_pct >= 10:
        return 40 + (margin_pct - 10) * 2
    elif margin_pct >= 0:
        return margin_pct * 4
    else:
        return 0


def _score_competition(express_pct: float, ad_density: float) -> float:
    """竞争评分: Express% 低 + 广告少 = 高分"""
    # Express%: 0-20% → 100, 20-50% → 60-100, 50-80% → 20-60, >80% → 0-20
    if express_pct <= 20:
        e_score = 100
    elif express_pct <= 50:
        e_score = 100 - (express_pct - 20) * (40 / 30)
    elif express_pct <= 80:
        e_score = 60 - (express_pct - 50) * (40 / 30)
    else:
        e_score = max(0, 20 - (express_pct - 80) * 1)

    # Ad density: 0% → 100, 10% → 60, 30% → 20, >50% → 0
    if ad_density <= 5:
        a_score = 100
    elif ad_density <= 15:
        a_score = 100 - (ad_density - 5) * 4
    elif ad_density <= 30:
        a_score = 60 - (ad_density - 15) * (40 / 15)
    else:
        a_score = max(0, 20 - (ad_density - 30) * 1)

    return e_score * 0.6 + a_score * 0.4


def _score_review_barrier(review_count: int) -> float:
    """评论壁垒评分: 评论越少越容易进入"""
    if review_count <= 5:
        return 100
    elif review_count <= 20:
        return 80 + (20 - review_count) * (20 / 15)
    elif review_count <= 100:
        return 40 + (100 - review_count) * (40 / 80)
    elif review_count <= 500:
        return 10 + (500 - review_count) * (30 / 400)
    else:
        return max(0, 10 - (review_count - 500) * 0.005)


def _score_risk(
    price_sar: float,
    size_tier: str,
    commission_pct: float,
) -> float:
    """风险评分: 低风险 = 高分"""
    score = 100

    # 高价退货风险: >200 SAR 扣分
    if price_sar > 500:
        score -= 30
    elif price_sar > 200:
        score -= 15
    elif price_sar > 100:
        score -= 5

    # 尺寸风险
    oversize_tiers = {"Oversize", "Extra Oversize", "Bulky"}
    if size_tier in oversize_tiers:
        score -= 25
    elif size_tier == "Standard Parcel":
        score -= 0  # normal
    # small is good, no deduction

    # 佣金风险: >15% 扣分
    if commission_pct > 25:
        score -= 30
    elif commission_pct > 20:
        score -= 20
    elif commission_pct > 15:
        score -= 10

    return max(0, score)


def _compute_rating(total_score: float) -> str:
    """总分 → S/A/B/C/D 评级"""
    if total_score >= 85:
        return "S"
    elif total_score >= 70:
        return "A"
    elif total_score >= 55:
        return "B"
    elif total_score >= 40:
        return "C"
    else:
        return "D"


def _risk_flags(
    price_sar: float,
    size_tier: str,
    commission_pct: float,
) -> list[str]:
    """生成风险标签列表"""
    flags = []
    if price_sar > 200:
        flags.append("HIGH_PRICE_RETURN")
    oversize_tiers = {"Oversize", "Extra Oversize", "Bulky"}
    if size_tier in oversize_tiers:
        flags.append("OVERSIZE")
    if commission_pct > 15:
        flags.append(f"HIGH_COMM_{commission_pct:.0f}%")
    return flags


def analyze_products(
    engine: NoonPricingEngine,
    input_path: Optional[str] = None,
    output_path: Optional[str] = None,
    top_n: int = 100,
) -> pd.DataFrame:
    """
    读取爬虫报告, 对产品进行全维度选品分析。

    查找顺序:
    1. input_path (如果指定)
    2. D:/claude noon/report_automotive_enhanced.xlsx
    3. D:/claude noon/noon-selection-tool/data/monitoring/categories/automotive/report_automotive.xlsx

    返回分析后的 DataFrame。
    """

    # ── 查找输入文件 ──
    search_paths = []
    if input_path:
        search_paths.append(Path(input_path))
    search_paths.extend([
        Path("D:/claude noon/noon-selection-tool/data/monitoring/categories/automotive/report_automotive.xlsx"),
        Path("D:/claude noon/report_automotive_enhanced.xlsx"),
    ])

    src_path = None
    for p in search_paths:
        if p.exists():
            src_path = p
            break

    if src_path is None:
        raise FileNotFoundError(
            f"找不到输入文件, 已搜索: {[str(p) for p in search_paths]}"
        )

    print(f"[INFO] 读取数据源: {src_path}")

    # ── 读取 Excel ──
    xls = pd.ExcelFile(src_path, engine="openpyxl")

    # 优先读 "All Products" sheet, fallback 到第一个 sheet
    if "All Products" in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name="All Products")
    else:
        df = pd.read_excel(xls, sheet_name=0)

    print(f"[INFO] 原始数据: {len(df)} 行, 列: {list(df.columns)}")

    # ── 读取子品类概览 (用于 Sheet 3) ──
    subcat_df = None
    if "Sub-Category Overview" in xls.sheet_names:
        subcat_df = pd.read_excel(xls, sheet_name="Sub-Category Overview")

    # ── 数据清洗 ──
    # 确保 price 列存在且为数值
    if "price" not in df.columns:
        raise ValueError("输入数据缺少 'price' 列")

    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df = df.dropna(subset=["price"])
    df = df[df["price"] > 0].copy()

    # 确保关键列存在 (填充默认值)
    col_defaults = {
        "rating": 0,
        "review_count": 0,
        "is_express": False,
        "is_ad": False,
        "is_bestseller": False,
        "_subcategory": "Unknown",
        "v6_category": "",
        "title": "",
        "product_id": "",
    }
    for col, default in col_defaults.items():
        if col not in df.columns:
            df[col] = default
        else:
            df[col] = df[col].fillna(default)

    df["review_count"] = pd.to_numeric(df["review_count"], errors="coerce").fillna(0).astype(int)
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce").fillna(0)

    # ── 取 Top N (按 price 降序, 过滤极端值) ──
    df = df[df["price"] <= 5000].copy()  # 排除极端高价
    df = df.sort_values("price", ascending=False).head(top_n).reset_index(drop=True)

    print(f"[INFO] 分析产品数: {len(df)}")

    # ── 计算子品类级竞争指标 ──
    subcat_stats = {}
    if "_subcategory" in df.columns:
        for subcat, grp in df.groupby("_subcategory"):
            total = len(grp)
            express_count = grp["is_express"].sum() if "is_express" in grp.columns else 0
            ad_count = grp["is_ad"].sum() if "is_ad" in grp.columns else 0
            subcat_stats[subcat] = {
                "express_pct": (express_count / total * 100) if total > 0 else 0,
                "ad_density": (ad_count / total * 100) if total > 0 else 0,
                "avg_reviews": grp["review_count"].mean(),
                "median_price": grp["price"].median(),
                "count": total,
            }

    # ── 逐产品计算利润 + 评级 ──
    results = []

    for idx, row in df.iterrows():
        try:
            price_sar = float(row["price"])
            cat_name = str(row.get("v6_category", ""))

            # 品类映射 (如果 v6_category 为空或无效)
            if not cat_name or cat_name not in CATEGORY_INDEX:
                subcat = str(row.get("_subcategory", ""))
                cat_name = map_category("Automotive", [subcat])

            # 获取默认尺寸/重量
            defaults = get_defaults(cat_name)
            l_cm = defaults["l"]
            w_cm = defaults["w"]
            h_cm = defaults["h"]
            wt_kg = defaults["wt"]
            cost_ratio = defaults["cost_ratio"]

            # 采购价估算
            purchase_cny = price_sar * engine.exchange_rate * cost_ratio

            # ── FBN 利润计算 ──
            fbn = engine.calculate_fbn(
                purchase_price_cny=purchase_cny,
                length_cm=l_cm, width_cm=w_cm, height_cm=h_cm,
                weight_kg=wt_kg, target_price_sar=price_sar,
                category_name=cat_name,
            )

            # ── FBP 利润计算 ──
            fbp = engine.calculate_fbp(
                purchase_price_cny=purchase_cny,
                length_cm=l_cm, width_cm=w_cm, height_cm=h_cm,
                weight_kg=wt_kg, target_price_sar=price_sar,
                category_name=cat_name,
            )

            # ── 竞争分析 ──
            subcat = str(row.get("_subcategory", "Unknown"))
            stats = subcat_stats.get(subcat, {
                "express_pct": 50, "ad_density": 10, "avg_reviews": 50,
            })
            express_pct = stats["express_pct"]
            ad_density = stats["ad_density"]
            review_count = int(row["review_count"])

            # ── 取两个模式中较优的利润率 ──
            best_margin_war = max(fbn["margin_war_pct"], fbp["margin_war_pct"])
            best_margin_peace = max(fbn["margin_peace_pct"], fbp["margin_peace_pct"])
            best_roi_war = max(fbn["roi_war_pct"], fbp["roi_war_pct"])
            commission_pct = fbn["commission_rate_pct"]

            # ── 评分 ──
            s_margin = _score_margin(best_margin_war)
            s_comp = _score_competition(express_pct, ad_density)
            s_review = _score_review_barrier(review_count)
            s_risk = _score_risk(price_sar, fbn["size_tier"], commission_pct)
            s_roi = min(100, max(0, best_roi_war))  # ROI% capped at 100

            total_score = (
                s_margin * WEIGHT_MARGIN
                + s_comp * WEIGHT_COMPETITION
                + s_review * WEIGHT_REVIEW
                + s_risk * WEIGHT_RISK
                + s_roi * WEIGHT_ROI
            )

            rating = _compute_rating(total_score)
            flags = _risk_flags(price_sar, fbn["size_tier"], commission_pct)

            results.append({
                # 基础信息
                "product_id": str(row.get("product_id", "")),
                "title": str(row.get("title", ""))[:80],
                "subcategory": subcat,
                "v6_category": cat_name,
                "price_sar": price_sar,
                "rating_stars": round(float(row["rating"]), 1),
                "review_count": review_count,
                "is_express": bool(row.get("is_express", False)),
                "is_ad": bool(row.get("is_ad", False)),
                "is_bestseller": bool(row.get("is_bestseller", False)),

                # 评级
                "grade": rating,
                "total_score": round(total_score, 1),
                "score_margin": round(s_margin, 1),
                "score_competition": round(s_comp, 1),
                "score_review": round(s_review, 1),
                "score_risk": round(s_risk, 1),

                # FBN 利润
                "fbn_margin_war_pct": fbn["margin_war_pct"],
                "fbn_margin_peace_pct": fbn["margin_peace_pct"],
                "fbn_profit_war_cny": fbn["net_profit_war_cny"],
                "fbn_profit_peace_cny": fbn["net_profit_peace_cny"],
                "fbn_roi_war_pct": fbn["roi_war_pct"],

                # FBP 利润
                "fbp_margin_war_pct": fbp["margin_war_pct"],
                "fbp_margin_peace_pct": fbp["margin_peace_pct"],
                "fbp_profit_war_cny": fbp["net_profit_war_cny"],
                "fbp_profit_peace_cny": fbp["net_profit_peace_cny"],
                "fbp_roi_war_pct": fbp["roi_war_pct"],

                # 最优利润
                "best_margin_war_pct": best_margin_war,
                "best_margin_peace_pct": best_margin_peace,

                # 竞争
                "express_pct": round(express_pct, 1),
                "ad_density_pct": round(ad_density, 1),

                # 风险
                "risk_flags": ", ".join(flags) if flags else "",
                "size_tier": fbn["size_tier"],
                "commission_pct": commission_pct,

                # FBN 费用明细
                "fbn_vat_sar": fbn["vat_sar"],
                "fbn_commission_sar": fbn["commission_sar"],
                "fbn_commission_vat_sar": fbn["commission_vat_sar"],
                "fbn_fee_sar": fbn["fbn_fee_sar"],
                "fbn_ads_sar": fbn["ads_sar"],
                "fbn_return_admin_sar": fbn["return_admin_sar"],
                "fbn_logistics_loss_sar": fbn["logistics_loss_sar"],
                "fbn_platform_costs_sar": fbn["platform_costs_sar"],
                "fbn_head_cost_war_cny": fbn["head_cost_war_cny"],
                "fbn_head_cost_peace_cny": fbn["head_cost_peace_cny"],
                "fbn_total_cost_war_cny": fbn["total_cost_war_cny"],
                "fbn_total_cost_peace_cny": fbn["total_cost_peace_cny"],

                # FBP 费用明细
                "fbp_vat_sar": fbp["vat_sar"],
                "fbp_commission_sar": fbp["commission_sar"],
                "fbp_commission_vat_sar": fbp["commission_vat_sar"],
                "fbp_delivery_fee_sar": fbp["fbp_fee_sar"],
                "fbp_reimbursement_sar": fbp["reimbursement_sar"],
                "fbp_ads_sar": fbp["ads_sar"],
                "fbp_return_admin_sar": fbp["return_admin_sar"],
                "fbp_logistics_loss_sar": fbp["logistics_loss_sar"],
                "fbp_platform_costs_sar": fbp["platform_costs_sar"],
                "fbp_head_cost_war_cny": fbp["head_cost_war_cny"],
                "fbp_head_cost_peace_cny": fbp["head_cost_peace_cny"],
                "fbp_total_cost_war_cny": fbp["total_cost_war_cny"],
                "fbp_total_cost_peace_cny": fbp["total_cost_peace_cny"],

                # 通用
                "purchase_est_cny": round(purchase_cny, 2),
                "revenue_cny": fbn["revenue_cny"],
                "volume_weight_kg": fbn["volume_weight_kg"],
                "fbn_cw_kg": fbn["fbn_chargeable_weight_kg"],
                "fbp_cw_kg": fbp["fbp_chargeable_weight_kg"],

                # URL
                "product_url": str(row.get("product_url", "")),
            })

        except Exception as e:
            print(f"[WARN] 跳过产品 {row.get('product_id', '?')}: {e}")
            continue

    result_df = pd.DataFrame(results)
    print(f"[INFO] 成功分析 {len(result_df)} 个产品")

    # ── 输出 Excel ──
    if output_path is None:
        output_path = "D:/claude noon/selection_report_automotive.xlsx"

    _write_report(result_df, subcat_df, output_path)

    return result_df


# ═══════════════════════════════════════════════════════════
#  2.3  Excel 报告输出
# ═══════════════════════════════════════════════════════════

# ── 样式常量 ──
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

RATING_FILLS = {
    "S": PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),  # dark green
    "A": PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid"),  # green
    "B": PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),  # blue
    "C": PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),  # yellow
    "D": PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"),  # red
}
RATING_FONTS = {
    "S": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
    "A": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
    "B": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
    "C": Font(name="Calibri", size=12, bold=True, color="000000"),
    "D": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
}


def _apply_header(ws, headers: list[str], col_widths: Optional[list[int]] = None):
    """写入表头行并应用样式"""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def _write_cell(ws, row: int, col: int, value, fmt: Optional[str] = None):
    """写入单元格并应用基础样式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _write_report(df: pd.DataFrame, subcat_df: Optional[pd.DataFrame], output_path: str):
    """输出三 Sheet 的选品报告"""

    wb = Workbook()

    # ════════════════════════════════════════════
    # Sheet 1: Selection Dashboard
    # ════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Selection Dashboard"

    headers1 = [
        "Grade", "Score", "Product ID", "Title", "Subcategory",
        "Price (SAR)", "Reviews", "Rating",
        "FBN Margin% (War)", "FBP Margin% (War)",
        "Best Margin% (War)", "Best Margin% (Peace)",
        "Express%", "Ad Density%", "Risk Flags",
        "Commission%", "Size Tier",
        "Is Express", "Is Ad", "Is Bestseller",
        "URL",
    ]
    widths1 = [8, 8, 24, 40, 22, 12, 10, 8, 14, 14, 14, 14, 10, 12, 24, 12, 16, 10, 8, 12, 50]
    _apply_header(ws1, headers1, widths1)

    # 按评级排序: S > A > B > C > D, 同级按 score 降序
    rating_order = {"S": 0, "A": 1, "B": 2, "C": 3, "D": 4}
    df_sorted = df.sort_values(
        by=["grade", "total_score"],
        key=lambda s: s.map(rating_order) if s.name == "grade" else -s,
        ascending=[True, True],
    ).reset_index(drop=True)

    for i, row in df_sorted.iterrows():
        r = i + 2
        grade = row["grade"]

        # Grade cell with color
        cell_grade = _write_cell(ws1, r, 1, grade)
        cell_grade.fill = RATING_FILLS.get(grade, PatternFill())
        cell_grade.font = RATING_FONTS.get(grade, Font())
        cell_grade.alignment = Alignment(horizontal="center", vertical="center")

        _write_cell(ws1, r, 2, row["total_score"], "0.0")
        _write_cell(ws1, r, 3, row["product_id"])
        _write_cell(ws1, r, 4, row["title"])
        _write_cell(ws1, r, 5, row["subcategory"])
        _write_cell(ws1, r, 6, row["price_sar"], "#,##0")
        _write_cell(ws1, r, 7, row["review_count"], "#,##0")
        _write_cell(ws1, r, 8, row["rating_stars"], "0.0")
        _write_cell(ws1, r, 9, row["fbn_margin_war_pct"], "0.00")
        _write_cell(ws1, r, 10, row["fbp_margin_war_pct"], "0.00")
        _write_cell(ws1, r, 11, row["best_margin_war_pct"], "0.00")
        _write_cell(ws1, r, 12, row["best_margin_peace_pct"], "0.00")
        _write_cell(ws1, r, 13, row["express_pct"], "0.0")
        _write_cell(ws1, r, 14, row["ad_density_pct"], "0.0")
        _write_cell(ws1, r, 15, row["risk_flags"])
        _write_cell(ws1, r, 16, row["commission_pct"], "0.0")
        _write_cell(ws1, r, 17, row["size_tier"])
        _write_cell(ws1, r, 18, "Yes" if row["is_express"] else "No")
        _write_cell(ws1, r, 19, "Yes" if row["is_ad"] else "No")
        _write_cell(ws1, r, 20, "Yes" if row["is_bestseller"] else "No")
        _write_cell(ws1, r, 21, row["product_url"])

        # 利润率条件着色: 正=绿底, 负=红底
        for col_idx in [9, 10, 11, 12]:
            cell = ws1.cell(row=r, column=col_idx)
            val = cell.value
            if val is not None and val < 0:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif val is not None and val >= 20:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    # ════════════════════════════════════════════
    # Sheet 2: Profit Details (FBN vs FBP 对比)
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet("Profit Details")

    headers2 = [
        "Product ID", "Title", "Price (SAR)", "V6 Category",
        "Purchase (CNY)", "Revenue (CNY)", "Vol Weight (kg)",
        # FBN 列
        "FBN CW (kg)", "FBN Size Tier",
        "FBN VAT", "FBN Commission", "FBN Comm VAT",
        "FBN Fee", "FBN Ads", "FBN Return Admin", "FBN Logistics Loss",
        "FBN Platform Total (SAR)",
        "FBN Head War (CNY)", "FBN Total Cost War (CNY)",
        "FBN Profit War (CNY)", "FBN Margin War%",
        "FBN Head Peace (CNY)", "FBN Total Cost Peace (CNY)",
        "FBN Profit Peace (CNY)", "FBN Margin Peace%",
        # FBP 列
        "FBP CW (kg)",
        "FBP VAT", "FBP Commission", "FBP Comm VAT",
        "FBP Delivery Fee", "FBP Reimb", "FBP Ads",
        "FBP Return Admin", "FBP Logistics Loss",
        "FBP Platform Total (SAR)",
        "FBP Head War (CNY)", "FBP Total Cost War (CNY)",
        "FBP Profit War (CNY)", "FBP Margin War%",
        "FBP Head Peace (CNY)", "FBP Total Cost Peace (CNY)",
        "FBP Profit Peace (CNY)", "FBP Margin Peace%",
    ]
    widths2 = [24, 35, 10, 20, 12, 12, 10] + [10] * (len(headers2) - 7)
    _apply_header(ws2, headers2, widths2)

    detail_cols = [
        ("product_id", None), ("title", None), ("price_sar", "#,##0"),
        ("v6_category", None), ("purchase_est_cny", "0.00"), ("revenue_cny", "0.00"),
        ("volume_weight_kg", "0.000"),
        # FBN
        ("fbn_cw_kg", "0.000"), ("size_tier", None),
        ("fbn_vat_sar", "0.00"), ("fbn_commission_sar", "0.00"),
        ("fbn_commission_vat_sar", "0.00"), ("fbn_fee_sar", "0.00"),
        ("fbn_ads_sar", "0.00"), ("fbn_return_admin_sar", "0.0000"),
        ("fbn_logistics_loss_sar", "0.0000"), ("fbn_platform_costs_sar", "0.00"),
        ("fbn_head_cost_war_cny", "0.00"), ("fbn_total_cost_war_cny", "0.00"),
        ("fbn_profit_war_cny", "0.00"), ("fbn_margin_war_pct", "0.00"),
        ("fbn_head_cost_peace_cny", "0.00"), ("fbn_total_cost_peace_cny", "0.00"),
        ("fbn_profit_peace_cny", "0.00"), ("fbn_margin_peace_pct", "0.00"),
        # FBP
        ("fbp_cw_kg", "0.000"),
        ("fbp_vat_sar", "0.00"), ("fbp_commission_sar", "0.00"),
        ("fbp_commission_vat_sar", "0.00"), ("fbp_delivery_fee_sar", "0.00"),
        ("fbp_reimbursement_sar", "0.00"), ("fbp_ads_sar", "0.00"),
        ("fbp_return_admin_sar", "0.0000"), ("fbp_logistics_loss_sar", "0.0000"),
        ("fbp_platform_costs_sar", "0.00"),
        ("fbp_head_cost_war_cny", "0.00"), ("fbp_total_cost_war_cny", "0.00"),
        ("fbp_profit_war_cny", "0.00"), ("fbp_margin_war_pct", "0.00"),
        ("fbp_head_cost_peace_cny", "0.00"), ("fbp_total_cost_peace_cny", "0.00"),
        ("fbp_profit_peace_cny", "0.00"), ("fbp_margin_peace_pct", "0.00"),
    ]

    for i, row in df_sorted.iterrows():
        r = i + 2
        for c, (col_name, fmt) in enumerate(detail_cols, 1):
            val = row.get(col_name, "")
            _write_cell(ws2, r, c, val, fmt)

    # ════════════════════════════════════════════
    # Sheet 3: Category Strategy
    # ════════════════════════════════════════════
    ws3 = wb.create_sheet("Category Strategy")

    headers3 = [
        "Subcategory", "Products Analyzed", "Express%", "Ad Density%",
        "Median Price (SAR)", "Avg Reviews",
        "Avg FBN Margin% (War)", "Avg FBP Margin% (War)",
        "S+A Count", "B Count", "C+D Count",
        "Recommended Mode", "Entry Difficulty", "Strategy",
    ]
    widths3 = [24, 14, 10, 12, 14, 12, 16, 16, 10, 10, 10, 16, 14, 50]
    _apply_header(ws3, headers3, widths3)

    # 按子品类聚合
    if len(df_sorted) > 0:
        subcat_groups = df_sorted.groupby("subcategory")
        r = 2
        for subcat, grp in subcat_groups:
            count = len(grp)
            express_pct = grp["express_pct"].mean()
            ad_density = grp["ad_density_pct"].mean()
            median_price = grp["price_sar"].median()
            avg_reviews = grp["review_count"].mean()
            avg_fbn_margin = grp["fbn_margin_war_pct"].mean()
            avg_fbp_margin = grp["fbp_margin_war_pct"].mean()

            sa_count = len(grp[grp["grade"].isin(["S", "A"])])
            b_count = len(grp[grp["grade"] == "B"])
            cd_count = len(grp[grp["grade"].isin(["C", "D"])])

            # 推荐模式
            if avg_fbp_margin > avg_fbn_margin + 2:
                rec_mode = "FBP"
            elif avg_fbn_margin > avg_fbp_margin + 2:
                rec_mode = "FBN"
            else:
                rec_mode = "Both OK"

            # 进入难度
            if express_pct > 50 and avg_reviews > 100:
                difficulty = "Hard"
            elif express_pct > 30 or avg_reviews > 50:
                difficulty = "Medium"
            else:
                difficulty = "Easy"

            # 策略建议
            strategies = []
            if avg_fbn_margin > 20 or avg_fbp_margin > 20:
                strategies.append("High margin, priority entry")
            elif avg_fbn_margin > 10 or avg_fbp_margin > 10:
                strategies.append("Moderate margin, selective entry")
            else:
                strategies.append("Low margin, cautious entry")

            if express_pct < 20:
                strategies.append("Low Express penetration, good opportunity")
            if avg_reviews < 20:
                strategies.append("Low review barrier")
            if median_price > 200:
                strategies.append("High-price niche, watch return risk")
            if ad_density < 5:
                strategies.append("Low ad competition")

            strategy_text = "; ".join(strategies)

            _write_cell(ws3, r, 1, subcat)
            _write_cell(ws3, r, 2, count, "#,##0")
            _write_cell(ws3, r, 3, round(express_pct, 1), "0.0")
            _write_cell(ws3, r, 4, round(ad_density, 1), "0.0")
            _write_cell(ws3, r, 5, round(median_price, 0), "#,##0")
            _write_cell(ws3, r, 6, round(avg_reviews, 0), "#,##0")
            _write_cell(ws3, r, 7, round(avg_fbn_margin, 2), "0.00")
            _write_cell(ws3, r, 8, round(avg_fbp_margin, 2), "0.00")
            _write_cell(ws3, r, 9, sa_count, "#,##0")
            _write_cell(ws3, r, 10, b_count, "#,##0")
            _write_cell(ws3, r, 11, cd_count, "#,##0")
            _write_cell(ws3, r, 12, rec_mode)
            _write_cell(ws3, r, 13, difficulty)
            _write_cell(ws3, r, 14, strategy_text)

            # 颜色: Easy=绿, Medium=黄, Hard=红
            diff_cell = ws3.cell(row=r, column=13)
            if difficulty == "Easy":
                diff_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif difficulty == "Medium":
                diff_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            else:
                diff_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

            r += 1

    # ── 保存 ──
    wb.save(output_path)
    print(f"[OK] 选品报告已生成: {output_path}")


# ═══════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("Noon 选品决策引擎 v1.0")
    print("=" * 60)

    engine = NoonPricingEngine(
        exchange_rate=1.90,
        ad_rate=0.05,
        return_rate=0.03,
        shipping_rate_cbm=2000,
        shipping_rate_cbm_peace=950,
    )

    # ── 快速验证 FBP 计算 ──
    print("\n--- FBP 引擎验证 ---")
    fbp_result = engine.calculate_fbp(
        purchase_price_cny=25,
        length_cm=15, width_cm=10, height_cm=8,
        weight_kg=0.8,
        target_price_sar=89,
        category_name="厨具/床品/卫浴/装饰 Home",
    )
    print(f"  FBP CW:           {fbp_result['fbp_chargeable_weight_kg']} kg")
    print(f"  FBP 配送费:       {fbp_result['fbp_fee_sar']} SAR")
    print(f"  FBP 补贴:         {fbp_result['reimbursement_sar']} SAR")
    print(f"  FBP 佣金:         {fbp_result['commission_sar']} SAR ({fbp_result['commission_rate_pct']}%)")
    print(f"  FBP 平台费合计:   {fbp_result['platform_costs_sar']} SAR")
    print(f"  FBP 战时利润率:   {fbp_result['margin_war_pct']}%")
    print(f"  FBP 和平利润率:   {fbp_result['margin_peace_pct']}%")

    # ── 运行选品分析 ──
    print("\n" + "=" * 60)
    print("运行 Automotive 品类选品分析...")
    print("=" * 60)

    try:
        result_df = analyze_products(engine)
        print(f"\n--- 评级分布 ---")
        if len(result_df) > 0:
            grade_dist = result_df["grade"].value_counts().sort_index()
            for grade, count in grade_dist.items():
                print(f"  {grade}: {count} 个产品")
            print(f"\n--- Top 5 产品 ---")
            top5 = result_df.sort_values("total_score", ascending=False).head(5)
            for _, p in top5.iterrows():
                print(f"  [{p['grade']}] {p['total_score']:.1f}分 | "
                      f"SAR {p['price_sar']:.0f} | "
                      f"FBN {p['fbn_margin_war_pct']:.1f}% / FBP {p['fbp_margin_war_pct']:.1f}% | "
                      f"{p['title'][:50]}")
    except Exception as e:
        print(f"[ERROR] 分析失败: {e}")
        import traceback
        traceback.print_exc()
