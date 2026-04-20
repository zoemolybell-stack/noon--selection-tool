#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
noon_seller_analyzer.py — Noon 卖家竞争分析工具

分析维度:
  1. 卖家/品牌集中度 (按子类目): CR1/CR3/CR5, HHI, 竞争格局分类
  2. 卖家类型分布: Express vs 普通, 高评价 vs 低评价
  3. 品牌集中度: 品牌数, Top品牌份额, 白牌率 (中国卖家机会)
  4. 进入难度: 综合评分 → Easy/Moderate/Hard/Very Hard

用法:
  python noon_seller_analyzer.py --input seller_data.xlsx --output seller_report.xlsx
  (若 input 不存在, 自动降级到 automotive 监控数据)
"""

import sys
import os
import argparse
import math
import re
from collections import Counter, defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, 'D:/claude noon')
from noon_config import CONFIG

# ============================================================
# 常量
# ============================================================

FALLBACK_PATH = r'D:\claude noon\noon-selection-tool\data\monitoring\categories\automotive\report_automotive.xlsx'
FALLBACK_SHEET = 'All Products'

DARK_BLUE_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

DIFFICULTY_COLORS = {
    'Easy':      PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
    'Moderate':  PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),
    'Hard':      PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    'Very Hard': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
}
DIFFICULTY_FONTS = {
    'Easy':      Font(name='Calibri', bold=True, color='006100'),
    'Moderate':  Font(name='Calibri', bold=True, color='806000'),
    'Hard':      Font(name='Calibri', bold=True, color='833C0B'),
    'Very Hard': Font(name='Calibri', bold=True, color='FFFFFF'),
}

# 品牌识别: 通用产品词 (遇到这些词就截断)
GENERIC_WORDS = {
    'car', 'auto', 'for', 'universal', 'with', 'and', 'set', 'kit', 'pack',
    'pcs', 'pair', 'inch', 'the', 'new', 'premium', 'professional', 'portable',
    'electric', 'digital', 'mini', 'led', 'lcd', 'usb', 'wireless',
    'waterproof', 'heavy', 'duty', 'stainless', 'steel', 'silicone', 'rubber',
    'plastic', 'metal', 'aluminum', 'carbon', 'fiber', 'black', 'red', 'blue',
    'white', 'green', 'grey', 'gray', 'silver', 'gold', 'multi', 'color',
    'high', 'quality', 'super', 'extra', 'large', 'small', 'big', 'medium',
    'motorcycle', 'truck', 'vehicle', 'suv', 'sedan', 'pickup',
    'front', 'rear', 'left', 'right', 'top', 'bottom', 'side', 'inner', 'outer',
    'compatible', 'replacement', 'upgrade', 'adjustable', 'foldable',
    'type', 'style', 'model', 'version', 'series', 'generation',
    '12v', '24v', '110v', '220v', '1pc', '2pcs', '3pcs', '4pcs', '5pcs',
    '6pcs', '8pcs', '10pcs', '12pcs', '20pcs', '50pcs', '100pcs',
}

# 已知汽配品牌 (用于优先匹配)
KNOWN_BRANDS = {
    'BOSCH', 'PHILIPS', 'OSRAM', 'DENSO', 'NGK', 'CASTROL', 'MOBIL', 'SHELL',
    'MICHELIN', 'BRIDGESTONE', 'CONTINENTAL', 'GOODYEAR', 'PIRELLI', 'DUNLOP',
    'HELLA', 'MANN', 'MAHLE', 'BREMBO', 'BILSTEIN', 'KYB', 'MOOG', 'TRW',
    'VALEO', 'SACHS', 'FEBI', 'MEYLE', 'LEMFORDER', 'SKF', 'GATES', 'DAYCO',
    'ACDelco', 'ACDELCO', 'MOTORCRAFT', 'MOPAR', 'GENUINE', 'OEM',
    'BASEUS', 'XIAOMI', 'ANKER', 'AUKEY', 'SCOSCHE', 'GARMIN', 'PIONEER',
    'KENWOOD', 'ALPINE', 'JBL', 'SONY', 'SAMSUNG', 'THINKWARE', 'VIOFO',
    '3M', 'CARPRO', 'MEGUIARS', 'SONAX', 'TURTLE', 'CHEMICAL', 'ARMOR',
    'VLAND', 'LASFIT', 'AUXBEAM', 'NILIGHT', 'SEALEY', 'DEWALT', 'STANLEY',
    'CRAFTSMAN', 'TEKTON', 'GEARWRENCH', 'SNAP-ON', 'WERA', 'KNIPEX',
    'AUTODECO', 'SMILYEEZ', 'ALTITUDECRAFT',
}
KNOWN_BRANDS_LOWER = {b.lower() for b in KNOWN_BRANDS}

# 白牌 / 无品牌 关键词
WHITE_LABEL_INDICATORS = {
    'generic', 'unbranded', 'no brand', 'oem', 'compatible', 'replacement',
    'aftermarket', 'unknown', 'white_label', '',
}


# ============================================================
# 品牌提取
# ============================================================

def extract_brand(title):
    """
    从产品标题中提取品牌名。
    启发式: 取标题开头1-3个看起来像品牌的词 (大写/CamelCase/已知品牌)。
    """
    if not title or not isinstance(title, str):
        return 'Unknown'

    title = title.strip()
    # 清理特殊字符
    title = re.sub(r'^[\[\(].*?[\]\)]\s*', '', title)  # 去掉开头方括号

    words = title.split()
    if not words:
        return 'Unknown'

    brand_parts = []

    for i, word in enumerate(words[:5]):  # 最多看前5个词
        clean = re.sub(r'[,\-\.\(\)\[\]/:;]', '', word).strip()
        if not clean:
            continue
        lower = clean.lower()

        # 遇到通用词就停
        if lower in GENERIC_WORDS:
            break
        # 遇到纯数字或度量单位也停
        if re.match(r'^\d+(\.\d+)?(mm|cm|m|kg|g|ml|l|pcs|pc|pack|set)?$', lower):
            break

        # 判断是否像品牌
        is_brand_like = False

        # 已知品牌
        if lower in KNOWN_BRANDS_LOWER:
            is_brand_like = True
        # 全大写且>=2字符 (如 BOSCH, NGK)
        elif clean.isupper() and len(clean) >= 2 and clean.isalpha():
            is_brand_like = True
        # 首字母大写 (如 Baseus, Carpro)
        elif clean[0].isupper() and len(clean) >= 2:
            is_brand_like = True
        # 含数字的品牌 (如 3M)
        elif clean in KNOWN_BRANDS:
            is_brand_like = True

        if is_brand_like:
            brand_parts.append(clean)
            if len(brand_parts) >= 3:
                break
        else:
            break

    if not brand_parts:
        return 'Unknown'

    brand = ' '.join(brand_parts)

    # 标准化: 常见品牌统一大小写
    brand_upper = brand.upper()
    for kb in KNOWN_BRANDS:
        if brand_upper == kb.upper():
            return kb
    return brand


def is_white_label(brand):
    """判断是否白牌/无品牌"""
    if not brand:
        return True
    lower = brand.lower().strip()
    if lower in WHITE_LABEL_INDICATORS:
        return True
    if lower == 'unknown':
        return True
    return False


# ============================================================
# 数据加载
# ============================================================

def load_data(input_path):
    """加载数据，支持优雅降级"""
    if input_path and os.path.exists(input_path):
        print(f"[INFO] 正在加载输入文件: {input_path}")
        try:
            df = pd.read_excel(input_path, sheet_name='All Products')
        except Exception:
            df = pd.read_excel(input_path, sheet_name=0)
        print(f"[INFO] 已加载 {len(df)} 行数据")
        return df

    # 降级到 automotive 数据
    print(f"[WARN] 输入文件 '{input_path}' 不存在，降级到 automotive 监控数据")
    if not os.path.exists(FALLBACK_PATH):
        print(f"[ERROR] 降级文件也不存在: {FALLBACK_PATH}")
        sys.exit(1)

    print(f"[INFO] 正在加载降级文件: {FALLBACK_PATH}")
    df = pd.read_excel(FALLBACK_PATH, sheet_name=FALLBACK_SHEET)
    print(f"[INFO] 已加载 {len(df)} 行数据 (automotive 降级模式)")
    return df


def prepare_data(df):
    """数据预处理：提取品牌，标准化字段"""
    # 确保必要列存在
    required = ['title', '_subcategory']
    for col in required:
        if col not in df.columns:
            print(f"[ERROR] 缺少必要列: {col}")
            sys.exit(1)

    # 提取品牌
    print("[INFO] 正在提取品牌信息...")
    df['brand'] = df['title'].apply(extract_brand)
    df['is_white_label'] = df['brand'].apply(is_white_label)

    # 确保 review_count 和 is_express 存在
    if 'review_count' not in df.columns:
        df['review_count'] = 0
    if 'is_express' not in df.columns:
        df['is_express'] = False
    if 'rating' not in df.columns:
        df['rating'] = None
    if 'price' not in df.columns:
        df['price'] = 0

    df['review_count'] = pd.to_numeric(df['review_count'], errors='coerce').fillna(0).astype(int)
    df['is_express'] = df['is_express'].fillna(False).astype(bool)
    df['price'] = pd.to_numeric(df['price'], errors='coerce').fillna(0)

    # 高评价标准: review_count >= 10
    df['is_high_review'] = df['review_count'] >= 10

    print(f"[INFO] 品牌提取完成: {df['brand'].nunique()} 个品牌, "
          f"白牌率 {df['is_white_label'].mean()*100:.1f}%")
    return df


# ============================================================
# 分析函数
# ============================================================

def calc_hhi(shares):
    """计算 HHI 指数 (0-10000)"""
    return sum(s ** 2 for s in shares)


def classify_competition(hhi, cr1):
    """根据 HHI 和 CR1 判断竞争格局"""
    if hhi > 2500 or cr1 > 50:
        return 'Monopoly'
    elif hhi > 1500 or cr1 > 30:
        return 'Oligopoly'
    elif hhi > 500:
        return 'Competitive'
    else:
        return 'Fragmented'


def analyze_subcategory_competition(df):
    """分析每个子类目的卖家/品牌集中度"""
    results = []
    subcats = df['_subcategory'].dropna().unique()

    for subcat in sorted(subcats):
        sub_df = df[df['_subcategory'] == subcat]
        n_products = len(sub_df)
        if n_products < 3:
            continue

        # 品牌集中度
        brand_counts = sub_df['brand'].value_counts()
        n_brands = len(brand_counts)
        brand_shares = (brand_counts / n_products * 100).values

        cr1 = brand_shares[0] if len(brand_shares) >= 1 else 0
        cr3 = sum(brand_shares[:3]) if len(brand_shares) >= 3 else sum(brand_shares)
        cr5 = sum(brand_shares[:5]) if len(brand_shares) >= 5 else sum(brand_shares)
        hhi = calc_hhi(brand_shares)

        competition_type = classify_competition(hhi, cr1)

        # Express 渗透率
        express_ratio = sub_df['is_express'].mean() * 100

        # 高评价比例
        high_review_ratio = sub_df['is_high_review'].mean() * 100

        # 白牌率
        white_label_ratio = sub_df['is_white_label'].mean() * 100

        # 平均评论数
        avg_reviews = sub_df['review_count'].mean()

        # 平均价格
        avg_price = sub_df[sub_df['price'] > 0]['price'].mean() if (sub_df['price'] > 0).any() else 0

        # Top品牌
        top_brand = brand_counts.index[0] if len(brand_counts) > 0 else 'N/A'
        top_brand_share = cr1

        # 进入难度评分 (0-100, 越高越难)
        # 4个维度各25分
        conc_score = min(25, hhi / 400)  # HHI 0-10000 → 0-25
        brand_score = min(25, (100 - white_label_ratio) / 4)  # 白牌率低=品牌壁垒高
        express_score = min(25, express_ratio / 4)  # Express高=门槛高
        review_score = min(25, min(avg_reviews, 100) / 4)  # 评论多=壁垒高

        difficulty_score = conc_score + brand_score + express_score + review_score

        if difficulty_score < 25:
            difficulty = 'Easy'
        elif difficulty_score < 50:
            difficulty = 'Moderate'
        elif difficulty_score < 75:
            difficulty = 'Hard'
        else:
            difficulty = 'Very Hard'

        results.append({
            'subcategory': subcat,
            'product_count': n_products,
            'brand_count': n_brands,
            'cr1_pct': round(cr1, 1),
            'cr3_pct': round(cr3, 1),
            'cr5_pct': round(cr5, 1),
            'hhi': round(hhi, 0),
            'competition_type': competition_type,
            'express_pct': round(express_ratio, 1),
            'high_review_pct': round(high_review_ratio, 1),
            'avg_reviews': round(avg_reviews, 1),
            'avg_price': round(avg_price, 1),
            'white_label_pct': round(white_label_ratio, 1),
            'top_brand': top_brand,
            'top_brand_share_pct': round(top_brand_share, 1),
            'difficulty_score': round(difficulty_score, 1),
            'difficulty': difficulty,
        })

    return pd.DataFrame(results)


def analyze_seller_profiles(df):
    """分析高频品牌/卖家的类目分布"""
    results = []

    brand_stats = df.groupby('brand').agg(
        product_count=('title', 'count'),
        subcategory_count=('_subcategory', 'nunique'),
        subcategories=('_subcategory', lambda x: ', '.join(sorted(x.unique())[:5])),
        avg_price=('price', 'mean'),
        avg_reviews=('review_count', 'mean'),
        max_reviews=('review_count', 'max'),
        express_ratio=('is_express', 'mean'),
        is_white_label=('is_white_label', 'first'),
    ).reset_index()

    # 按产品数排序，取Top品牌
    brand_stats = brand_stats.sort_values('product_count', ascending=False)

    # 过滤掉 Unknown
    brand_stats = brand_stats[brand_stats['brand'] != 'Unknown']

    for _, row in brand_stats.head(200).iterrows():
        subcats_display = row['subcategories']
        if row['subcategory_count'] > 5:
            subcats_display += f' ... (+{row["subcategory_count"]-5} more)'

        results.append({
            'brand': row['brand'],
            'product_count': row['product_count'],
            'subcategory_count': row['subcategory_count'],
            'subcategories': subcats_display,
            'avg_price': round(row['avg_price'], 1),
            'avg_reviews': round(row['avg_reviews'], 1),
            'max_reviews': int(row['max_reviews']),
            'express_pct': round(row['express_ratio'] * 100, 1),
            'is_white_label': 'Yes' if row['is_white_label'] else 'No',
        })

    return pd.DataFrame(results)


def analyze_brand_landscape(df):
    """品牌分布分析 — 按子类目维度"""
    results = []
    subcats = df['_subcategory'].dropna().unique()

    for subcat in sorted(subcats):
        sub_df = df[df['_subcategory'] == subcat]
        n_products = len(sub_df)
        if n_products < 3:
            continue

        brand_counts = sub_df['brand'].value_counts()
        n_brands = len(brand_counts)
        n_white_label = sub_df['is_white_label'].sum()
        white_label_ratio = n_white_label / n_products * 100

        # Top 3 品牌
        top_brands = []
        for brand_name, count in brand_counts.head(3).items():
            share = count / n_products * 100
            top_brands.append(f"{brand_name} ({share:.0f}%)")
        top_brands_str = ' | '.join(top_brands)

        # 品牌多样性指数 (Shannon entropy, normalized)
        probs = brand_counts.values / n_products
        entropy = -sum(p * math.log2(p) for p in probs if p > 0)
        max_entropy = math.log2(n_brands) if n_brands > 1 else 1
        diversity_index = round(entropy / max_entropy * 100, 1) if max_entropy > 0 else 0

        # 中国卖家机会评分 (白牌率高 + 品牌分散 = 机会大)
        opportunity = round(white_label_ratio * 0.5 + diversity_index * 0.3 +
                            min(30, (100 - brand_counts.iloc[0] / n_products * 100) * 0.3), 1)

        results.append({
            'subcategory': subcat,
            'product_count': n_products,
            'brand_count': n_brands,
            'white_label_count': int(n_white_label),
            'white_label_pct': round(white_label_ratio, 1),
            'top_brands': top_brands_str,
            'diversity_index': diversity_index,
            'cn_seller_opportunity': round(opportunity, 1),
        })

    return pd.DataFrame(results)


# ============================================================
# Excel 输出
# ============================================================

def style_header(ws, n_cols):
    """给表头添加深蓝色样式"""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = DARK_BLUE_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=45):
    """自动列宽"""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            lengths.append(len(val))
        width = min(max(max(lengths) + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def write_sheet(wb, sheet_name, df, difficulty_col=None):
    """将 DataFrame 写入 workbook 的一个 sheet"""
    ws = wb.create_sheet(title=sheet_name)

    # 写表头
    for c, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=c, value=col_name)
    style_header(ws, len(df.columns))

    # 写数据
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=r, column=c, value=row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')

            # 百分比列自动格式
            if col_name.endswith('_pct'):
                cell.number_format = '0.0'
            elif col_name == 'hhi':
                cell.number_format = '#,##0'

    # Difficulty 列上色
    if difficulty_col and difficulty_col in df.columns:
        col_idx = list(df.columns).index(difficulty_col) + 1
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=col_idx)
            val = str(cell.value)
            if val in DIFFICULTY_COLORS:
                cell.fill = DIFFICULTY_COLORS[val]
                cell.font = DIFFICULTY_FONTS[val]
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 冻结首行
    ws.freeze_panes = 'A2'

    auto_width(ws)
    return ws


def generate_report(comp_df, seller_df, brand_df, output_path):
    """生成 Excel 报告"""
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    # Sheet 1: Subcategory Competition
    write_sheet(wb, 'Subcategory Competition', comp_df, difficulty_col='difficulty')

    # Sheet 2: Seller Profiles
    write_sheet(wb, 'Seller Profiles', seller_df)

    # Sheet 3: Brand Landscape
    write_sheet(wb, 'Brand Landscape', brand_df)

    wb.save(output_path)
    print(f"[OK] Excel 报告已保存: {output_path}")


# ============================================================
# Markdown 洞察报告
# ============================================================

def generate_insights(comp_df, seller_df, brand_df, output_path):
    """生成 Markdown 文本报告"""
    lines = []
    lines.append('# Noon Seller Competition Analysis Report')
    lines.append(f'\nGenerated: {datetime.now().strftime("%Y-%m-%d %H:%M")}\n')

    # --- 总览 ---
    lines.append('## 1. Overview\n')
    n_subcats = len(comp_df)
    n_brands_total = seller_df['brand'].nunique() if len(seller_df) > 0 else 0
    avg_white_label = comp_df['white_label_pct'].mean() if len(comp_df) > 0 else 0
    lines.append(f'- **Subcategories analyzed**: {n_subcats}')
    lines.append(f'- **Unique brands identified**: {n_brands_total}')
    lines.append(f'- **Average white-label ratio**: {avg_white_label:.1f}%')

    # 竞争格局分布
    if len(comp_df) > 0:
        comp_dist = comp_df['competition_type'].value_counts()
        lines.append(f'- **Competition landscape**:')
        for ctype, count in comp_dist.items():
            lines.append(f'  - {ctype}: {count} subcategories ({count/n_subcats*100:.0f}%)')

    # 进入难度分布
    if len(comp_df) > 0:
        diff_dist = comp_df['difficulty'].value_counts()
        lines.append(f'- **Entry difficulty**:')
        for d, count in diff_dist.items():
            lines.append(f'  - {d}: {count} subcategories ({count/n_subcats*100:.0f}%)')

    # --- 最佳机会 ---
    lines.append('\n## 2. Top Opportunities (Easy Entry)\n')
    easy = comp_df[comp_df['difficulty'] == 'Easy'].sort_values('white_label_pct', ascending=False)
    if len(easy) > 0:
        lines.append('| Subcategory | Products | Brands | White-label% | HHI | Avg Price |')
        lines.append('|---|---|---|---|---|---|')
        for _, row in easy.head(15).iterrows():
            lines.append(f"| {row['subcategory']} | {row['product_count']} | "
                         f"{row['brand_count']} | {row['white_label_pct']:.1f}% | "
                         f"{row['hhi']:.0f} | {row['avg_price']:.0f} |")
    else:
        lines.append('No subcategories classified as Easy entry.\n')

    # --- 高壁垒类目 ---
    lines.append('\n## 3. High Barrier Subcategories (Hard / Very Hard)\n')
    hard = comp_df[comp_df['difficulty'].isin(['Hard', 'Very Hard'])].sort_values(
        'difficulty_score', ascending=False)
    if len(hard) > 0:
        lines.append('| Subcategory | Products | CR1% | HHI | Express% | Difficulty |')
        lines.append('|---|---|---|---|---|---|')
        for _, row in hard.head(15).iterrows():
            lines.append(f"| {row['subcategory']} | {row['product_count']} | "
                         f"{row['cr1_pct']:.1f}% | {row['hhi']:.0f} | "
                         f"{row['express_pct']:.1f}% | {row['difficulty']} |")
    else:
        lines.append('No subcategories classified as Hard or Very Hard.\n')

    # --- Top 品牌 ---
    lines.append('\n## 4. Top Brands by Product Count\n')
    if len(seller_df) > 0:
        lines.append('| Brand | Products | Subcategories | Avg Price | Express% | White-label |')
        lines.append('|---|---|---|---|---|---|')
        for _, row in seller_df.head(20).iterrows():
            lines.append(f"| {row['brand']} | {row['product_count']} | "
                         f"{row['subcategory_count']} | {row['avg_price']:.0f} | "
                         f"{row['express_pct']:.1f}% | {row['is_white_label']} |")

    # --- 中国卖家机会 ---
    lines.append('\n## 5. China Seller Opportunity (High White-label + Fragmented)\n')
    if len(brand_df) > 0:
        opp = brand_df.sort_values('cn_seller_opportunity', ascending=False)
        lines.append('| Subcategory | Products | White-label% | Diversity | Opportunity Score |')
        lines.append('|---|---|---|---|---|')
        for _, row in opp.head(15).iterrows():
            lines.append(f"| {row['subcategory']} | {row['product_count']} | "
                         f"{row['white_label_pct']:.1f}% | {row['diversity_index']:.0f} | "
                         f"{row['cn_seller_opportunity']:.1f} |")

    # --- Key Takeaways ---
    lines.append('\n## 6. Key Takeaways\n')

    if len(comp_df) > 0:
        easy_pct = len(comp_df[comp_df['difficulty'] == 'Easy']) / len(comp_df) * 100
        frag_pct = len(comp_df[comp_df['competition_type'] == 'Fragmented']) / len(comp_df) * 100
        lines.append(f'1. **{easy_pct:.0f}% of subcategories** have Easy entry difficulty — '
                      'favorable for new sellers.')
        lines.append(f'2. **{frag_pct:.0f}% of subcategories** are Fragmented — '
                      'no dominant brand, lots of room for competition.')
        lines.append(f'3. **Average white-label ratio is {avg_white_label:.0f}%** — '
                      + ('high ratio means strong private-label opportunity for Chinese sellers.'
                         if avg_white_label > 50
                         else 'moderate ratio, focus on quality differentiation.'))
        if len(easy) > 0:
            best = easy.iloc[0]
            lines.append(f'4. **Best opportunity**: {best["subcategory"]} — '
                         f'{best["product_count"]} products, {best["white_label_pct"]:.0f}% white-label, '
                         f'HHI {best["hhi"]:.0f}.')

    text = '\n'.join(lines)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(text)
    print(f"[OK] 洞察报告已保存: {output_path}")


# ============================================================
# 主流程
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Noon 卖家竞争分析工具')
    parser.add_argument('--input', '-i', default='seller_data.xlsx',
                        help='输入数据文件 (默认: seller_data.xlsx)')
    parser.add_argument('--output', '-o', default='seller_report.xlsx',
                        help='输出报告文件 (默认: seller_report.xlsx)')
    args = parser.parse_args()

    # 1. 加载数据
    df = load_data(args.input)

    # 2. 预处理
    df = prepare_data(df)

    # 3. 分析
    print("[INFO] 正在分析子类目竞争格局...")
    comp_df = analyze_subcategory_competition(df)
    print(f"[INFO] 完成: {len(comp_df)} 个子类目")

    print("[INFO] 正在分析品牌/卖家画像...")
    seller_df = analyze_seller_profiles(df)
    print(f"[INFO] 完成: {len(seller_df)} 个品牌")

    print("[INFO] 正在分析品牌分布...")
    brand_df = analyze_brand_landscape(df)
    print(f"[INFO] 完成: {len(brand_df)} 个子类目")

    # 4. 输出
    output_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(output_dir, exist_ok=True)

    generate_report(comp_df, seller_df, brand_df, args.output)

    insights_path = os.path.join(output_dir, 'seller_insights.md')
    generate_insights(comp_df, seller_df, brand_df, insights_path)

    # 5. 摘要
    print("\n" + "=" * 60)
    print("  分析完成!")
    print("=" * 60)
    print(f"  子类目数:  {len(comp_df)}")
    print(f"  品牌数:    {len(seller_df)}")
    if len(comp_df) > 0:
        for d in ['Easy', 'Moderate', 'Hard', 'Very Hard']:
            cnt = len(comp_df[comp_df['difficulty'] == d])
            if cnt:
                print(f"  {d:10s}: {cnt} 子类目")
    print(f"\n  报告: {args.output}")
    print(f"  洞察: {insights_path}")
    print("=" * 60)


if __name__ == '__main__':
    main()
