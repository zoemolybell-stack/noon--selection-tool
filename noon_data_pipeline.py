#!/usr/bin/env python3
"""
Noon Automotive Data Cleaning & Enhancement Pipeline
=====================================================
Reads report_automotive.xlsx, cleans data, computes multi-dimensional
scores per subcategory, and outputs report_automotive_enhanced.xlsx
with 5 sheets: Enhanced Sub-Categories, Cleaned Products, Smart Top 100,
Category Opportunities, Risk Alerts.
"""

import sys
sys.path.insert(0, 'D:/claude noon/noon-selection-tool')

import json
import logging
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# ─── Paths ────────────────────────────────────────────────────────────────────
INPUT_PATH = Path(r'D:\claude noon\noon-selection-tool\data\monitoring\categories\automotive\report_automotive.xlsx')
OUTPUT_PATH = Path(r'D:\claude noon\report_automotive_enhanced.xlsx')

# ─── Style constants ─────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


# ═══════════════════════════════════════════════════════════════════════════════
# 1.1  clean_data
# ═══════════════════════════════════════════════════════════════════════════════
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Clean product-level data and add derived columns."""
    logger.info(f"Starting clean_data – input rows: {len(df)}")
    df = df.copy()

    # 1. Filter price < 5 SAR
    before = len(df)
    df = df[df['price'] >= 5].copy()
    logger.info(f"  Removed {before - len(df)} rows with price < 5 SAR")

    # 2. Filter margin_war_pct < -100
    before = len(df)
    df = df[df['margin_war_pct'] >= -100].copy()
    logger.info(f"  Removed {before - len(df)} rows with margin < -100%")

    # 3. Standardize sold_recently: convert to int, None→0
    def _parse_sold(val):
        try:
            if val is None or (isinstance(val, float) and np.isnan(val)):
                return 0
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    df['sold_recently'] = df['sold_recently'].apply(_parse_sold)

    # 4. Fill missing rating→0, review_count→0
    df['rating'] = df['rating'].fillna(0).astype(float)
    df['review_count'] = df['review_count'].fillna(0).astype(int)

    # 5. Create has_sales bool
    df['has_sales'] = df['sold_recently'] > 0

    # 6. Create price_band
    conditions = [
        df['price'] < 30,
        df['price'] < 80,
        df['price'] < 200,
        df['price'] < 500,
        df['price'] >= 500,
    ]
    labels = ['Budget(<30)', 'Low(30-80)', 'Mid(80-200)', 'High(200-500)', 'Premium(>500)']
    df['price_band'] = np.select(conditions, labels, default='Mid(80-200)')

    logger.info(f"  Output rows: {len(df)}")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# Helper: min-max normalise to 0-100
# ═══════════════════════════════════════════════════════════════════════════════
def _norm(series: pd.Series) -> pd.Series:
    """Min-max normalise a series to 0–100. Handles constant series."""
    mn, mx = series.min(), series.max()
    if mx == mn:
        return pd.Series(50.0, index=series.index)
    return ((series - mn) / (mx - mn) * 100).clip(0, 100)


# ═══════════════════════════════════════════════════════════════════════════════
# 1.2  enhance_features  (per-subcategory scoring)
# ═══════════════════════════════════════════════════════════════════════════════
def enhance_features(df: pd.DataFrame) -> pd.DataFrame:
    """Compute competition / demand / profit / opportunity / operability scores
    per subcategory and merge them back onto every product row."""
    logger.info("Starting enhance_features")

    scores_rows = []
    grouped = df.groupby('_subcategory', sort=False)

    for subcat, grp in grouped:
        try:
            n = len(grp)

            # ── competition_score ────────────────────────────────────────
            express_pct = grp['is_express'].sum() / n * 100
            has_reviews_pct = (grp['review_count'] > 0).sum() / n * 100
            avg_review_count = grp['review_count'].mean()
            ad_density = grp['is_ad'].sum() / n * 100
            bestseller_density = grp['is_bestseller'].sum() / n * 100

            # ── demand signals ──────────────────────────────────────────
            sold_sum = grp['sold_recently'].sum()
            avg_reviews = grp['review_count'].mean()
            product_count = n

            # ── profit signals ──────────────────────────────────────────
            median_margin = grp['margin_war_pct'].median()
            positive_margin_pct = (grp['margin_war_pct'] > 0).sum() / n * 100
            has_orig = grp['original_price'].notna().sum()
            price_stability = 1 - (has_orig / n) if n > 0 else 0.5

            # ── operability inputs ──────────────────────────────────────
            median_commission = grp['commission_pct'].median()
            median_price = grp['price'].median()
            # size_tier: map to numeric (smaller = better)
            size_map = {
                'Small Envelope': 5, 'Large Envelope': 4, 'Standard Parcel': 3,
                'Oversize': 2, 'Large Oversize': 1, 'Extra Oversize': 0, 'Bulky': 0,
            }
            size_num = grp['size_tier'].map(size_map).fillna(2).mean()

            scores_rows.append({
                '_subcategory': subcat,
                # raw intermediaries (used later for normalisation)
                '_express_pct': express_pct,
                '_has_reviews_pct': has_reviews_pct,
                '_avg_review_count': avg_review_count,
                '_ad_density': ad_density,
                '_bestseller_density': bestseller_density,
                '_sold_sum': sold_sum,
                '_avg_reviews': avg_reviews,
                '_product_count': product_count,
                '_median_margin': median_margin,
                '_positive_margin_pct': positive_margin_pct,
                '_price_stability': price_stability,
                '_median_commission': median_commission,
                '_median_price': median_price,
                '_size_num': size_num,
            })
        except Exception as e:
            logger.warning(f"  Skipping subcategory '{subcat}': {e}")

    sc = pd.DataFrame(scores_rows)

    # ── Normalise & compute composite scores ────────────────────────────────
    # competition_score
    sc['competition_score'] = (
        _norm(sc['_express_pct']) * 0.30
        + _norm(sc['_has_reviews_pct']) * 0.20
        + _norm(sc['_avg_review_count']) * 0.20
        + _norm(sc['_ad_density']) * 0.15
        + _norm(sc['_bestseller_density']) * 0.15
    ).round(2)

    # demand_score
    sc['demand_score'] = (
        _norm(sc['_sold_sum']) * 0.40
        + _norm(sc['_avg_reviews']) * 0.30
        + _norm(sc['_product_count']) * 0.30
    ).round(2)

    # profit_score
    sc['profit_score'] = (
        _norm(sc['_median_margin']) * 0.50
        + _norm(sc['_positive_margin_pct']) * 0.30
        + _norm(sc['_price_stability']) * 0.20
    ).round(2)

    # opportunity_score
    sc['opportunity_score'] = (
        sc['profit_score'] * 0.35
        + sc['demand_score'] * 0.35
        + (100 - sc['competition_score']) * 0.30
    ).round(2)

    # operability_score
    # commission: lower is better → invert
    comm_score = _norm(sc['_median_commission'].max() - sc['_median_commission'])
    # price band: mid-high (80-500) best → distance from 200 penalised
    price_band_score = _norm(100 - (sc['_median_price'] - 200).abs() / 5)
    # size: higher _size_num = smaller = better
    size_score = _norm(sc['_size_num'])
    sc['operability_score'] = (
        comm_score * 0.40 + price_band_score * 0.30 + size_score * 0.30
    ).round(2)

    # keep only what we need for merging
    score_cols = ['_subcategory', 'competition_score', 'demand_score',
                  'profit_score', 'opportunity_score', 'operability_score']
    merge_df = sc[score_cols]

    df = df.merge(merge_df, on='_subcategory', how='left')
    logger.info(f"  Scores computed for {len(merge_df)} subcategories")
    return df, sc  # return sc for subcategory sheet


# ═══════════════════════════════════════════════════════════════════════════════
# 1.3  enhance_subcategory_overview
# ═══════════════════════════════════════════════════════════════════════════════
def enhance_subcategory_overview(df: pd.DataFrame, subcat_df: pd.DataFrame,
                                  scores_df: pd.DataFrame) -> pd.DataFrame:
    """Augment the Sub-Category Overview sheet with sales, distribution, and scores."""
    logger.info("Starting enhance_subcategory_overview")
    subcat_df = subcat_df.copy()

    # Aggregate from cleaned products
    agg = df.groupby('_subcategory').agg(
        total_sold_recently=('sold_recently', 'sum'),
        products_with_sales=('has_sales', 'sum'),
        top_product_sales=('sold_recently', 'max'),
    ).reset_index()

    # Price band distribution as JSON string
    pb_dist = (
        df.groupby('_subcategory')['price_band']
        .value_counts()
        .unstack(fill_value=0)
    )
    pb_json = pb_dist.apply(lambda row: json.dumps(row.to_dict(), ensure_ascii=False), axis=1)
    pb_json = pb_json.reset_index()
    pb_json.columns = ['_subcategory', 'price_band_distribution']

    # Seller concentration heuristic: ratio of products whose title shares
    # the most common first word (brand proxy) — higher = more concentrated
    def _seller_concentration(grp):
        try:
            first_words = grp['title'].str.split().str[0].str.lower()
            if len(first_words) == 0:
                return 0.0
            top_freq = first_words.value_counts().iloc[0]
            return round(top_freq / len(first_words) * 100, 2)
        except Exception:
            return 0.0

    conc = df.groupby('_subcategory').apply(_seller_concentration, include_groups=False).reset_index()
    conc.columns = ['_subcategory', 'seller_concentration']

    # Merge everything onto subcat_df
    subcat_df = subcat_df.merge(agg, left_on='Subcategory', right_on='_subcategory', how='left')
    subcat_df.drop(columns=['_subcategory'], inplace=True, errors='ignore')
    subcat_df = subcat_df.merge(pb_json, left_on='Subcategory', right_on='_subcategory', how='left')
    subcat_df.drop(columns=['_subcategory'], inplace=True, errors='ignore')
    subcat_df = subcat_df.merge(conc, left_on='Subcategory', right_on='_subcategory', how='left')
    subcat_df.drop(columns=['_subcategory'], inplace=True, errors='ignore')

    # Merge scores
    score_cols = ['_subcategory', 'competition_score', 'demand_score',
                  'profit_score', 'opportunity_score', 'operability_score']
    subcat_df = subcat_df.merge(
        scores_df[score_cols],
        left_on='Subcategory', right_on='_subcategory', how='left'
    )
    subcat_df.drop(columns=['_subcategory'], inplace=True, errors='ignore')

    # Fill NaN scores with 0
    for col in ['total_sold_recently', 'products_with_sales', 'top_product_sales',
                'seller_concentration', 'competition_score', 'demand_score',
                'profit_score', 'opportunity_score', 'operability_score']:
        subcat_df[col] = subcat_df[col].fillna(0)

    # Final rank by opportunity_score descending
    subcat_df = subcat_df.sort_values('opportunity_score', ascending=False).reset_index(drop=True)
    subcat_df.insert(0, 'final_rank', range(1, len(subcat_df) + 1))

    logger.info(f"  Enhanced {len(subcat_df)} subcategories")
    return subcat_df


# ═══════════════════════════════════════════════════════════════════════════════
# Smart Top 100 — multi-dimensional ranking
# ═══════════════════════════════════════════════════════════════════════════════
def build_smart_top100(df: pd.DataFrame) -> pd.DataFrame:
    """Rank products by a composite score (not just margin)."""
    logger.info("Building Smart Top 100")
    t = df.copy()

    # Composite product score
    margin_n = _norm(t['margin_war_pct'])
    sales_n = _norm(t['sold_recently'])
    review_n = _norm(t['review_count'])
    opp_n = t['opportunity_score'].fillna(50)  # subcategory opportunity

    t['smart_score'] = (
        margin_n * 0.30
        + sales_n * 0.25
        + review_n * 0.15
        + opp_n * 0.30
    ).round(2)

    t = t.sort_values('smart_score', ascending=False).head(100).reset_index(drop=True)
    t.insert(0, 'rank', range(1, len(t) + 1))

    cols = ['rank', 'product_id', 'title', '_subcategory', 'price', 'price_band',
            'margin_war_pct', 'margin_peace_pct', 'net_profit_war_cny',
            'sold_recently', 'review_count', 'rating', 'is_express', 'is_bestseller',
            'smart_score', 'opportunity_score', 'competition_score', 'demand_score',
            'profit_score', 'operability_score', 'product_url']
    return t[[c for c in cols if c in t.columns]]


# ═══════════════════════════════════════════════════════════════════════════════
# Category Opportunities — high opp + low competition + demand verified
# ═══════════════════════════════════════════════════════════════════════════════
def build_category_opportunities(subcat_enhanced: pd.DataFrame) -> pd.DataFrame:
    """Subcategories with high opportunity, low competition, and verified demand."""
    logger.info("Building Category Opportunities")
    c = subcat_enhanced.copy()

    # Thresholds: opportunity top 40%, competition bottom 50%, demand > 0 evidence
    opp_thresh = c['opportunity_score'].quantile(0.60)
    comp_thresh = c['competition_score'].quantile(0.50)

    mask = (
        (c['opportunity_score'] >= opp_thresh)
        & (c['competition_score'] <= comp_thresh)
        & (c['total_sold_recently'] > 0)
    )
    result = c[mask].sort_values('opportunity_score', ascending=False).reset_index(drop=True)
    logger.info(f"  Found {len(result)} opportunity subcategories")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Risk Alerts
# ═══════════════════════════════════════════════════════════════════════════════
def build_risk_alerts(df: pd.DataFrame) -> pd.DataFrame:
    """Flag products with negative margin, ultra-high competition subcategories,
    or oversized items."""
    logger.info("Building Risk Alerts")

    alerts = []

    for idx, row in df.iterrows():
        try:
            reasons = []
            if row.get('margin_war_pct', 0) < 0:
                reasons.append(f"Negative margin ({row['margin_war_pct']:.1f}%)")
            if row.get('competition_score', 0) >= 80:
                reasons.append(f"Ultra-high competition ({row['competition_score']:.0f})")
            if row.get('size_tier', '') in ('Extra Oversize', 'Bulky', 'Large Oversize', 'Oversize'):
                reasons.append(f"Oversized ({row['size_tier']})")

            if reasons:
                alerts.append({
                    'product_id': row['product_id'],
                    'title': row['title'],
                    '_subcategory': row['_subcategory'],
                    'price': row['price'],
                    'margin_war_pct': row['margin_war_pct'],
                    'size_tier': row['size_tier'],
                    'competition_score': row.get('competition_score', None),
                    'risk_reasons': ' | '.join(reasons),
                    'product_url': row.get('product_url', ''),
                })
        except Exception as e:
            logger.debug(f"  Skipping row {idx}: {e}")

    result = pd.DataFrame(alerts)
    if len(result) > 0:
        result = result.sort_values('margin_war_pct', ascending=True).reset_index(drop=True)
    logger.info(f"  Found {len(result)} risk alerts")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Excel writer with formatting
# ═══════════════════════════════════════════════════════════════════════════════
def _style_sheet(ws, df_columns: list, margin_cols: list = None,
                 score_cols: list = None):
    """Apply formatting to a worksheet."""
    # Header row styling
    for col_idx in range(1, len(df_columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-width (capped)
    for col_idx, col_name in enumerate(df_columns, 1):
        max_len = len(str(col_name)) + 2
        # sample a few data rows for width
        for row_idx in range(2, min(52, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Conditional formatting: green/red for margin columns
    if margin_cols:
        for col_name in margin_cols:
            if col_name in df_columns:
                col_letter = get_column_letter(df_columns.index(col_name) + 1)
                rng = f'{col_letter}2:{col_letter}{ws.max_row}'
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator='greaterThan', formula=['0'], fill=GREEN_FILL))
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator='lessThan', formula=['0'], fill=RED_FILL))

    # Color scale for score columns (red low → green high)
    if score_cols:
        for col_name in score_cols:
            if col_name in df_columns:
                col_letter = get_column_letter(df_columns.index(col_name) + 1)
                rng = f'{col_letter}2:{col_letter}{ws.max_row}'
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type='min', start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B',
                ))


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame,
                 margin_cols: list = None, score_cols: list = None):
    """Write a DataFrame to a named sheet with formatting."""
    ws = wb.create_sheet(title=sheet_name)
    columns = list(df.columns)

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(columns, 1):
            try:
                val = row[col_name]
                # Convert numpy types to native Python
                if isinstance(val, (np.integer,)):
                    val = int(val)
                elif isinstance(val, (np.floating,)):
                    val = float(val)
                elif isinstance(val, (np.bool_,)):
                    val = bool(val)
                elif pd.isna(val):
                    val = None
                ws.cell(row=row_idx, column=col_idx, value=val)
            except Exception:
                ws.cell(row=row_idx, column=col_idx, value=str(row[col_name]))

    _style_sheet(ws, columns, margin_cols=margin_cols, score_cols=score_cols)
    logger.info(f"  Sheet '{sheet_name}': {len(df)} rows × {len(columns)} cols")


def write_output(subcat_enhanced: pd.DataFrame, cleaned_products: pd.DataFrame,
                 smart_top100: pd.DataFrame, category_opps: pd.DataFrame,
                 risk_alerts: pd.DataFrame, output_path: Path):
    """Write all 5 sheets to the output xlsx."""
    logger.info(f"Writing output to {output_path}")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    score_col_names = ['competition_score', 'demand_score', 'profit_score',
                       'opportunity_score', 'operability_score', 'smart_score']
    margin_col_names = ['margin_war_pct', 'margin_peace_pct', 'net_profit_war_cny',
                        'Median Margin %']

    _write_sheet(wb, 'Enhanced Sub-Categories', subcat_enhanced,
                 margin_cols=['Median Margin %'], score_cols=score_col_names)

    # For cleaned products, limit columns to keep file manageable
    product_cols = [
        'product_id', 'title', '_subcategory', 'price', 'price_band',
        'original_price', 'rating', 'review_count', 'is_express', 'is_bestseller',
        'is_ad', 'image_count', 'sold_recently', 'has_sales',
        'v6_category', 'size_tier', 'purchase_est_cny',
        'margin_war_pct', 'margin_peace_pct', 'net_profit_war_cny',
        'fbn_fee_sar', 'commission_pct',
        'competition_score', 'demand_score', 'profit_score',
        'opportunity_score', 'operability_score',
        'product_url',
    ]
    existing_cols = [c for c in product_cols if c in cleaned_products.columns]
    _write_sheet(wb, 'Cleaned Products', cleaned_products[existing_cols],
                 margin_cols=margin_col_names, score_cols=score_col_names)

    _write_sheet(wb, 'Smart Top 100', smart_top100,
                 margin_cols=margin_col_names, score_cols=score_col_names)

    _write_sheet(wb, 'Category Opportunities', category_opps,
                 margin_cols=['Median Margin %'], score_cols=score_col_names)

    _write_sheet(wb, 'Risk Alerts', risk_alerts,
                 margin_cols=margin_col_names, score_cols=score_col_names)

    wb.save(str(output_path))
    logger.info(f"Done — saved to {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    logger.info(f"Reading {INPUT_PATH}")

    # Read sheets
    products_raw = pd.read_excel(INPUT_PATH, sheet_name='All Products')
    subcat_raw = pd.read_excel(INPUT_PATH, sheet_name='Sub-Category Overview')
    logger.info(f"  All Products: {len(products_raw)} rows")
    logger.info(f"  Sub-Category Overview: {len(subcat_raw)} rows")

    # 1.1 Clean
    products_clean = clean_data(products_raw)

    # 1.2 Enhance features (scores per subcategory)
    products_enhanced, scores_df = enhance_features(products_clean)

    # 1.3 Enhance subcategory overview
    subcat_enhanced = enhance_subcategory_overview(products_enhanced, subcat_raw, scores_df)

    # Build derived sheets
    smart_top100 = build_smart_top100(products_enhanced)
    category_opps = build_category_opportunities(subcat_enhanced)
    risk_alerts = build_risk_alerts(products_enhanced)

    # 1.4 Write output
    write_output(
        subcat_enhanced=subcat_enhanced,
        cleaned_products=products_enhanced,
        smart_top100=smart_top100,
        category_opps=category_opps,
        risk_alerts=risk_alerts,
        output_path=OUTPUT_PATH,
    )


if __name__ == '__main__':
    main()
