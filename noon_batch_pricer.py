#!/usr/bin/env python3
"""
noon_batch_pricer.py — Noon 批量反向定价计算器

根据目标利润率反推售价，输出含定价建议、成本拆解、敏感度分析的 Excel 报告。

Usage:
    python noon_batch_pricer.py --input products_to_price.csv --target-margin 0.20 --mode FBN --country KSA --output pricing_plan.xlsx
"""

import argparse
import sys
import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, 'D:/claude noon')
from noon_config import (
    CONFIG, get_commission, lookup_fbn_fee, lookup_fbp_fee,
    get_fbp_reimbursement, calc_chargeable_weight, calc_full_profit,
)


# ============================================================
# Styling constants
# ============================================================
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(color="9C6500")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")


# ============================================================
# Helpers
# ============================================================

def calc_head_cost(row, country, mode):
    """Calculate head freight cost in CNY based on mode."""
    cfg = CONFIG[country]
    if mode in ('FBN', 'FBP_OVERSEAS'):
        # Sea freight CBM-based
        vol_m3 = (row['length_cm'] * row['width_cm'] * row['height_cm']) / 1_000_000
        return vol_m3 * cfg['fbn_head_rate']
    else:
        # FBP direct mail — per kg
        return row['weight_kg'] * cfg['fbp_head_rate']


def find_target_price(target_margin, row, country, mode,
                      lo=1.0, hi=5000.0, precision=0.5):
    """
    Binary search to find the selling price that achieves target_margin.
    Precision: 0.5 SAR/AED.
    """
    cw = calc_chargeable_weight(
        row['length_cm'], row['width_cm'], row['height_cm'], row['weight_kg']
    )

    calc_mode = 'FBN' if mode == 'FBN' else 'FBP'

    for _ in range(200):  # enough iterations for convergence
        mid = round((lo + hi) / 2 * 2) / 2  # snap to 0.5
        head_cny = calc_head_cost(row, country, mode)
        result = calc_full_profit(
            price=mid,
            purchase_cny=row['purchase_cny'],
            chargeable_weight=cw,
            v6_category=row.get('v6_category'),
            commission_pct=row.get('commission_pct'),
            country=country,
            head_cost_cny=head_cny,
            mode=calc_mode,
        )
        current_margin = result['margin']

        if abs(current_margin - target_margin) < 0.001:
            return mid, result
        if current_margin < target_margin:
            lo = mid + precision
        else:
            hi = mid - precision

        if lo > hi:
            break

    # Return best found
    head_cny = calc_head_cost(row, country, mode)
    result = calc_full_profit(
        price=mid,
        purchase_cny=row['purchase_cny'],
        chargeable_weight=cw,
        v6_category=row.get('v6_category'),
        commission_pct=row.get('commission_pct'),
        country=country,
        head_cost_cny=head_cny,
        mode=calc_mode,
    )
    return mid, result


def get_risk_flags(price, competitor_price, country):
    """Generate risk flags comparing recommended price to competitor."""
    flags = []
    if pd.notna(competitor_price) and competitor_price > 0:
        ratio = price / competitor_price
        if ratio > 1.3:
            flags.append(f"ABOVE_COMPETITOR ({ratio:.0%})")
        elif ratio < 0.7:
            flags.append(f"BELOW_COMPETITOR ({ratio:.0%})")

    # Free shipping thresholds
    currency = CONFIG[country]['currency']
    if country == 'KSA' and price < 100:
        flags.append(f"Below free-ship 100 {currency}")
    elif country == 'UAE' and price < 100:
        flags.append(f"Below free-ship 100 {currency}")

    return "; ".join(flags) if flags else "OK"


# ============================================================
# Excel writing helpers
# ============================================================

def style_header_row(ws, num_cols, row=1):
    """Apply dark blue header style to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=30):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def add_margin_conditional_formatting(ws, col_letter, min_row, max_row):
    """Add green/yellow/red conditional formatting for margin columns."""
    rng = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.20"],
                        fill=GREEN_FILL, font=GREEN_FONT)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="between", formula=["0.10", "0.1999"],
                        fill=YELLOW_FILL, font=YELLOW_FONT)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0.10"],
                        fill=RED_FILL, font=RED_FONT)
    )


def fmt_pct(val):
    """Format a float as percentage string."""
    return f"{val * 100:.1f}%"


# ============================================================
# Sheet builders
# ============================================================

def build_pricing_recommendations(wb, products, scenarios, country):
    """Sheet 1: Pricing Recommendations."""
    ws = wb.active
    ws.title = "Pricing Recommendations"

    currency = CONFIG[country]['currency']
    headers = [
        "SKU", "Product Name",
        f"Purchase (CNY)",
        f"Conservative Price ({currency})", "Conservative Margin",
        f"Standard Price ({currency})", "Standard Margin",
        f"Aggressive Price ({currency})", "Aggressive Margin",
        f"Competitor Price ({currency})", "vs Competitor (Std)",
        "Risk Flags",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for i, row in products.iterrows():
        cons = scenarios[i]['conservative']
        std = scenarios[i]['standard']
        aggr = scenarios[i]['aggressive']
        competitor = row.get('competitor_price_sar', None)
        if pd.isna(competitor):
            competitor = None

        vs_comp = ""
        if competitor and competitor > 0:
            vs_comp = f"{std['price'] / competitor:.0%}"

        ws.append([
            row['sku'], row['product_name'],
            round(row['purchase_cny'], 2),
            cons['price'], cons['result']['margin'],
            std['price'], std['result']['margin'],
            aggr['price'], aggr['result']['margin'],
            competitor if competitor else "",
            vs_comp,
            scenarios[i]['risk_flags'],
        ])

    # Format margin columns as percentage
    for r in range(2, len(products) + 2):
        for c in [5, 7, 9]:  # margin columns
            cell = ws.cell(row=r, column=c)
            cell.number_format = '0.0%'

    # Conditional formatting on margin columns
    max_row = len(products) + 1
    for col_letter in ['E', 'G', 'I']:
        add_margin_conditional_formatting(ws, col_letter, 2, max_row)

    # Borders for data area
    for r in range(1, max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.freeze_panes = "A2"
    auto_width(ws)


def build_cost_breakdown(wb, products, scenarios, country, mode_label, sheet_name):
    """Sheet 2/3: FBN or FBP Cost Breakdown for standard scenario."""
    ws = wb.create_sheet(title=sheet_name)
    currency = CONFIG[country]['currency']

    headers = [
        "SKU", "Product Name",
        f"Selling Price ({currency})",
        f"VAT ({currency})", f"Commission ({currency})", "Commission Rate",
        f"Commission VAT ({currency})", f"Fulfillment Fee ({currency})",
        f"Reimbursement ({currency})",
        f"Ads Fee ({currency})", f"Return Admin ({currency})",
        f"Logistics Loss ({currency})",
        f"Head Freight (CNY)", f"Total Cost (CNY)", f"Revenue (CNY)",
        f"Net Profit (CNY)", "Margin", "ROI",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for i, row in products.iterrows():
        r = scenarios[i]['standard']['result']
        ws.append([
            row['sku'], row['product_name'],
            r['price'], r['vat'], r['commission'], r['commission_rate'],
            r['commission_vat'], r['fulfillment_fee'], r['reimbursement'],
            r['ads_fee'], r['return_admin'], r['log_loss'],
            r['head_cost_cny'], r['total_cost_cny'], r['revenue_cny'],
            r['net_profit_cny'], r['margin'], r['roi'],
        ])

    # Format percentage columns
    max_row = len(products) + 1
    for r_idx in range(2, max_row + 1):
        ws.cell(row=r_idx, column=6).number_format = '0.0%'   # commission rate
        ws.cell(row=r_idx, column=17).number_format = '0.0%'  # margin
        ws.cell(row=r_idx, column=18).number_format = '0.0%'  # ROI

    add_margin_conditional_formatting(ws, 'Q', 2, max_row)

    for r_idx in range(1, max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c).border = THIN_BORDER

    ws.freeze_panes = "A2"
    auto_width(ws)


def build_sensitivity_analysis(wb, products, country, mode):
    """Sheet 4: Sensitivity Analysis — price +-20% in 5% steps."""
    ws = wb.create_sheet(title="Sensitivity Analysis")

    steps = [-20, -15, -10, -5, 0, 5, 10, 15, 20]
    currency = CONFIG[country]['currency']

    headers = ["SKU", "Product Name", f"Base Price ({currency})"]
    for s in steps:
        sign = "+" if s >= 0 else ""
        headers.append(f"{sign}{s}% Price")
        headers.append(f"{sign}{s}% Margin")
    ws.append(headers)
    style_header_row(ws, len(headers))

    calc_mode = 'FBN' if mode == 'FBN' else 'FBP'

    for i, row in products.iterrows():
        base_price = products.loc[i, '_std_price']
        cw = calc_chargeable_weight(
            row['length_cm'], row['width_cm'], row['height_cm'], row['weight_kg']
        )
        head_cny = calc_head_cost(row, country, mode)

        data = [row['sku'], row['product_name'], base_price]
        for s in steps:
            p = round(base_price * (1 + s / 100) * 2) / 2  # snap to 0.5
            if p < 1:
                p = 1.0
            result = calc_full_profit(
                price=p,
                purchase_cny=row['purchase_cny'],
                chargeable_weight=cw,
                v6_category=row.get('v6_category'),
                commission_pct=row.get('commission_pct'),
                country=country,
                head_cost_cny=head_cny,
                mode=calc_mode,
            )
            data.append(round(p, 1))
            data.append(result['margin'])

        ws.append(data)

    max_row = len(products) + 1

    # Format margin columns (every other column starting from col 5)
    margin_cols = []
    for idx, s in enumerate(steps):
        price_col = 4 + idx * 2
        margin_col = 5 + idx * 2
        margin_cols.append(margin_col)
        col_letter = get_column_letter(margin_col)
        for r_idx in range(2, max_row + 1):
            ws.cell(row=r_idx, column=margin_col).number_format = '0.0%'
        add_margin_conditional_formatting(ws, col_letter, 2, max_row)

    for r_idx in range(1, max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c).border = THIN_BORDER

    ws.freeze_panes = "A2"
    auto_width(ws)


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Noon 批量反向定价计算器 — 根据目标利润率反推售价"
    )
    parser.add_argument("--input", required=True, help="Input CSV file path")
    parser.add_argument("--target-margin", type=float, default=0.20,
                        help="Target profit margin (0-1), default 0.20")
    parser.add_argument("--mode", default="FBN", choices=["FBN", "FBP"],
                        help="Fulfillment mode: FBN or FBP (default: FBN)")
    parser.add_argument("--country", default="KSA", choices=["KSA", "UAE"],
                        help="Market: KSA or UAE (default: KSA)")
    parser.add_argument("--output", default="pricing_plan.xlsx",
                        help="Output Excel file path (default: pricing_plan.xlsx)")
    args = parser.parse_args()

    target = args.target_margin
    mode = args.mode
    country = args.country

    # Read input CSV
    print(f"Reading input: {args.input}")
    df = pd.read_csv(args.input)
    required_cols = ['sku', 'product_name', 'purchase_cny', 'weight_kg',
                     'length_cm', 'width_cm', 'height_cm']
    for col in required_cols:
        if col not in df.columns:
            print(f"ERROR: Missing required column '{col}' in input CSV.")
            sys.exit(1)

    # Fill optional columns
    if 'v6_category' not in df.columns:
        df['v6_category'] = '其他通用 General/Other'
    if 'commission_pct' not in df.columns:
        df['commission_pct'] = None
    if 'competitor_price_sar' not in df.columns:
        df['competitor_price_sar'] = None

    currency = CONFIG[country]['currency']
    print(f"Market: {country} ({currency}), Mode: {mode}, Target margin: {target:.0%}")
    print(f"Products: {len(df)}")
    print()

    # Three scenarios
    conservative_target = target + 0.05
    standard_target = target
    aggressive_target = target - 0.05

    scenarios = {}

    for i, row in df.iterrows():
        print(f"  [{i+1}/{len(df)}] {row['sku']} — {row['product_name']}")

        # Conservative
        c_price, c_result = find_target_price(conservative_target, row, country, mode)
        # Standard
        s_price, s_result = find_target_price(standard_target, row, country, mode)
        # Aggressive
        a_price, a_result = find_target_price(aggressive_target, row, country, mode)

        # Also compute the other mode for cost breakdown
        other_mode = 'FBP' if mode == 'FBN' else 'FBN'
        other_calc_mode = 'FBP' if mode == 'FBN' else 'FBN'
        cw = calc_chargeable_weight(
            row['length_cm'], row['width_cm'], row['height_cm'], row['weight_kg']
        )
        head_cny = calc_head_cost(row, country, mode)
        other_head_cny = calc_head_cost(row, country, other_mode)
        other_result = calc_full_profit(
            price=s_price,
            purchase_cny=row['purchase_cny'],
            chargeable_weight=cw,
            v6_category=row.get('v6_category'),
            commission_pct=row.get('commission_pct'),
            country=country,
            head_cost_cny=other_head_cny,
            mode=other_calc_mode,
        )

        competitor = row.get('competitor_price_sar', None)
        if pd.isna(competitor):
            competitor = None
        risk = get_risk_flags(s_price, competitor, country)

        scenarios[i] = {
            'conservative': {'price': c_price, 'result': c_result},
            'standard': {'price': s_price, 'result': s_result},
            'aggressive': {'price': a_price, 'result': a_result},
            'other_result': other_result,
            'risk_flags': risk,
        }

        # Store standard price for sensitivity analysis
        df.at[i, '_std_price'] = s_price

        print(f"    Conservative: {c_price:.1f} {currency} ({fmt_pct(c_result['margin'])})"
              f"  Standard: {s_price:.1f} {currency} ({fmt_pct(s_result['margin'])})"
              f"  Aggressive: {a_price:.1f} {currency} ({fmt_pct(a_result['margin'])})")
        if risk != "OK":
            print(f"    WARNING: {risk}")

    # Build Excel
    print(f"\nGenerating {args.output} ...")
    wb = Workbook()

    # Sheet 1: Pricing Recommendations
    build_pricing_recommendations(wb, df, scenarios, country)

    # Sheet 2: FBN Cost Breakdown (use primary mode results for standard scenario)
    # We need to generate FBN results at the standard price for each product
    fbn_scenarios = {}
    for i, row in df.iterrows():
        s_price = scenarios[i]['standard']['price']
        cw = calc_chargeable_weight(
            row['length_cm'], row['width_cm'], row['height_cm'], row['weight_kg']
        )
        head_cny_fbn = calc_head_cost(row, country, 'FBN')
        fbn_result = calc_full_profit(
            price=s_price,
            purchase_cny=row['purchase_cny'],
            chargeable_weight=cw,
            v6_category=row.get('v6_category'),
            commission_pct=row.get('commission_pct'),
            country=country,
            head_cost_cny=head_cny_fbn,
            mode='FBN',
        )
        fbn_scenarios[i] = {'standard': {'result': fbn_result}}

    build_cost_breakdown(wb, df, fbn_scenarios, country, "FBN", "FBN Cost Breakdown")

    # Sheet 3: FBP Cost Breakdown
    fbp_scenarios = {}
    for i, row in df.iterrows():
        s_price = scenarios[i]['standard']['price']
        cw = calc_chargeable_weight(
            row['length_cm'], row['width_cm'], row['height_cm'], row['weight_kg']
        )
        head_cny_fbp = calc_head_cost(row, country, 'FBP_OVERSEAS')
        fbp_result = calc_full_profit(
            price=s_price,
            purchase_cny=row['purchase_cny'],
            chargeable_weight=cw,
            v6_category=row.get('v6_category'),
            commission_pct=row.get('commission_pct'),
            country=country,
            head_cost_cny=head_cny_fbp,
            mode='FBP',
        )
        fbp_scenarios[i] = {'standard': {'result': fbp_result}}

    build_cost_breakdown(wb, df, fbp_scenarios, country, "FBP", "FBP Cost Breakdown")

    # Sheet 4: Sensitivity Analysis
    build_sensitivity_analysis(wb, df, country, mode)

    wb.save(args.output)
    print(f"Done! Output saved to: {args.output}")
    print(f"\nSheets: Pricing Recommendations | FBN Cost Breakdown | FBP Cost Breakdown | Sensitivity Analysis")


if __name__ == '__main__':
    main()
