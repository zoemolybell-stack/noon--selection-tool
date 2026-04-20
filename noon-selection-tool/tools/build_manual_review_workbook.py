import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ZERO_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
LOW_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
TRUNCATED_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
ERROR_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


ISSUE_PRIORITY = {
    "all_deduped": 1,
    "zero_unique": 2,
    "low_unique": 3,
    "runtime_error": 4,
    "truncated_multi_source": 5,
}


ISSUE_LABELS = {
    "all_deduped": "有原始结果但去重后为 0",
    "zero_unique": "结果为 0",
    "low_unique": "结果偏低",
    "runtime_error": "运行异常/超时",
    "truncated_multi_source": "组合叶子超出 100 后被截断",
}


ISSUE_FILLS = {
    "all_deduped": ZERO_FILL,
    "zero_unique": ZERO_FILL,
    "low_unique": LOW_FILL,
    "runtime_error": ERROR_FILL,
    "truncated_multi_source": TRUNCATED_FILL,
}


def autosize_columns(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 48)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def normalize_subcategory_name(path: Path) -> str:
    stem = path.stem.replace("_", " ").strip()
    return stem


def issue_rows_from_subcategory(category: str, sub_path: Path, export_path: str | None):
    data = json.loads(sub_path.read_text(encoding="utf-8"))
    subcategory = data.get("subcategory") or normalize_subcategory_name(sub_path)
    product_count = data.get("product_count", 0) or 0
    raw_count = data.get("raw_count", 0) or 0
    deduped_full_count = data.get("deduped_full_count")
    truncated_count = data.get("truncated_count", 0) or 0
    issues = []

    if raw_count > 0 and deduped_full_count == 0:
        issues.append(("all_deduped", "页面有抓取结果，但去重后全部被相邻叶子吃掉，建议核查边界是否重复。"))
    elif product_count == 0:
        issues.append(("zero_unique", "当前叶子最终没有保留任何商品，建议核查 URL 是否空页或解析器是否失效。"))

    if 0 < product_count < 30:
        issues.append(("low_unique", "当前叶子商品量偏低，建议核查 Noon 前台是否本身货少，或 URL 是否过窄。"))

    if truncated_count > 0:
        issues.append(("truncated_multi_source", "该叶子来自多源页，去重后仍超过 100，已按口径截断到前 100，建议抽检相关性。"))

    rows = []
    for issue_type, reason in issues:
        rows.append(
            {
                "issue_type": issue_type,
                "category": category,
                "subcategory": subcategory,
                "product_count": product_count,
                "raw_count": raw_count,
                "deduped_full_count": deduped_full_count,
                "truncated_count": truncated_count,
                "reason": reason,
                "url": data.get("url", ""),
                "source_urls": ", ".join(data.get("source_urls", [])) if isinstance(data.get("source_urls"), list) else "",
                "subcategory_json": str(sub_path),
                "category_excel": export_path or "",
                "products": data.get("products", []),
            }
        )
    return rows


def runtime_issue_rows(summary: dict):
    rows = []
    for item in summary.get("results", []):
        export_path = item.get("excel_path")
        for error in item.get("result", {}).get("errors", []):
            rows.append(
                {
                    "issue_type": "runtime_error",
                    "category": item.get("category"),
                    "subcategory": error.get("subcategory", ""),
                    "product_count": None,
                    "raw_count": None,
                    "deduped_full_count": None,
                    "truncated_count": None,
                    "reason": error.get("error", "").strip(),
                    "url": "",
                    "source_urls": "",
                    "subcategory_json": "",
                    "category_excel": export_path or "",
                    "products": [],
                }
            )
    return rows


def build_review_rows(batch_dir: Path):
    summary_path = batch_dir / "batch_scan_summary.json"
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    export_map = {item["category"]: item.get("excel_path") for item in summary.get("results", [])}
    rows = []

    monitoring_dir = batch_dir / "monitoring" / "categories"
    for category_dir in sorted(monitoring_dir.iterdir()):
        if not category_dir.is_dir():
            continue
        category = category_dir.name
        for sub_path in sorted((category_dir / "subcategory").glob("*.json")):
            rows.extend(issue_rows_from_subcategory(category, sub_path, export_map.get(category)))

    rows.extend(runtime_issue_rows(summary))
    rows.sort(key=lambda row: (ISSUE_PRIORITY[row["issue_type"]], row["category"], row["subcategory"]))
    return rows


def write_readme(ws):
    ws.append(["颜色", "含义", "说明"])
    style_header(ws)
    legend = [
        ("红色", ISSUE_LABELS["all_deduped"], "有原始抓取但最终为 0，通常说明叶子边界重复或页不干净。", ZERO_FILL),
        ("红色", ISSUE_LABELS["zero_unique"], "最终结果为 0，通常说明空页、解析失效或 URL 不稳定。", ZERO_FILL),
        ("橙色", ISSUE_LABELS["low_unique"], "结果偏低，建议确认是不是 Noon 前台本身样本很少。", LOW_FILL),
        ("黄色", ISSUE_LABELS["runtime_error"], "本轮有超时或运行异常，建议复跑或核查该页。", ERROR_FILL),
        ("蓝色", ISSUE_LABELS["truncated_multi_source"], "组合叶子来自多源页，已截断到前 100，建议抽检相关性。", TRUNCATED_FILL),
    ]
    for color_name, label, note, fill in legend:
        ws.append([color_name, label, note])
        for cell in ws[ws.max_row]:
            cell.fill = fill
    autosize_columns(ws)


def write_summary(ws, review_rows):
    headers = [
        "priority",
        "issue_type",
        "issue_label",
        "category",
        "subcategory",
        "product_count",
        "raw_count",
        "deduped_full_count",
        "truncated_count",
        "reason",
        "url",
        "source_urls",
        "category_excel",
        "subcategory_json",
    ]
    ws.append(headers)
    style_header(ws)

    for row in review_rows:
        values = [
            ISSUE_PRIORITY[row["issue_type"]],
            row["issue_type"],
            ISSUE_LABELS[row["issue_type"]],
            row["category"],
            row["subcategory"],
            row["product_count"],
            row["raw_count"],
            row["deduped_full_count"],
            row["truncated_count"],
            row["reason"],
            row["url"],
            row["source_urls"],
            row["category_excel"],
            row["subcategory_json"],
        ]
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.fill = ISSUE_FILLS[row["issue_type"]]
    autosize_columns(ws)


def write_products(ws, review_rows):
    headers = [
        "issue_label",
        "category",
        "subcategory",
        "search_rank",
        "product_id",
        "title",
        "price",
        "rating",
        "review_count",
        "delivery_type",
        "is_ad",
        "sold_recently_text",
        "ranking_signal_text",
        "product_url",
    ]
    ws.append(headers)
    style_header(ws)

    for row in review_rows:
        if not row["products"]:
            continue
        for product in row["products"]:
            ws.append(
                [
                    ISSUE_LABELS[row["issue_type"]],
                    row["category"],
                    row["subcategory"],
                    product.get("search_rank"),
                    product.get("product_id"),
                    product.get("title"),
                    product.get("price"),
                    product.get("rating"),
                    product.get("review_count"),
                    product.get("delivery_type"),
                    product.get("is_ad"),
                    product.get("sold_recently_text"),
                    product.get("ranking_signal_text"),
                    product.get("product_url"),
                ]
            )
            for cell in ws[ws.max_row]:
                cell.fill = ISSUE_FILLS[row["issue_type"]]
    autosize_columns(ws)


def parse_args():
    parser = argparse.ArgumentParser(description="生成人工复核标注 Excel")
    parser.add_argument("--batch-dir", type=Path, required=True, help="批量扫描输出目录")
    parser.add_argument("--output", type=Path, default=None, help="输出 Excel 路径")
    return parser.parse_args()


def main():
    args = parse_args()
    batch_dir = args.batch_dir
    output_path = args.output or (batch_dir / "manual_review_flags.xlsx")
    review_rows = build_review_rows(batch_dir)

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "Readme"
    write_readme(ws_readme)

    ws_summary = wb.create_sheet("Review_Summary")
    write_summary(ws_summary, review_rows)

    ws_products = wb.create_sheet("Review_Products")
    write_products(ws_products, review_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
