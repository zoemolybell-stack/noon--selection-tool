"""
Automotive 类目完整爬取 + 报告生成

步骤:
1. 爬取完整类目树
2. 每个子类目最多取前500条商品
3. 全局 product_id 去重
4. 对全部商品运行V6利润精算 + 产品级分析
5. 输出报告 report_automotive.xlsx
"""
import asyncio
import json
import logging
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from scrapers.noon_category_crawler import NoonCategoryCrawler
from analysis.pricing_engine import calculate_ksa_fbn_profit
from config.category_mapping import map_category
from config.cost_defaults import get_defaults
from config.settings import Settings

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("category_crawl.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("category-crawl")

# Excel 样式
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def calc_product_profit(product: dict, settings: Settings) -> dict:
    """对单个商品调用 V6 引擎计算利润"""
    price = product.get("price", 0)
    if not price or price <= 0:
        return {
            "purchase_est_cny": None,
            "margin_war_pct": None,
            "margin_peace_pct": None,
            "net_profit_war_cny": None,
            "fbn_fee_sar": None,
            "commission_pct": None,
            "size_tier": None,
            "v6_category": None,
        }

    # 从子类目名推断 V6 品类
    sub_cat = product.get("_subcategory", "")
    v6_cat = map_category(sub_cat, None)
    defaults = get_defaults(v6_cat)

    purchase_cny = price * settings.ksa_exchange_rate * defaults["cost_ratio"]

    try:
        result = calculate_ksa_fbn_profit(
            purchase_price_cny=purchase_cny,
            length_cm=defaults["l"],
            width_cm=defaults["w"],
            height_cm=defaults["h"],
            weight_kg=defaults["wt"],
            target_price_sar=price,
            category_name=v6_cat,
            shipping_rate_cbm=settings.shipping_rate_cbm,
            shipping_rate_cbm_peacetime=settings.shipping_rate_cbm_peacetime,
            exchange_rate=settings.ksa_exchange_rate,
            ad_rate=settings.default_ad_rate,
            return_rate=settings.default_return_rate,
        )
        return {
            "purchase_est_cny": round(purchase_cny, 1),
            "margin_war_pct": result["margin_war_pct"],
            "margin_peace_pct": result["margin_peace_pct"],
            "net_profit_war_cny": result["net_profit_war_cny"],
            "fbn_fee_sar": result["fbn_fee_sar"],
            "commission_pct": result["commission_rate_pct"],
            "size_tier": result["size_tier"],
            "v6_category": v6_cat,
        }
    except Exception:
        return {
            "purchase_est_cny": round(purchase_cny, 1),
            "margin_war_pct": None,
            "margin_peace_pct": None,
            "net_profit_war_cny": None,
            "fbn_fee_sar": None,
            "commission_pct": None,
            "size_tier": None,
            "v6_category": v6_cat,
        }


def classify_opportunity(row) -> str:
    """识别机会类型"""
    tags = []
    reviews = row.get("review_count", 0) or 0
    rating = row.get("rating") or 0
    images = row.get("image_count", 0) or 0
    margin = row.get("margin_war_pct")
    sold = row.get("sold_recently", "")

    sold_num = 0
    if sold:
        try:
            sold_num = int(str(sold).replace(",", "").replace("+", ""))
        except (ValueError, TypeError):
            pass

    has_demand = reviews >= 20 or sold_num >= 50
    poor_listing = images <= 3 or (rating > 0 and rating < 4.3)
    if has_demand and poor_listing:
        tags.append("Potential")

    if reviews < 50:
        tags.append("Blue Ocean")

    if margin is not None and margin > 20:
        tags.append("High Margin")

    return " | ".join(tags) if tags else "Regular"


def _style_header(ws, col_count):
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def generate_category_report(
    all_products: list[dict],
    subcategories: list[dict],
    output_path: Path,
    settings: Settings,
):
    """生成类目分析 Excel 报告"""
    logger.info(f"开始利润精算: {len(all_products)} 条商品")

    # 计算利润
    enriched = []
    for p in all_products:
        profit = calc_product_profit(p, settings)
        enriched.append({**p, **profit})

    df = pd.DataFrame(enriched)
    if df.empty:
        logger.error("无商品数据")
        return

    # ── Sheet 1: 子类目概览 ──
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sub-Category Overview"

    sub_stats = []
    for sub_name in df["_subcategory"].unique():
        sub_df = df[df["_subcategory"] == sub_name]
        sub_stats.append({
            "Subcategory": sub_name,
            "Products": len(sub_df),
            "Express %": round(sub_df["is_express"].sum() / len(sub_df) * 100, 1) if len(sub_df) > 0 else 0,
            "Median Price (SAR)": round(sub_df["price"].median(), 1),
            "Avg Reviews": round(sub_df["review_count"].mean(), 1),
            "Avg Rating": round(sub_df["rating"].dropna().mean(), 2) if sub_df["rating"].notna().any() else 0,
            "Ad Products": sub_df["is_ad"].sum(),
            "Median Margin %": round(sub_df["margin_war_pct"].dropna().median(), 1) if sub_df["margin_war_pct"].notna().any() else 0,
        })

    sub_stats_df = pd.DataFrame(sub_stats).sort_values("Products", ascending=False)

    for r_idx, row in enumerate(dataframe_to_rows(sub_stats_df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=val)
    _style_header(ws1, len(sub_stats_df.columns))
    ws1.freeze_panes = "A2"

    # ── Sheet 2: 全部商品列表（按利润率降序）──
    ws2 = wb.create_sheet("All Products")
    product_cols = [
        "product_id", "title", "_subcategory", "price", "original_price",
        "rating", "review_count", "is_express", "is_bestseller", "is_ad",
        "image_count", "sold_recently", "v6_category", "size_tier",
        "purchase_est_cny", "margin_war_pct", "margin_peace_pct",
        "net_profit_war_cny", "fbn_fee_sar", "commission_pct", "product_url",
    ]
    existing_cols = [c for c in product_cols if c in df.columns]
    df_sorted = df[existing_cols].sort_values("margin_war_pct", ascending=False, na_position="last")

    for r_idx, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
    _style_header(ws2, len(existing_cols))
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Top 100 产品推荐 ──
    ws3 = wb.create_sheet("Top 100 Recommendations")
    df["opportunity_type"] = df.apply(classify_opportunity, axis=1)
    # 过滤有利润数据的，按利润率降序取 Top 100
    top_df = df[df["margin_war_pct"].notna()].sort_values(
        "margin_war_pct", ascending=False
    ).head(100).reset_index(drop=True)
    top_df["rank"] = range(1, len(top_df) + 1)

    top_cols = [
        "rank", "product_id", "title", "_subcategory", "price",
        "margin_war_pct", "margin_peace_pct", "net_profit_war_cny",
        "opportunity_type", "review_count", "rating", "is_express",
        "is_ad", "sold_recently", "product_url",
    ]
    top_existing = [c for c in top_cols if c in top_df.columns]

    for r_idx, row in enumerate(dataframe_to_rows(top_df[top_existing], index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            ws3.cell(row=r_idx, column=c_idx, value=val)
    _style_header(ws3, len(top_existing))
    ws3.freeze_panes = "A2"

    # ── Sheet 4: 广告商品分析 ──
    ws4 = wb.create_sheet("Ad Products Analysis")
    ad_df = df[df["is_ad"] == True].sort_values("price", ascending=False).reset_index(drop=True)

    ad_cols = [
        "product_id", "title", "_subcategory", "price", "margin_war_pct",
        "review_count", "rating", "is_express", "sold_recently", "product_url",
    ]
    ad_existing = [c for c in ad_cols if c in ad_df.columns]

    if not ad_df.empty:
        for r_idx, row in enumerate(dataframe_to_rows(ad_df[ad_existing], index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                ws4.cell(row=r_idx, column=c_idx, value=val)
    else:
        ws4.cell(row=1, column=1, value="No ad products found")
    _style_header(ws4, len(ad_existing))
    ws4.freeze_panes = "A2"

    # 调整列宽
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    logger.info(f"报告已保存: {output_path}")

    # 返回 top_df 用于 Amazon 对比
    return top_df


async def main():
    settings = Settings()
    data_dir = settings.data_dir
    category = "automotive"

    # 1. 完整类目爬取
    logger.info("=" * 60)
    logger.info("  Step 1: Automotive 类目完整爬取")
    logger.info("=" * 60)

    crawler = NoonCategoryCrawler(category, data_dir, max_products_per_sub=500)
    summary = await crawler.run()

    if "error" in summary:
        logger.error(f"爬取失败: {summary['error']}")
        return

    # 2. 加载所有商品
    all_products_path = data_dir / "monitoring" / "categories" / category / "all_products.json"
    all_products = json.loads(all_products_path.read_text(encoding="utf-8"))
    logger.info(f"总商品数: {len(all_products)}")

    # 3. 加载类目树
    tree_path = data_dir / "monitoring" / "categories" / category / "category_tree.json"
    subcategories = json.loads(tree_path.read_text(encoding="utf-8"))

    # 4. 生成报告
    logger.info("=" * 60)
    logger.info("  Step 2: V6 利润精算 + 报告生成")
    logger.info("=" * 60)

    report_path = data_dir / "monitoring" / "categories" / category / "report_automotive.xlsx"
    top_df = generate_category_report(all_products, subcategories, report_path, settings)

    # 5. 打印汇总
    print("\n" + "=" * 60)
    print(f"  Automotive 类目爬取完成")
    print("=" * 60)
    print(f"  子类目数: {summary.get('total_subcategories', 0)}")
    print(f"  完成: {summary.get('completed', 0)}, 跳过(续跑): {summary.get('skipped_resume', 0)}")
    print(f"  独立商品: {summary.get('total_unique_products', 0)}")
    print(f"  耗时: {summary.get('duration_minutes', 0)} 分钟")
    print(f"  错误数: {len(summary.get('errors', []))}")
    print(f"  报告: {report_path}")
    print("=" * 60)

    return top_df


if __name__ == "__main__":
    asyncio.run(main())
