"""
第二轮精采集脚本 — 25个高价值关键词

流程:
1. 对缺失Noon数据的关键词补爬Noon搜索
2. Alibaba.com FOB采购价爬取
3. Noon详情页尺寸/重量（Top5商品）
4. V6定价引擎精算（真实采购价 vs 估算采购价对比）
5. 输出精算报告Excel
"""
import asyncio
import json
import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime
from statistics import median

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

from config.settings import Settings
from config.category_mapping import map_category
from config.cost_defaults import get_defaults
from analysis.pricing_engine import calculate_ksa_fbn_profit

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("refine")

# ═══ 关键词列表 ═══
KEYWORDS = {
    "运动健身": [
        "cardio training", "exercise machines", "yoga mat",
        "fitness equipment", "treadmill", "exercise bike",
    ],
    "健康营养": [
        "smoking cessation", "energy formula", "heart health supplements",
        "sports nutrition bars", "amino acids",
    ],
    "汽车配件": [
        "jump starter batteries", "dash cameras", "car audio",
        "full car covers", "exterior accessories",
    ],
    "美妆护肤": [
        "skincare cleansers", "haircare", "body gift set",
        "after shaves", "electronic personal care",
    ],
    "母婴": [
        "stroller travel systems", "double and twin strollers", "baby stroller",
    ],
}

ALL_KEYWORDS = []
KW_GROUP = {}
for group, kws in KEYWORDS.items():
    for kw in kws:
        ALL_KEYWORDS.append(kw)
        KW_GROUP[kw] = group

# 全量扫描快照 — Noon 数据源
FULL_SCAN_SNAP = Path("D:/claude noon/noon-selection-tool/data/snapshots/2026-03-17_114142")

# USD→CNY 汇率
USD_CNY = 7.25


async def main():
    settings = Settings()
    settings.set_runtime_scope("keyword")
    snap_dir = settings.snapshot_dir
    logger.info(f"精采集 Snapshot: {settings.snapshot_id}")
    logger.info(f"数据目录: {snap_dir}")

    for sub in ["noon", "amazon", "alibaba", "noon_details", "processed"]:
        (snap_dir / sub).mkdir(parents=True, exist_ok=True)

    # ═══ Step 0: 复制全量扫描的Noon数据到本次snapshot ═══
    logger.info("═══ Step 0: 复制全量扫描Noon数据 ═══")
    noon_src = FULL_SCAN_SNAP / "noon"
    noon_dst = snap_dir / "noon"
    copied = 0
    missing_noon = []
    for kw in ALL_KEYWORDS:
        safe = kw.replace(" ", "_")
        src_file = noon_src / f"{safe}.json"
        dst_file = noon_dst / f"{safe}.json"
        if src_file.exists():
            dst_file.write_text(src_file.read_text(encoding="utf-8"), encoding="utf-8")
            copied += 1
        else:
            missing_noon.append(kw)
    logger.info(f"已复制 {copied} 个, 缺失 {len(missing_noon)} 个: {missing_noon}")

    # ═══ Step 0.5: 补爬缺失的Noon关键词 ═══
    if missing_noon:
        logger.info(f"═══ Step 0.5: 补爬 {len(missing_noon)} 个缺失Noon关键词 ═══")
        try:
            from scrapers.noon_scraper import NoonScraper
            noon_scraper = NoonScraper(settings, skip_suggestions=True)
            # 手动设置 data_dir 到当前 snapshot
            noon_scraper.data_dir = noon_dst
            await noon_scraper.run(missing_noon)
            logger.info("Noon补爬完成")
        except Exception as e:
            logger.error(f"Noon补爬失败: {e}")
            traceback.print_exc()

    # ═══ Step 1: Alibaba FOB 采购价 ═══
    logger.info(f"\n═══ Step 1: Alibaba.com FOB采购价 ═══")
    alibaba_data = {}

    # 1a. 优先从 CSV 手动价格文件加载
    csv_path = Path("D:/claude noon/noon-selection-tool/data/alibaba_prices.csv")
    csv_prices = {}
    if csv_path.exists():
        from scrapers.alibaba_scraper import AlibabaCsvImporter
        importer = AlibabaCsvImporter(csv_path)
        csv_prices = importer.prices
        for kw in ALL_KEYWORDS:
            price = importer.get_price(kw)
            if price:
                alibaba_data[kw] = {
                    "keyword": kw,
                    "median_price_usd": price,
                    "products": [],
                    "total_results": 0,
                    "source": "csv_manual",
                }
        if csv_prices:
            logger.info(f"CSV手动价格: 已加载 {len(csv_prices)} 条, 匹配 {len(alibaba_data)} 个关键词")

    # 1b. 对没有CSV价格的关键词尝试爬取
    missing_ali = [kw for kw in ALL_KEYWORDS if kw not in alibaba_data]
    if missing_ali:
        logger.info(f"尝试Alibaba爬取剩余 {len(missing_ali)} 个关键词...")
        try:
            from scrapers.alibaba_scraper import AlibabaScraper
            alibaba = AlibabaScraper(settings)
            alibaba.data_dir = snap_dir / "alibaba"
            results = await alibaba.run(missing_ali)
            scraped_ok = 0
            for r in results:
                if r and r.get("keyword") and r.get("median_price_usd"):
                    alibaba_data[r["keyword"]] = r
                    scraped_ok += 1
            logger.info(f"Alibaba爬取: {scraped_ok}/{len(missing_ali)} 获得有效价格")
        except Exception as e:
            logger.error(f"Alibaba爬取失败: {e}")
            traceback.print_exc()

    # 1c. 从已有JSON文件加载
    alibaba_dir = snap_dir / "alibaba"
    for kw in ALL_KEYWORDS:
        if kw not in alibaba_data:
            safe = kw.replace(" ", "_").replace("/", "_").replace("\\", "_")
            fpath = alibaba_dir / f"{safe}.json"
            if fpath.exists():
                try:
                    data = json.loads(fpath.read_text(encoding="utf-8"))
                    if data.get("median_price_usd"):
                        alibaba_data[kw] = data
                except Exception:
                    pass

    ali_with_price = sum(1 for v in alibaba_data.values() if v.get("median_price_usd"))
    logger.info(f"Alibaba价格汇总: {ali_with_price}/{len(ALL_KEYWORDS)} 有真实采购价")

    # ═══ Step 2: Noon 详情页尺寸（Top5商品） ═══
    logger.info(f"\n═══ Step 2: Noon详情页尺寸采集 ═══")
    noon_details = {}  # kw -> [{url, dims}]

    # 先从上次精采集加载已有的详情页数据
    PREV_SNAP = Path("D:/claude noon/noon-selection-tool/data/snapshots/2026-03-17_220143")
    prev_details_dir = PREV_SNAP / "noon_details"
    reused = 0
    if prev_details_dir.exists():
        for kw in ALL_KEYWORDS:
            safe = kw.replace(" ", "_")
            prev_file = prev_details_dir / f"{safe}.json"
            if prev_file.exists():
                try:
                    data = json.loads(prev_file.read_text(encoding="utf-8"))
                    if data:
                        noon_details[kw] = data
                        # 同时复制到当前 snapshot
                        dst = snap_dir / "noon_details" / f"{safe}.json"
                        dst.write_text(prev_file.read_text(encoding="utf-8"), encoding="utf-8")
                        reused += 1
                except Exception:
                    pass
    if reused:
        logger.info(f"复用上次详情页数据: {reused} 个关键词")

    # 对没有详情页数据的关键词爬取
    missing_details = [kw for kw in ALL_KEYWORDS if kw not in noon_details]
    if missing_details:
        logger.info(f"需要爬取详情页: {len(missing_details)} 个关键词")
    else:
        logger.info("所有详情页数据已有，跳过爬取")

    if missing_details:
        try:
            from scrapers.noon_scraper import NoonScraper
            noon_detail_scraper = NoonScraper(settings)
            await noon_detail_scraper.start_browser()

            for kw in missing_details:
                safe = kw.replace(" ", "_")
                noon_file = noon_dst / f"{safe}.json"
                if not noon_file.exists():
                    continue
                rec = json.loads(noon_file.read_text(encoding="utf-8"))
                products = rec.get("products", [])[:5]
                kw_dims = []
                for p in products:
                    url = p.get("product_url", "")
                    if not url:
                        continue
                    try:
                        dims = await noon_detail_scraper.scrape_product_dimensions(url)
                        if dims:
                            kw_dims.append({"url": url, "title": p.get("title", ""), **dims})
                            logger.info(f"  [{kw}] 尺寸: {dims}")
                    except Exception as e:
                        logger.debug(f"  [{kw}] 详情页失败: {e}")
                noon_details[kw] = kw_dims
                detail_file = snap_dir / "noon_details" / f"{safe}.json"
                detail_file.write_text(json.dumps(kw_dims, ensure_ascii=False, indent=2), encoding="utf-8")

            await noon_detail_scraper.stop_browser()
        except Exception as e:
            logger.error(f"Noon详情页采集失败: {e}")
            traceback.print_exc()

    dims_found = sum(1 for v in noon_details.values() if v)
    logger.info(f"详情页汇总: {dims_found}/{len(ALL_KEYWORDS)} 关键词有尺寸数据")

    # ═══ Step 3: V6 精算 ═══
    logger.info(f"\n═══ Step 3: V6定价引擎精算 ═══")
    rows = []
    product_rows = []

    for kw in ALL_KEYWORDS:
        safe = kw.replace(" ", "_")
        noon_file = noon_dst / f"{safe}.json"
        if not noon_file.exists():
            logger.warning(f"[{kw}] 无Noon数据，跳过")
            continue

        rec = json.loads(noon_file.read_text(encoding="utf-8"))
        products = rec.get("products", [])
        if not products:
            logger.warning(f"[{kw}] Noon产品列表为空，跳过")
            continue

        # 品类映射
        primary_cat = rec.get("primary_category", "")
        cat_path = rec.get("category_path", [])
        v6_cat = map_category(primary_cat, cat_path if isinstance(cat_path, list) else None)
        defaults = get_defaults(v6_cat)

        # Noon 价格统计
        prices = [p["price"] for p in products if p.get("price") and p["price"] > 0]
        noon_median_price = median(prices) if prices else 0

        # Alibaba FOB 价格
        ali = alibaba_data.get(kw, {})
        alibaba_median_usd = ali.get("median_price_usd")
        alibaba_products = ali.get("products", [])
        alibaba_median_cny = round(alibaba_median_usd * USD_CNY, 2) if alibaba_median_usd else None

        # 估算采购价（售价比例法）
        est_purchase_cny = round(noon_median_price * settings.ksa_exchange_rate * defaults["cost_ratio"], 2)

        # 真实采购价（优先Alibaba）
        real_purchase_cny = alibaba_median_cny if alibaba_median_cny else est_purchase_cny
        purchase_source = "Alibaba FOB" if alibaba_median_cny else "估算(售价比例)"

        # 尺寸/重量 — 优先详情页数据
        detail_dims = noon_details.get(kw, [])
        if detail_dims:
            # 取第一个有完整尺寸的
            for dd in detail_dims:
                if "length_cm" in dd and "width_cm" in dd and "height_cm" in dd:
                    l, w, h = dd["length_cm"], dd["width_cm"], dd["height_cm"]
                    wt = dd.get("weight_kg", defaults["wt"])
                    dims_source = "详情页实测"
                    break
            else:
                l, w, h, wt = defaults["l"], defaults["w"], defaults["h"], defaults["wt"]
                dims_source = "品类默认"
        else:
            l, w, h, wt = defaults["l"], defaults["w"], defaults["h"], defaults["wt"]
            dims_source = "品类默认"

        # V6精算 — 真实采购价
        calc_real = calculate_ksa_fbn_profit(
            purchase_price_cny=real_purchase_cny,
            length_cm=l, width_cm=w, height_cm=h, weight_kg=wt,
            target_price_sar=noon_median_price,
            category_name=v6_cat,
            shipping_rate_cbm=settings.shipping_rate_cbm,
            shipping_rate_cbm_peacetime=settings.shipping_rate_cbm_peacetime,
            exchange_rate=settings.ksa_exchange_rate,
            ad_rate=settings.default_ad_rate,
            return_rate=settings.default_return_rate,
        )

        # V6精算 — 估算采购价（对比用）
        calc_est = calculate_ksa_fbn_profit(
            purchase_price_cny=est_purchase_cny,
            length_cm=l, width_cm=w, height_cm=h, weight_kg=wt,
            target_price_sar=noon_median_price,
            category_name=v6_cat,
            shipping_rate_cbm=settings.shipping_rate_cbm,
            shipping_rate_cbm_peacetime=settings.shipping_rate_cbm_peacetime,
            exchange_rate=settings.ksa_exchange_rate,
            ad_rate=settings.default_ad_rate,
            return_rate=settings.default_return_rate,
        )

        row = {
            "group": KW_GROUP[kw],
            "keyword": kw,
            "v6_category": v6_cat,
            "noon_total_results": rec.get("total_results", 0),
            "noon_express_results": rec.get("express_results", 0),
            "noon_median_price_sar": round(noon_median_price, 2),
            "noon_product_count": len(products),
            # Alibaba
            "alibaba_median_usd": alibaba_median_usd or "",
            "alibaba_median_cny": alibaba_median_cny or "",
            "alibaba_product_count": len(alibaba_products),
            "purchase_source": purchase_source,
            # 采购价对比
            "est_purchase_cny": est_purchase_cny,
            "real_purchase_cny": real_purchase_cny,
            # 尺寸
            "dims_source": dims_source,
            "length_cm": l, "width_cm": w, "height_cm": h, "weight_kg": wt,
            "size_tier": calc_real["size_tier"],
            # V6精算 — 真实采购价
            "commission_rate_pct": calc_real["commission_rate_pct"],
            "fbn_fee_sar": calc_real["fbn_fee_sar"],
            "margin_war_pct": calc_real["margin_war_pct"],
            "margin_peace_pct": calc_real["margin_peace_pct"],
            "net_profit_war_cny": calc_real["net_profit_war_cny"],
            "net_profit_peace_cny": calc_real["net_profit_peace_cny"],
            "roi_war_pct": calc_real["roi_war_pct"],
            "total_cost_war_cny": calc_real["total_cost_war_cny"],
            "revenue_cny": calc_real["revenue_cny"],
            # V6精算 — 估算采购价（对比）
            "est_margin_war_pct": calc_est["margin_war_pct"],
            "est_margin_peace_pct": calc_est["margin_peace_pct"],
            "est_net_profit_war_cny": calc_est["net_profit_war_cny"],
            # 差异
            "margin_diff_pp": round(calc_real["margin_war_pct"] - calc_est["margin_war_pct"], 2),
        }
        rows.append(row)
        logger.info(
            f"  [{kw}] 售价={noon_median_price:.0f}SAR, "
            f"采购={real_purchase_cny:.0f}CNY({purchase_source}), "
            f"战时利润={calc_real['margin_war_pct']:.1f}%, "
            f"和平={calc_real['margin_peace_pct']:.1f}%"
        )

        # ═══ Top 3 产品推荐 ═══
        # 按利润率排序
        product_calcs = []
        for p in products[:10]:  # 取前10个算，选出Top3
            pp = p.get("price", 0)
            if not pp or pp <= 0:
                continue
            # 用真实采购价比例推算单品采购价
            if alibaba_median_cny and noon_median_price > 0:
                p_purchase = round(alibaba_median_cny * (pp / noon_median_price), 2)
            else:
                p_purchase = round(pp * settings.ksa_exchange_rate * defaults["cost_ratio"], 2)

            p_calc = calculate_ksa_fbn_profit(
                purchase_price_cny=p_purchase,
                length_cm=l, width_cm=w, height_cm=h, weight_kg=wt,
                target_price_sar=pp,
                category_name=v6_cat,
                shipping_rate_cbm=settings.shipping_rate_cbm,
                shipping_rate_cbm_peacetime=settings.shipping_rate_cbm_peacetime,
                exchange_rate=settings.ksa_exchange_rate,
                ad_rate=settings.default_ad_rate,
                return_rate=settings.default_return_rate,
            )
            product_calcs.append((p, p_purchase, p_calc))

        # 按战时利润率降序
        product_calcs.sort(key=lambda x: x[2]["margin_war_pct"], reverse=True)
        for rank, (p, p_purchase, p_calc) in enumerate(product_calcs[:3], 1):
            product_rows.append({
                "group": KW_GROUP[kw],
                "keyword": kw,
                "product_rank": rank,
                "title": p.get("title", "")[:80],
                "price_sar": p.get("price", 0),
                "purchase_cny": p_purchase,
                "margin_war_pct": p_calc["margin_war_pct"],
                "margin_peace_pct": p_calc["margin_peace_pct"],
                "net_profit_war_cny": p_calc["net_profit_war_cny"],
                "review_count": p.get("review_count", 0),
                "rating": p.get("rating", ""),
                "is_express": p.get("is_express", False),
                "is_bestseller": p.get("is_bestseller", False),
                "fbn_fee_sar": p_calc["fbn_fee_sar"],
                "commission_rate_pct": p_calc["commission_rate_pct"],
                "size_tier": p_calc["size_tier"],
                "product_url": p.get("product_url", ""),
            })

    # ═══ Step 4: 生成Excel报告 ═══
    logger.info(f"\n═══ Step 4: 生成精算Excel报告 ═══")
    df = pd.DataFrame(rows)
    product_df = pd.DataFrame(product_rows)

    report_path = snap_dir / f"refine_report_{settings.snapshot_id}.xlsx"
    generate_refine_report(df, product_df, report_path)
    logger.info(f"报告已保存: {report_path}")

    # 保存中间数据
    try:
        df.to_parquet(snap_dir / "processed" / "refine_scored.parquet", index=False)
    except ImportError:
        df.to_csv(snap_dir / "processed" / "refine_scored.csv", index=False, encoding="utf-8-sig")

    # ═══ 打印汇总 ═══
    print("\n" + "=" * 70)
    print("  第二轮精采集结果汇总")
    print("=" * 70)
    print(f"  关键词总数: {len(ALL_KEYWORDS)}")
    print(f"  Alibaba有FOB价: {sum(1 for r in rows if r['alibaba_median_usd'])} 个")
    print(f"  详情页有尺寸: {sum(1 for r in rows if r['dims_source'] == '详情页实测')} 个")
    print(f"  产品推荐: {len(product_rows)} 个")
    print()

    # 按品类组打印
    for group in KEYWORDS:
        group_rows = [r for r in rows if r["group"] == group]
        if not group_rows:
            continue
        print(f"  ── {group} ──")
        for r in sorted(group_rows, key=lambda x: x["margin_war_pct"], reverse=True):
            ali_info = f"FOB ${r['alibaba_median_usd']}" if r['alibaba_median_usd'] else "估算"
            print(
                f"    {r['keyword']:<30} "
                f"售价{r['noon_median_price_sar']:>6.0f}SAR  "
                f"采购{r['real_purchase_cny']:>6.0f}CNY({ali_info:<12})  "
                f"战时{r['margin_war_pct']:>6.1f}%  "
                f"和平{r['margin_peace_pct']:>6.1f}%  "
                f"利润{r['net_profit_war_cny']:>7.0f}CNY"
            )
        print()

    print(f"  报告文件: {report_path}")
    print("=" * 70)


def generate_refine_report(df: pd.DataFrame, product_df: pd.DataFrame, output_path: Path):
    """生成精算报告Excel"""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def auto_width(ws, max_width=30):
        for col_cells in ws.columns:
            length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value) if cell.value else ""
                    length = max(length, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(length + 3, max_width)

    # ═══ Sheet 1: 精算总览 ═══
    ws1 = wb.active
    ws1.title = "精算总览"
    headers1 = [
        "品类组", "关键词", "V6品类", "Noon结果数", "Express数",
        "Noon中位价SAR", "Alibaba FOB(USD)", "Alibaba(CNY)",
        "估算采购CNY", "真实采购CNY", "采购来源",
        "尺寸来源", "尺寸档", "佣金率%", "FBN出库费SAR",
        "战时利润率%", "和平利润率%", "战时净利润CNY", "和平净利润CNY",
        "战时ROI%", "总成本CNY", "收入CNY",
        "估算战时利润率%", "估算和平利润率%", "利润差异pp",
    ]
    ws1.append(headers1)
    style_header(ws1, len(headers1))

    for _, r in df.iterrows():
        row_data = [
            r["group"], r["keyword"], r["v6_category"],
            r["noon_total_results"], r["noon_express_results"],
            r["noon_median_price_sar"], r["alibaba_median_usd"], r["alibaba_median_cny"],
            r["est_purchase_cny"], r["real_purchase_cny"], r["purchase_source"],
            r["dims_source"], r["size_tier"], r["commission_rate_pct"], r["fbn_fee_sar"],
            r["margin_war_pct"], r["margin_peace_pct"],
            r["net_profit_war_cny"], r["net_profit_peace_cny"],
            r["roi_war_pct"], r["total_cost_war_cny"], r["revenue_cny"],
            r["est_margin_war_pct"], r["est_margin_peace_pct"], r["margin_diff_pp"],
        ]
        ws1.append(row_data)
        row_idx = ws1.max_row
        # 利润率着色
        margin_cell = ws1.cell(row=row_idx, column=16)
        if isinstance(margin_cell.value, (int, float)):
            if margin_cell.value >= 30:
                margin_cell.fill = green_fill
            elif margin_cell.value < 0:
                margin_cell.fill = red_fill
            else:
                margin_cell.fill = yellow_fill

    auto_width(ws1)

    # ═══ Sheet 2: 产品推荐 ═══
    ws2 = wb.create_sheet("Top3产品推荐")
    headers2 = [
        "品类组", "关键词", "排名", "商品标题", "售价SAR",
        "采购价CNY", "战时利润率%", "和平利润率%", "战时净利润CNY",
        "评论数", "评分", "Express", "Best Seller",
        "FBN费SAR", "佣金率%", "尺寸档", "Noon链接",
    ]
    ws2.append(headers2)
    style_header(ws2, len(headers2))

    for _, r in product_df.iterrows():
        ws2.append([
            r["group"], r["keyword"], r["product_rank"], r["title"], r["price_sar"],
            r["purchase_cny"], r["margin_war_pct"], r["margin_peace_pct"], r["net_profit_war_cny"],
            r["review_count"], r["rating"],
            "Yes" if r["is_express"] else "",
            "Yes" if r["is_bestseller"] else "",
            r["fbn_fee_sar"], r["commission_rate_pct"], r["size_tier"], r["product_url"],
        ])
        row_idx = ws2.max_row
        margin_cell = ws2.cell(row=row_idx, column=7)
        if isinstance(margin_cell.value, (int, float)):
            if margin_cell.value >= 30:
                margin_cell.fill = green_fill
            elif margin_cell.value < 0:
                margin_cell.fill = red_fill

    auto_width(ws2, max_width=40)

    # ═══ Sheet 3: Alibaba供应商明细 ═══
    ws3 = wb.create_sheet("Alibaba供应商明细")
    headers3 = [
        "关键词", "供应商", "产品标题", "FOB低价USD", "FOB高价USD",
        "最小起订", "交易量",
    ]
    ws3.append(headers3)
    style_header(ws3, len(headers3))

    alibaba_dir = Path(str(df.iloc[0]["keyword"])) if not df.empty else None
    # 从snapshot目录加载alibaba数据
    settings = Settings()
    settings.set_runtime_scope("keyword")
    ali_dir = settings.snapshot_dir / "alibaba"
    for kw in ALL_KEYWORDS:
        safe = kw.replace(" ", "_").replace("/", "_").replace("\\", "_")
        fpath = ali_dir / f"{safe}.json"
        if not fpath.exists():
            continue
        try:
            data = json.loads(fpath.read_text(encoding="utf-8"))
            for p in data.get("products", [])[:5]:
                ws3.append([
                    kw, p.get("supplier_name", ""), p.get("title", "")[:60],
                    p.get("price_low", ""), p.get("price_high", ""),
                    p.get("min_order", ""), p.get("transaction_count", 0),
                ])
        except Exception:
            continue

    auto_width(ws3, max_width=40)

    # ═══ Sheet 4: 采购价对比 ═══
    ws4 = wb.create_sheet("采购价对比分析")
    headers4 = [
        "品类组", "关键词", "Noon中位售价SAR", "售价CNY",
        "估算采购CNY", "Alibaba采购CNY", "采购价差异CNY", "采购价差异%",
        "估算利润率%", "真实利润率%", "利润率差异pp", "判断",
    ]
    ws4.append(headers4)
    style_header(ws4, len(headers4))

    for _, r in df.iterrows():
        est = r["est_purchase_cny"]
        real = r["real_purchase_cny"]
        diff_cny = round(real - est, 2)
        diff_pct = round((real - est) / est * 100, 1) if est > 0 else 0
        judgment = ""
        if r["margin_diff_pp"] > 5:
            judgment = "估算偏悲观(实际更赚)"
        elif r["margin_diff_pp"] < -5:
            judgment = "估算偏乐观(实际更亏)"
        elif r["alibaba_median_usd"] == "":
            judgment = "无Alibaba数据"
        else:
            judgment = "估算接近"

        ws4.append([
            r["group"], r["keyword"],
            r["noon_median_price_sar"],
            round(r["noon_median_price_sar"] * 1.90, 2),
            est, r["alibaba_median_cny"] if r["alibaba_median_cny"] else "N/A",
            diff_cny if r["alibaba_median_cny"] else "N/A",
            diff_pct if r["alibaba_median_cny"] else "N/A",
            r["est_margin_war_pct"], r["margin_war_pct"],
            r["margin_diff_pp"], judgment,
        ])

    auto_width(ws4)

    wb.save(output_path)
    logger.info(f"Excel报告已保存: {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
