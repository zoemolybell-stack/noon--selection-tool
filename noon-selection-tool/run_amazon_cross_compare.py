"""
步骤8: Amazon 交叉对比

对 Automotive Top 100 商品标题，去 Amazon.sa 搜索同款，
记录价格/评论/BSR，结果追加到报告 Sheet 5。
"""
import asyncio
import json
import logging
import random
import re
import sys
import time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from playwright.async_api import async_playwright

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("amazon_cross.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("amazon-cross")

AMAZON_BASE = "https://www.amazon.sa"

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


async def search_amazon_for_product(page, title: str) -> dict:
    """
    在 Amazon.sa 搜索商品标题，返回最匹配的结果。
    """
    # 截取标题前8个词作为搜索词，避免太长
    words = title.split()[:8]
    query = " ".join(words)

    result = {
        "search_query": query,
        "amazon_title": "",
        "amazon_price_sar": None,
        "amazon_rating": None,
        "amazon_review_count": 0,
        "amazon_is_bestseller": False,
        "amazon_is_choice": False,
        "amazon_product_url": "",
        "amazon_asin": "",
        "match_found": False,
    }

    try:
        url = f"{AMAZON_BASE}/s?k={query}"
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await page.wait_for_timeout(3000)

        # 取第一个搜索结果
        cards = page.locator('[data-component-type="s-search-result"]')
        count = await cards.count()
        if count == 0:
            return result

        card = cards.first
        # 标题
        title_el = card.locator('h2 span')
        if await title_el.count() > 0:
            result["amazon_title"] = (await title_el.first.inner_text()).strip()

        # ASIN
        asin = await card.get_attribute("data-asin")
        if asin:
            result["amazon_asin"] = asin
            result["amazon_product_url"] = f"{AMAZON_BASE}/dp/{asin}"

        # 价格
        price_el = card.locator('.a-price .a-offscreen')
        if await price_el.count() > 0:
            text = await price_el.first.inner_text()
            nums = re.findall(r'[\d.]+', text.replace(",", ""))
            if nums:
                result["amazon_price_sar"] = float(nums[0])

        # 评分
        rating_el = card.locator('[class*="a-icon-alt"]')
        if await rating_el.count() > 0:
            text = await rating_el.first.inner_text()
            nums = re.findall(r'[\d.]+', text)
            if nums:
                result["amazon_rating"] = float(nums[0])

        # 评论数
        review_el = card.locator('a[href*="customerReviews"] span, span[class*="a-size-base"][dir="auto"]')
        if await review_el.count() > 0:
            text = await review_el.first.inner_text()
            nums = re.findall(r'[\d,]+', text)
            if nums:
                result["amazon_review_count"] = int(nums[0].replace(",", ""))

        # Best Seller
        bsr_el = card.locator('[class*="best-seller"], span:has-text("Best Seller")')
        result["amazon_is_bestseller"] = await bsr_el.count() > 0

        # Amazon's Choice
        choice_el = card.locator('[class*="amazons-choice"], span:has-text("Amazon\'s Choice")')
        result["amazon_is_choice"] = await choice_el.count() > 0

        result["match_found"] = bool(result["amazon_title"])

    except Exception as e:
        logger.warning(f"Amazon 搜索失败 [{query[:40]}]: {e}")
        result["error"] = str(e)

    return result


async def main():
    data_dir = Path("data/monitoring/categories/automotive")
    report_path = data_dir / "report_automotive.xlsx"

    if not report_path.exists():
        logger.error(f"报告文件不存在: {report_path}")
        return

    # 加载已有报告的 Top 100
    wb = load_workbook(report_path)
    ws_top = wb["Top 100 Recommendations"]

    # 读取 Top 100 的标题列
    headers = [cell.value for cell in ws_top[1]]
    title_col = headers.index("title") + 1 if "title" in headers else None
    pid_col = headers.index("product_id") + 1 if "product_id" in headers else None
    price_col = headers.index("price") + 1 if "price" in headers else None
    margin_col = headers.index("margin_war_pct") + 1 if "margin_war_pct" in headers else None

    if not title_col:
        logger.error("Top 100 sheet 中未找到 title 列")
        return

    top_products = []
    for row in ws_top.iter_rows(min_row=2, max_row=ws_top.max_row):
        title = row[title_col - 1].value
        pid = row[pid_col - 1].value if pid_col else ""
        price = row[price_col - 1].value if price_col else 0
        margin = row[margin_col - 1].value if margin_col else 0
        if title:
            top_products.append({
                "title": title,
                "noon_product_id": pid,
                "noon_price_sar": price,
                "noon_margin_pct": margin,
            })

    logger.info(f"Top 100 商品加载完成: {len(top_products)} 条")

    # 检查断点续跑
    progress_file = data_dir / "amazon_cross_progress.json"
    completed_results = []
    completed_pids = set()
    if progress_file.exists():
        completed_results = json.loads(progress_file.read_text(encoding="utf-8"))
        completed_pids = {r["noon_product_id"] for r in completed_results if r.get("noon_product_id")}
        logger.info(f"断点续跑: 已完成 {len(completed_pids)} 条")

    # 启动浏览器
    pw = await async_playwright().start()
    browser = await pw.chromium.launch(
        headless=False,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-http2",
            "--no-sandbox",
            "--window-position=-9999,-9999",
        ],
    )
    ctx = await browser.new_context(
        viewport={"width": 1920, "height": 1080},
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        locale="en-US",
    )
    try:
        from playwright_stealth import Stealth
        stealth = Stealth()
        await stealth.apply_stealth_async(ctx)
    except Exception:
        pass

    page = await ctx.new_page()
    logger.info("浏览器已启动，开始 Amazon 交叉对比")

    start_time = time.time()
    results = list(completed_results)

    try:
        for i, prod in enumerate(top_products, 1):
            pid = prod["noon_product_id"]
            if pid in completed_pids:
                logger.info(f"[{i}/{len(top_products)}] 跳过已完成: {pid}")
                continue

            logger.info(f"[{i}/{len(top_products)}] 搜索: {prod['title'][:50]}...")

            amazon_result = await search_amazon_for_product(page, prod["title"])

            row_data = {
                **prod,
                **amazon_result,
            }

            # 价差计算
            if amazon_result.get("amazon_price_sar") and prod.get("noon_price_sar"):
                row_data["price_diff_pct"] = round(
                    (prod["noon_price_sar"] - amazon_result["amazon_price_sar"])
                    / amazon_result["amazon_price_sar"] * 100, 1
                )
            else:
                row_data["price_diff_pct"] = None

            results.append(row_data)

            # 每条保存进度
            progress_file.write_text(
                json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8"
            )

            if amazon_result["match_found"]:
                price_info = f"SAR {amazon_result['amazon_price_sar']}" if amazon_result['amazon_price_sar'] else "N/A"
                logger.info(f"  Found: {price_info}, reviews={amazon_result['amazon_review_count']}")
            else:
                logger.info(f"  No match found")

            # 随机延迟
            await asyncio.sleep(random.uniform(3, 5))

    except Exception as e:
        logger.error(f"Amazon 对比中断: {e}")
    finally:
        await page.close()
        await ctx.close()
        await browser.close()
        await pw.stop()

    elapsed = time.time() - start_time
    logger.info(f"Amazon 对比完成: {len(results)} 条, 耗时 {elapsed/60:.1f} 分钟")

    # 写入 Sheet 5
    df_cross = pd.DataFrame(results)
    ws5 = wb.create_sheet("Amazon Cross Compare")

    cross_cols = [
        "noon_product_id", "title", "noon_price_sar", "noon_margin_pct",
        "search_query", "match_found",
        "amazon_title", "amazon_price_sar", "amazon_rating",
        "amazon_review_count", "amazon_is_bestseller", "amazon_is_choice",
        "price_diff_pct", "amazon_product_url",
    ]
    existing_cols = [c for c in cross_cols if c in df_cross.columns]

    # 写表头
    for c_idx, col_name in enumerate(existing_cols, 1):
        cell = ws5.cell(row=1, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 写数据
    for r_idx, (_, row) in enumerate(df_cross[existing_cols].iterrows(), 2):
        for c_idx, val in enumerate(row, 1):
            ws5.cell(row=r_idx, column=c_idx, value=val)

    ws5.freeze_panes = "A2"

    # 调整列宽
    for col in ws5.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws5.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(report_path)
    logger.info(f"Sheet 5 已追加到报告: {report_path}")

    # 打印汇总
    matched = sum(1 for r in results if r.get("match_found"))
    print("\n" + "=" * 60)
    print("  Amazon 交叉对比完成")
    print("=" * 60)
    print(f"  Top 100 商品: {len(top_products)}")
    print(f"  Amazon 匹配: {matched}")
    print(f"  未匹配: {len(results) - matched}")
    if df_cross["price_diff_pct"].notna().any():
        avg_diff = df_cross["price_diff_pct"].dropna().mean()
        print(f"  Noon vs Amazon 平均价差: {avg_diff:+.1f}%")
    print(f"  报告: {report_path}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
