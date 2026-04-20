#!/usr/bin/env python3
"""
Noon V6 Pro 定价表 Excel 生成器
4种模式: FBN, 海外仓FBP, 直邮FBP, Global(UAE→KSA)
3个Sheet: Config, UAE_Pricing, KSA_Pricing
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
from copy import copy

# ============================================================
# STYLE DEFINITIONS
# ============================================================
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

COLORS = {
    'input': 'E2EFDA',       # light green
    'shared': 'F2F2F2',      # light gray
    'fbn': 'DAEEF3',         # light blue
    'ow_fbp': 'FDE9D9',      # light orange
    'dm_fbp': 'D5E8D4',      # light teal
    'global': 'F2DCDB',      # light red
    'summary': 'E4DFEC',     # light purple
    'header_dark': '1B3A5C',  # dark blue
    'header_mid': '4472C4',   # medium blue
    'editable': 'FFF2CC',    # light yellow
    'positive': 'C6EFCE',
    'negative': 'FFC7CE',
    'config_title': '2F5496',
}

FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') for k, v in COLORS.items()}

FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
FONT_NORMAL = Font(name='Calibri', size=10)
FONT_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_GREEN = Font(name='Calibri', size=10, bold=True, color='00B894')
FONT_RED = Font(name='Calibri', size=10, bold=True, color='D63031')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# WPS-COMPATIBLE FORMULA GENERATORS (no LET, no AGGREGATE, no CSE)
# ============================================================

def gen_size_tier_formula(r, l_col, w_col, h_col):
    """Generate FBN size tier formula using MEDIAN/MAX/MIN (no LET)."""
    mx = f'MAX({l_col}{r},{w_col}{r},{h_col}{r})'
    md = f'MEDIAN({l_col}{r},{w_col}{r},{h_col}{r})'
    mn = f'MIN({l_col}{r},{w_col}{r},{h_col}{r})'
    tiers = [
        (20, 15, 1, "Small Envelope"),
        (33, 23, 2.5, "Standard Envelope"),
        (33, 23, 5, "Large Envelope"),
        (45, 34, 26, "Standard Parcel"),
        (130, 34, 26, "Oversize"),
        (130, 130, 130, "Extra Oversize"),
    ]
    result = '"Bulky"'
    for max_l, max_w, max_h, name in reversed(tiers):
        result = f'IF(AND({mx}<={max_l},{md}<={max_w},{mn}<={max_h}),"{name}",{result})'
    return f'={result}'


def gen_fbn_fee_formula(fees_data, r, tier_col, cw_col, price_col):
    """Generate nested IF formula for FBN outbound fee lookup (no LET/AGGREGATE).
    fees_data: list of (tier_name, max_weight, low_fee, high_fee)
    """
    from collections import OrderedDict
    tiers = OrderedDict()
    for tier, max_wt, low, high in fees_data:
        if tier not in tiers:
            tiers[tier] = []
        tiers[tier].append((max_wt, low, high))

    def weight_chain(entries):
        """Build IF(cw<=wt, IF(price<=25,low,high), IF(cw<=wt2,...))"""
        if not entries:
            return '0'
        wt, low, high = entries[0]
        fee = f'IF({price_col}{r}<=25,{low},{high})'
        if len(entries) == 1:
            return fee
        rest = weight_chain(entries[1:])
        return f'IF({cw_col}{r}<={wt},{fee},{rest})'

    # Build outer tier chain
    result = '0'
    for tier_name in reversed(tiers):
        wt_formula = weight_chain(tiers[tier_name])
        result = f'IF({tier_col}{r}="{tier_name}",{wt_formula},{result})'
    return f'={result}'


def gen_comm_formula(r, cat_col, price_col, type_col, base_col, limit_col, tier_col):
    """Generate commission formula without LET — repeats MATCH calls for WPS compat."""
    m = f'MATCH({cat_col}{r},INDEX(Comm_Table,,1),0)'
    vtype = f'INDEX(Comm_Table,{m},{type_col})'
    vbase = f'INDEX(Comm_Table,{m},{base_col})'
    vlimit = f'INDEX(Comm_Table,{m},{limit_col})'
    vtier = f'INDEX(Comm_Table,{m},{tier_col})'
    p = f'{price_col}{r}'
    return (
        f'=IFERROR(IF({vtype}="Fixed",{p}*{vbase},'
        f'IF({vtype}="Threshold",IF({p}<={vlimit},{p}*{vbase},{p}*{vtier}),'
        f'MIN({p},{vlimit})*{vbase}+MAX(0,{p}-{vlimit})*{vtier})),0)'
    )


def gen_fbp_fee_formula(r, cw_col, dropoff_col, pickup_col):
    """Generate FBP delivery fee formula without LET."""
    return (
        f'=IFERROR(INDEX(FBP_Fees,'
        f'MATCH({cw_col}{r},INDEX(FBP_Fees,,1),1),'
        f'IF(FBP_DeliveryMode="Drop-off",{dropoff_col},{pickup_col})),0)'
    )


def style_cell(ws, row, col, value=None, font=FONT_NORMAL, fill=None, alignment=ALIGN_CENTER, border=THIN_BORDER, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell

def style_header_row(ws, row, col_start, col_end, headers, fill, font=FONT_HEADER):
    for i, h in enumerate(headers):
        style_cell(ws, row, col_start + i, h, font=font, fill=fill)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================
# COMMISSION DATA (from official docs - FBN/FBP split)
# ============================================================
# Format: (name, uae_fbn_type, uae_fbn_base, uae_fbn_limit, uae_fbn_tier,
#                uae_fbp_type, uae_fbp_base, uae_fbp_limit, uae_fbp_tier,
#                ksa_fbn_type, ksa_fbn_base, ksa_fbn_limit, ksa_fbn_tier,
#                ksa_fbp_type, ksa_fbp_base, ksa_fbp_limit, ksa_fbp_tier)

CATEGORIES = [
    # Fashion
    ("服装/鞋履 Apparel/Footwear", "Fixed",0.27,0,0, "Fixed",0.27,0,0, "Fixed",0.27,0,0, "Fixed",0.27,0,0),
    ("旅行箱包 Bags:Luggage", "Fixed",0.20,0,0, "Fixed",0.20,0,0, "Fixed",0.20,0,0, "Fixed",0.20,0,0),
    ("其他包袋 Bags:Other", "Fixed",0.25,0,0, "Fixed",0.25,0,0, "Fixed",0.25,0,0, "Fixed",0.25,0,0),
    ("手表 Watches", "Sliding",0.15,5000,0.05, "Sliding",0.15,5000,0.05, "Sliding",0.15,5000,0.05, "Sliding",0.15,5000,0.05),
    ("眼镜 Eyewear", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("高级珠宝 Jewelry:Fine", "Sliding",0.16,1000,0.05, "Sliding",0.16,1000,0.05, "Sliding",0.15,1000,0.05, "Sliding",0.15,1000,0.05),
    ("时尚饰品 Jewelry:Fashion", "Fixed",0.25,0,0, "Fixed",0.25,0,0, "Fixed",0.25,0,0, "Fixed",0.25,0,0),
    ("金条/金币 Gold Bars", "Fixed",0.05,0,0, "Fixed",0.05,0,0, "Fixed",0.05,0,0, "Fixed",0.05,0,0),
    ("银条 Silver Bars", "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0),
    # Electronics
    ("手机/平板 Mobiles/Tablets", "Threshold",0.06,500,0.05, "Threshold",0.06,500,0.05, "Threshold",0.05,750,0.045, "Threshold",0.06,750,0.055),
    ("苹果/三星国际版", "Fixed",0.06,0,0, "Fixed",0.06,0,0, "Fixed",0.06,0,0, "Fixed",0.06,0,0),
    ("SIM卡 SIM Cards", "Fixed",0.04,0,0, "Fixed",0.04,0,0, "Fixed",0.04,0,0, "Fixed",0.04,0,0),
    ("触控笔 Stylus Pens", "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0),
    ("笔记本/台式机 Laptops", "Fixed",0.06,0,0, "Fixed",0.065,0,0, "Fixed",0.06,0,0, "Fixed",0.065,0,0),
    ("内存卡 Memory Cards", "Fixed",0.07,0,0, "Fixed",0.075,0,0, "Fixed",0.07,0,0, "Fixed",0.075,0,0),
    ("U盘 USB Flash", "Threshold",0.15,250,0.08, "Threshold",0.155,250,0.085, "Threshold",0.15,250,0.08, "Threshold",0.155,250,0.085),
    ("硬盘/SSD Storage", "Fixed",0.06,0,0, "Fixed",0.065,0,0, "Fixed",0.06,0,0, "Fixed",0.065,0,0),
    ("网络配件 Network", "Fixed",0.06,0,0, "Fixed",0.065,0,0, "Fixed",0.06,0,0, "Fixed",0.065,0,0),
    ("显示器/显卡 Monitors/GPU", "Fixed",0.06,0,0, "Fixed",0.065,0,0, "Fixed",0.06,0,0, "Fixed",0.065,0,0),
    ("其他电脑硬件 PC Hardware", "Fixed",0.13,0,0, "Fixed",0.135,0,0, "Fixed",0.13,0,0, "Fixed",0.135,0,0),
    ("软件 Software", "Fixed",0.15,0,0, "Fixed",0.155,0,0, "Fixed",0.15,0,0, "Fixed",0.155,0,0),
    ("电脑包 Laptop Bags", "Fixed",0.15,0,0, "Fixed",0.155,0,0, "Fixed",0.15,0,0, "Fixed",0.155,0,0),
    ("打印机/扫描仪 Printers", "Fixed",0.06,0,0, "Fixed",0.065,0,0, "Fixed",0.06,0,0, "Fixed",0.065,0,0),
    ("其他办公电子 Office Electronics", "Fixed",0.13,0,0, "Fixed",0.135,0,0, "Fixed",0.13,0,0, "Fixed",0.135,0,0),
    ("手机配件≤50 Phone Acc≤50", "Fixed",0.20,0,0, "Fixed",0.20,0,0, "Fixed",0.20,0,0, "Fixed",0.20,0,0),
    ("手机配件>50 Phone Acc>50", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("耳机 Headphones", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.20,0,0, "Fixed",0.20,0,0),
    ("AR/VR眼镜", "Fixed",0.12,0,0, "Fixed",0.12,0,0, "Fixed",0.12,0,0, "Fixed",0.12,0,0),
    ("智能手表/穿戴 Smartwatches", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("相机/镜头 Cameras", "Fixed",0.06,0,0, "Fixed",0.06,0,0, "Threshold",0.10,250,0.08, "Threshold",0.10,250,0.08),
    ("电视 TVs", "Fixed",0.05,0,0, "Fixed",0.05,0,0, "Fixed",0.08,0,0, "Fixed",0.08,0,0),
    ("投影/流媒体 Projectors", "Fixed",0.05,0,0, "Fixed",0.05,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0),
    ("游戏主机 Consoles", "Fixed",0.06,0,0, "Fixed",0.06,0,0, "Fixed",0.05,0,0, "Fixed",0.05,0,0),
    ("视频游戏 Video Games", "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.14,0,0, "Fixed",0.14,0,0),
    ("游戏配件 Gaming Acc", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("大型家电 Large Appliances", "Threshold",0.15,50,0.05, "Threshold",0.15,50,0.05, "Threshold",0.12,250,0.06, "Threshold",0.12,250,0.06),
    ("小型家电 Small Appliances", "Fixed",0.13,0,0, "Fixed",0.13,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    # Beauty
    ("香水 Fragrance", "Fixed",0.14,0,0, "Fixed",0.14,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("彩妆-白牌 Makeup Generic", "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("彩妆-品牌 Makeup Branded", "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15, "Fixed",0.10,0,0, "Fixed",0.10,0,0),
    ("个护-通用 Personal Care Generic", "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("个护-品牌 Personal Care Branded", "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15, "Threshold",0.09,50,0.125, "Threshold",0.09,50,0.125),
    ("健康营养品≤50 Health≤50", "Fixed",0.08,0,0, "Fixed",0.08,0,0, "Fixed",0.085,0,0, "Fixed",0.085,0,0),
    ("健康营养品>50 Health>50", "Fixed",0.14,0,0, "Fixed",0.14,0,0, "Fixed",0.11,0,0, "Fixed",0.11,0,0),
    # Home
    ("清洁卫浴 Cleaning", "Fixed",0.09,0,0, "Fixed",0.09,0,0, "Fixed",0.09,0,0, "Fixed",0.09,0,0),
    ("厨具/床品/卫浴/装饰 Home", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.14,0,0, "Fixed",0.14,0,0),
    ("家具 Furniture", "Sliding",0.15,750,0.10, "Sliding",0.15,750,0.10, "Sliding",0.15,750,0.10, "Sliding",0.15,750,0.10),
    # Other
    ("运动户外 Sports", "Threshold",0.20,30,0.13, "Threshold",0.20,30,0.13, "Threshold",0.20,30,0.13, "Threshold",0.20,30,0.13),
    ("玩具 Toys", "Fixed",0.14,0,0, "Fixed",0.14,0,0, "Fixed",0.14,0,0, "Fixed",0.14,0,0),
    ("母婴 Baby Products", "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15, "Threshold",0.08,50,0.15),
    ("图书 Books", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.13,0,0, "Fixed",0.13,0,0),
    ("办公文具 Office Supplies", "Fixed",0.12,0,0, "Fixed",0.12,0,0, "Fixed",0.12,0,0, "Fixed",0.12,0,0),
    ("汽配电子 Auto Electronics", "Fixed",0.05,0,0, "Fixed",0.05,0,0, "Fixed",0.07,0,0, "Fixed",0.07,0,0),
    ("汽配零件 Auto Parts", "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0),
    ("杂货非食品 Grocery Non-Food", "Fixed",0.09,0,0, "Fixed",0.09,0,0, "Fixed",0.09,0,0, "Fixed",0.09,0,0),
    ("礼品卡 Gift Cards", "Fixed",0.05,0,0, "Fixed",0.05,0,0, "Fixed",0.10,0,0, "Fixed",0.10,0,0),
    ("宠物用品 Pet Supplies", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
    ("其他通用 General/Other", "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0, "Fixed",0.15,0,0),
]

# ============================================================
# FBN SIZE TIERS
# ============================================================
FBN_SIZE_TIERS = [
    # (name, max_l, max_w, max_h, pkg_wt_kg, max_chg_wt)
    ("Small Envelope", 20, 15, 1, 0.02, 0.1),
    ("Standard Envelope", 33, 23, 2.5, 0.04, 0.5),
    ("Large Envelope", 33, 23, 5, 0.04, 1.0),
    ("Standard Parcel", 45, 34, 26, 0.10, 12.0),
    ("Oversize", 130, 34, 26, 0.24, 30.0),
    ("Extra Oversize", 130, 130, 130, 0.24, 30.0),
    ("Bulky", 999, 999, 999, 0.40, 999),
]

# ============================================================
# FBN OUTBOUND FEE TABLES (expanded discrete rows)
# ============================================================
# (size_tier, max_weight, uae_low, uae_high) for UAE
FBN_UAE_FEES = [
    ("Small Envelope", 0.1, 5.0, 7.0),
    ("Standard Envelope", 0.1, 5.5, 7.5),
    ("Standard Envelope", 0.25, 6.0, 8.0),
    ("Standard Envelope", 0.5, 6.0, 8.0),
    ("Large Envelope", 1.0, 6.5, 8.5),
    ("Standard Parcel", 0.25, 7.0, 8.5),
    ("Standard Parcel", 0.5, 7.0, 9.0),
    ("Standard Parcel", 1.0, 8.0, 10.0),
    ("Standard Parcel", 1.5, 8.5, 10.5),
    ("Standard Parcel", 2.0, 9.0, 11.0),
    ("Standard Parcel", 3.0, 10.0, 12.0),
    ("Standard Parcel", 4.0, 11.0, 13.0),
    ("Standard Parcel", 5.0, 12.0, 14.0),
    ("Standard Parcel", 6.0, 13.0, 15.0),
    ("Standard Parcel", 7.0, 14.0, 16.0),
    ("Standard Parcel", 8.0, 15.0, 17.0),
    ("Standard Parcel", 9.0, 16.0, 18.0),
    ("Standard Parcel", 10.0, 17.0, 19.0),
    ("Standard Parcel", 11.0, 18.0, 20.0),
    ("Standard Parcel", 12.0, 19.0, 21.0),
    ("Oversize", 1.0, 10.0, 12.0),
    ("Oversize", 2.0, 11.0, 13.0),
    ("Oversize", 3.0, 12.0, 14.0),
    ("Oversize", 4.0, 13.0, 14.5),
    ("Oversize", 5.0, 14.0, 15.5),
    ("Oversize", 6.0, 15.0, 16.5),
    ("Oversize", 7.0, 16.0, 17.5),
    ("Oversize", 8.0, 17.0, 18.5),
    ("Oversize", 9.0, 18.0, 19.5),
    ("Oversize", 10.0, 19.0, 20.5),
    ("Oversize", 15.0, 23.5, 25.0),
    ("Oversize", 20.0, 28.0, 30.0),
    ("Oversize", 25.0, 33.0, 34.5),
    ("Oversize", 30.0, 37.5, 39.5),
    ("Extra Oversize", 1.0, 11.0, 13.0),
    ("Extra Oversize", 2.0, 12.0, 14.0),
    ("Extra Oversize", 3.0, 13.0, 15.0),
    ("Extra Oversize", 4.0, 14.0, 16.0),
    ("Extra Oversize", 5.0, 15.0, 17.0),
    ("Extra Oversize", 6.0, 16.0, 18.0),
    ("Extra Oversize", 7.0, 17.0, 19.0),
    ("Extra Oversize", 8.0, 18.0, 20.0),
    ("Extra Oversize", 9.0, 19.0, 21.0),
    ("Extra Oversize", 10.0, 20.0, 22.0),
    ("Extra Oversize", 15.0, 25.0, 27.0),
    ("Extra Oversize", 20.0, 30.0, 32.0),
    ("Extra Oversize", 25.0, 35.0, 37.0),
    ("Extra Oversize", 30.0, 40.0, 42.0),
    ("Bulky", 20.0, 33.0, 35.0),
    ("Bulky", 21.0, 35.0, 37.0),
    ("Bulky", 22.0, 37.0, 39.0),
    ("Bulky", 23.0, 39.0, 41.0),
    ("Bulky", 24.0, 41.0, 43.0),
    ("Bulky", 25.0, 43.0, 45.0),
    ("Bulky", 26.0, 45.0, 47.0),
    ("Bulky", 27.0, 47.0, 49.0),
    ("Bulky", 28.0, 49.0, 51.0),
    ("Bulky", 29.0, 51.0, 53.0),
    ("Bulky", 30.0, 53.0, 55.0),
]

FBN_KSA_FEES = [
    ("Small Envelope", 0.1, 5.5, 7.5),
    ("Standard Envelope", 0.1, 6.0, 8.0),
    ("Standard Envelope", 0.25, 6.0, 8.0),
    ("Standard Envelope", 0.5, 6.5, 8.5),
    ("Large Envelope", 1.0, 7.0, 9.0),
    ("Standard Parcel", 0.25, 7.0, 9.0),
    ("Standard Parcel", 0.5, 7.5, 9.5),
    ("Standard Parcel", 1.0, 8.0, 10.0),
    ("Standard Parcel", 1.5, 8.5, 11.5),
    ("Standard Parcel", 2.0, 9.0, 12.0),
    ("Standard Parcel", 3.0, 10.0, 13.0),
    ("Standard Parcel", 4.0, 11.0, 14.0),
    ("Standard Parcel", 5.0, 12.0, 15.0),
    ("Standard Parcel", 6.0, 13.0, 16.0),
    ("Standard Parcel", 7.0, 14.0, 17.0),
    ("Standard Parcel", 8.0, 15.0, 18.0),
    ("Standard Parcel", 9.0, 16.0, 19.0),
    ("Standard Parcel", 10.0, 17.0, 20.0),
    ("Standard Parcel", 11.0, 18.0, 21.0),
    ("Standard Parcel", 12.0, 19.0, 22.0),
    ("Oversize", 1.0, 10.0, 14.0),
    ("Oversize", 2.0, 11.0, 15.0),
    ("Oversize", 3.0, 12.0, 16.0),
    ("Oversize", 4.0, 13.0, 17.0),
    ("Oversize", 5.0, 14.0, 18.0),
    ("Oversize", 6.0, 15.0, 19.0),
    ("Oversize", 7.0, 16.0, 20.0),
    ("Oversize", 8.0, 17.0, 21.0),
    ("Oversize", 9.0, 18.0, 22.0),
    ("Oversize", 10.0, 19.0, 23.0),
    ("Oversize", 15.0, 24.0, 28.0),
    ("Oversize", 20.0, 29.0, 33.0),
    ("Oversize", 25.0, 34.0, 38.0),
    ("Oversize", 30.0, 39.0, 43.0),
    ("Extra Oversize", 4.0, 15.5, 18.5),
    ("Extra Oversize", 5.0, 16.5, 19.5),
    ("Extra Oversize", 6.0, 17.5, 20.5),
    ("Extra Oversize", 7.0, 18.5, 21.5),
    ("Extra Oversize", 8.0, 19.5, 22.5),
    ("Extra Oversize", 9.0, 20.5, 23.5),
    ("Extra Oversize", 10.0, 21.5, 24.5),
    ("Extra Oversize", 15.0, 25.0, 29.0),
    ("Extra Oversize", 20.0, 30.0, 34.0),
    ("Extra Oversize", 25.0, 35.0, 39.0),
    ("Extra Oversize", 30.0, 39.0, 44.0),
    ("Bulky", 20.0, 33.0, 37.0),
    ("Bulky", 21.0, 35.0, 39.0),
    ("Bulky", 22.0, 37.0, 41.0),
    ("Bulky", 23.0, 39.0, 43.0),
    ("Bulky", 24.0, 41.0, 45.0),
    ("Bulky", 25.0, 43.0, 47.0),
    ("Bulky", 26.0, 45.0, 49.0),
    ("Bulky", 27.0, 47.0, 51.0),
    ("Bulky", 28.0, 49.0, 53.0),
    ("Bulky", 29.0, 51.0, 55.0),
    ("Bulky", 30.0, 53.0, 57.0),
    ("Bulky", 35.0, 61.0, 65.0),
    ("Bulky", 40.0, 70.0, 75.0),
]

# ============================================================
# FBP DELIVERY FEE TABLE (expanded discrete rows)
# ============================================================
# (max_weight, uae_dropoff, uae_pickup, ksa_dropoff, ksa_pickup)
FBP_FEES = [
    (0.25, 12.0, 14.0, 17.0, 19.0),
    (0.50, 12.5, 14.5, 17.5, 19.5),
    (1.00, 13.0, 15.0, 18.5, 20.5),
    (1.50, 13.5, 15.5, 19.5, 21.5),
    (2.00, 14.0, 16.0, 20.5, 22.5),
    (3.00, 15.5, 17.5, 27.0, 29.0),
    (4.00, 16.5, 18.5, 28.0, 30.0),
    (5.00, 17.5, 19.5, 29.0, 31.0),
    (6.00, 18.5, 20.5, 30.0, 32.0),
    (7.00, 19.5, 21.5, 31.0, 33.0),
    (8.00, 20.5, 22.5, 32.0, 34.0),
    (9.00, 21.5, 23.5, 33.0, 35.0),
    (10.0, 22.5, 24.5, 34.0, 36.0),
    (15.0, 30.0, 32.0, 39.0, 41.0),
    (20.0, 34.0, 36.0, 43.0, 45.0),
    (25.0, 38.0, 40.0, 47.0, 49.0),
    (30.0, 42.0, 44.0, 51.0, 53.0),
]

# ============================================================
# GCC CUSTOMS REFERENCE TABLE
# ============================================================
GCC_CUSTOMS_REF = [
    ("电子/3C (第84-85章)", "5%", "标准税率"),
    ("服装/鞋类 (第61-64章)", "5%", "部分纺织品12%"),
    ("美妆/个护 (第33章)", "5%", "标准税率"),
    ("家居/厨具 (第39-76章)", "5%", "标准税率"),
    ("家具 (第94章)", "5%-15%", "⚠ 部分家具15%"),
    ("帐篷/遮阳篷", "15%", "户外品类注意"),
    ("部分纺织品", "12%", "约294项保护性商品"),
    ("玩具 (第95章)", "5%", "标准税率"),
    ("箱包 (第42章)", "5%", "标准税率"),
    ("手表 (第91章)", "5%", "标准税率"),
]

GCC_VAT_RATES = [
    ("KSA 沙特", "15%"),
    ("UAE 阿联酋", "5%"),
    ("Bahrain 巴林", "10%"),
    ("Oman 阿曼", "5%"),
    ("Kuwait 科威特", "0% (暂未实施)"),
    ("Qatar 卡塔尔", "0% (暂未实施)"),
]


# ============================================================
# BUILD CONFIG SHEET
# ============================================================
def build_config(wb):
    ws = wb.active
    ws.title = "Config"

    # --- Block 1: Global Variables ---
    r = 1
    style_cell(ws, r, 1, "全局变量 GLOBAL VARIABLES", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 5):
        style_cell(ws, r, c, fill=FILLS['config_title'])

    globals_data = [
        ("UAE_VAT", 0.05, "阿联酋增值税", "UAE_VAT"),
        ("KSA_VAT", 0.15, "沙特增值税", "KSA_VAT"),
        ("UAE_Exch (AED→CNY)", 1.96, "阿联酋汇率", "UAE_Exch"),
        ("KSA_Exch (SAR→CNY)", 1.92, "沙特汇率", "KSA_Exch"),
        ("UAE_Restock (AED)", 0, "退货上架费", "UAE_Restock"),
        ("KSA_Restock (SAR)", 0, "退货上架费", "KSA_Restock"),
        ("FBP_DeliveryMode", "Drop-off", "FBP交货方式切换", "FBP_DeliveryMode"),
        ("UAE_StorageRate", 1.5, "AED/CBF/月", "UAE_StorageRate"),
        ("KSA_StorageRate", 2.5, "SAR/CBF/月", "KSA_StorageRate"),
        ("USD_AED", 3.67, "美元/迪拉姆汇率", "USD_AED"),
        ("USD_SAR", 3.75, "美元/里亚尔汇率", "USD_SAR"),
        ("CBF_Divisor", 28317, "立方英尺换算", "CBF_Divisor"),
        ("GCC_DefaultDuty", 0.05, "GCC默认关税率5%", "GCC_DefaultDuty"),
        ("KSA_ClearanceFeeRate", 0.0015, "清关服务费率0.15%", "KSA_ClearanceFeeRate"),
        ("KSA_ClearanceFeeMin", 15, "清关服务费最低SAR", "KSA_ClearanceFeeMin"),
        ("KSA_ClearanceFeeMax", 500, "清关服务费最高SAR", "KSA_ClearanceFeeMax"),
        ("CrossBorderFeeRate", 0.05, "跨境费=5%", "CrossBorderFeeRate"),
    ]

    style_header_row(ws, r+1, 1, 4, ["参数名称", "值", "说明", "Named Range"], FILLS['header_mid'])
    for i, (name, val, desc, nr) in enumerate(globals_data):
        row = r + 2 + i
        style_cell(ws, row, 1, name, font=FONT_BOLD, fill=FILLS['editable'], alignment=ALIGN_LEFT)
        style_cell(ws, row, 2, val, fill=FILLS['editable'])
        style_cell(ws, row, 3, desc, alignment=ALIGN_LEFT)
        style_cell(ws, row, 4, nr, font=Font(name='Calibri', size=9, color='808080'))
        # Define named range
        dn = DefinedName(nr, attr_text=f"Config!$B${row}")
        wb.defined_names.add(dn)

    # FBP delivery mode dropdown
    dv_mode = DataValidation(type="list", formula1='"Drop-off,Pick-up"', allow_blank=False)
    dv_mode.error = "请选择 Drop-off 或 Pick-up"
    ws.add_data_validation(dv_mode)
    dv_mode.add(ws.cell(row=r+2+6, column=2))  # FBP_DeliveryMode cell

    set_col_width(ws, 1, 28)
    set_col_width(ws, 2, 14)
    set_col_width(ws, 3, 22)
    set_col_width(ws, 4, 22)

    # --- Block 2: Commission Table ---
    comm_start = r + 2 + len(globals_data) + 2  # leave a gap
    style_cell(ws, comm_start, 1, "佣金表 COMMISSION TABLE (FBN/FBP分列)", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 18):
        style_cell(ws, comm_start, c, fill=FILLS['config_title'])

    comm_headers = [
        "类目 Category",
        "UAE_FBN Type", "UAE_FBN Base", "UAE_FBN Limit", "UAE_FBN Tier",
        "UAE_FBP Type", "UAE_FBP Base", "UAE_FBP Limit", "UAE_FBP Tier",
        "KSA_FBN Type", "KSA_FBN Base", "KSA_FBN Limit", "KSA_FBN Tier",
        "KSA_FBP Type", "KSA_FBP Base", "KSA_FBP Limit", "KSA_FBP Tier",
    ]
    style_header_row(ws, comm_start+1, 1, 17, comm_headers, FILLS['header_mid'])

    comm_data_start = comm_start + 2
    for i, cat in enumerate(CATEGORIES):
        row = comm_data_start + i
        style_cell(ws, row, 1, cat[0], font=FONT_NORMAL, alignment=ALIGN_LEFT)
        for j in range(1, 17):
            val = cat[j]
            fmt = '0.0%' if isinstance(val, float) and val < 1 else None
            style_cell(ws, row, j+1, val, num_fmt=fmt)

    comm_data_end = comm_data_start + len(CATEGORIES) - 1
    # Named range for comm table
    wb.defined_names.add(DefinedName("Comm_Table", attr_text=f"Config!$A${comm_data_start}:$Q${comm_data_end}"))

    # Set commission column widths
    set_col_width(ws, 1, 32)
    for c in range(2, 18):
        set_col_width(ws, c, 13)

    # --- Block 3: FBN Size Tiers ---
    tier_start = comm_data_end + 3
    style_cell(ws, tier_start, 1, "FBN 尺寸分层定义 SIZE TIERS", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 7):
        style_cell(ws, tier_start, c, fill=FILLS['config_title'])

    tier_headers = ["层级名称", "最大L(cm)", "最大W(cm)", "最大H(cm)", "包装附加(kg)", "最大计费重(kg)"]
    style_header_row(ws, tier_start+1, 1, 6, tier_headers, FILLS['header_mid'])

    tier_data_start = tier_start + 2
    for i, t in enumerate(FBN_SIZE_TIERS):
        row = tier_data_start + i
        for j, val in enumerate(t):
            style_cell(ws, row, j+1, val, alignment=ALIGN_LEFT if j==0 else ALIGN_CENTER)

    tier_data_end = tier_data_start + len(FBN_SIZE_TIERS) - 1
    wb.defined_names.add(DefinedName("FBN_SizeTiers", attr_text=f"Config!$A${tier_data_start}:$F${tier_data_end}"))

    # --- Block 4: FBN UAE Fees ---
    fbn_uae_start = tier_data_end + 3
    style_cell(ws, fbn_uae_start, 1, "FBN 出库费 UAE (AED/件)", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 5):
        style_cell(ws, fbn_uae_start, c, fill=FILLS['config_title'])

    fbn_headers = ["尺寸层级", "最大重量(kg)", "ASP≤25", "ASP>25"]
    style_header_row(ws, fbn_uae_start+1, 1, 4, fbn_headers, FILLS['header_mid'])

    fbn_uae_data_start = fbn_uae_start + 2
    for i, f in enumerate(FBN_UAE_FEES):
        row = fbn_uae_data_start + i
        for j, val in enumerate(f):
            style_cell(ws, row, j+1, val, alignment=ALIGN_LEFT if j==0 else ALIGN_CENTER)

    fbn_uae_data_end = fbn_uae_data_start + len(FBN_UAE_FEES) - 1
    wb.defined_names.add(DefinedName("FBN_UAE_Fees", attr_text=f"Config!$A${fbn_uae_data_start}:$D${fbn_uae_data_end}"))

    # --- Block 5: FBN KSA Fees ---
    fbn_ksa_start = fbn_uae_data_end + 3
    style_cell(ws, fbn_ksa_start, 1, "FBN 出库费 KSA (SAR/件)", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 5):
        style_cell(ws, fbn_ksa_start, c, fill=FILLS['config_title'])

    style_header_row(ws, fbn_ksa_start+1, 1, 4, fbn_headers, FILLS['header_mid'])

    fbn_ksa_data_start = fbn_ksa_start + 2
    for i, f in enumerate(FBN_KSA_FEES):
        row = fbn_ksa_data_start + i
        for j, val in enumerate(f):
            style_cell(ws, row, j+1, val, alignment=ALIGN_LEFT if j==0 else ALIGN_CENTER)

    fbn_ksa_data_end = fbn_ksa_data_start + len(FBN_KSA_FEES) - 1
    wb.defined_names.add(DefinedName("FBN_KSA_Fees", attr_text=f"Config!$A${fbn_ksa_data_start}:$D${fbn_ksa_data_end}"))

    # --- Block 6: FBP Fees ---
    fbp_start = fbn_ksa_data_end + 3
    style_cell(ws, fbp_start, 1, "FBP 配送费 (Drop-off / Pick-up)", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 6):
        style_cell(ws, fbp_start, c, fill=FILLS['config_title'])

    fbp_headers = ["最大重量(kg)", "UAE Drop-off", "UAE Pick-up", "KSA Drop-off", "KSA Pick-up"]
    style_header_row(ws, fbp_start+1, 1, 5, fbp_headers, FILLS['header_mid'])

    fbp_data_start = fbp_start + 2
    for i, f in enumerate(FBP_FEES):
        row = fbp_data_start + i
        for j, val in enumerate(f):
            style_cell(ws, row, j+1, val)

    fbp_data_end = fbp_data_start + len(FBP_FEES) - 1
    wb.defined_names.add(DefinedName("FBP_Fees", attr_text=f"Config!$A${fbp_data_start}:$E${fbp_data_end}"))

    # --- Block 7: FBP Reimbursement ---
    reimb_start = fbp_data_end + 3
    style_cell(ws, reimb_start, 1, "FBP 运费补贴规则 REIMBURSEMENT", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 4):
        style_cell(ws, reimb_start, c, fill=FILLS['config_title'])

    reimb_headers = ["国家/条件", "订单金额范围", "补贴金额"]
    style_header_row(ws, reimb_start+1, 1, 3, reimb_headers, FILLS['header_mid'])

    reimb_data = [
        ("UAE", "< 100 AED", "10 AED"),
        ("UAE", "≥ 100 AED", "0"),
        ("KSA", "< 100 SAR", "12 SAR"),
        ("KSA", "100-500 SAR", "6 SAR"),
        ("KSA", "> 500 SAR", "0"),
    ]
    for i, (c, rng, amt) in enumerate(reimb_data):
        row = reimb_start + 2 + i
        style_cell(ws, row, 1, c, alignment=ALIGN_LEFT)
        style_cell(ws, row, 2, rng, alignment=ALIGN_LEFT)
        style_cell(ws, row, 3, amt, alignment=ALIGN_LEFT)

    # --- Block 8: GCC Customs Reference ---
    customs_start = reimb_start + 2 + len(reimb_data) + 2
    style_cell(ws, customs_start, 1, "GCC 关税税率参考表 (仅供查阅)", font=FONT_TITLE, fill=FILLS['config_title'])
    for c in range(2, 4):
        style_cell(ws, customs_start, c, fill=FILLS['config_title'])

    customs_headers = ["品类/HS编码章节", "关税税率", "备注"]
    style_header_row(ws, customs_start+1, 1, 3, customs_headers, FILLS['header_mid'])

    for i, (cat, rate, note) in enumerate(GCC_CUSTOMS_REF):
        row = customs_start + 2 + i
        style_cell(ws, row, 1, cat, alignment=ALIGN_LEFT)
        style_cell(ws, row, 2, rate)
        style_cell(ws, row, 3, note, alignment=ALIGN_LEFT)

    # --- Block 9: GCC VAT Rates ---
    vat_start = customs_start + 2 + len(GCC_CUSTOMS_REF) + 2
    style_cell(ws, vat_start, 1, "GCC 各国 VAT 税率", font=FONT_TITLE, fill=FILLS['config_title'])
    style_cell(ws, vat_start, 2, fill=FILLS['config_title'])

    vat_headers = ["国家", "VAT税率"]
    style_header_row(ws, vat_start+1, 1, 2, vat_headers, FILLS['header_mid'])

    for i, (country, rate) in enumerate(GCC_VAT_RATES):
        row = vat_start + 2 + i
        style_cell(ws, row, 1, country, alignment=ALIGN_LEFT)
        style_cell(ws, row, 2, rate)

    # Freeze panes
    ws.freeze_panes = 'A2'

    return ws


# ============================================================
# BUILD PRICING SHEET (UAE or KSA) — V2 layout
# ============================================================
def build_pricing_sheet(wb, country="UAE"):
    """Build a pricing sheet for UAE or KSA with profit summary right after input"""
    ws = wb.create_sheet(title=f"{country}_Pricing")

    is_uae = (country == "UAE")
    currency = "AED" if is_uae else "SAR"
    vat_ref = "UAE_VAT" if is_uae else "KSA_VAT"
    exch_ref = "UAE_Exch" if is_uae else "KSA_Exch"
    restock_ref = "UAE_Restock" if is_uae else "KSA_Restock"

    # Commission column references in Comm_Table (1-indexed within the named range)
    fbn_comm_type_col = 2 if is_uae else 10
    fbp_comm_type_col = 6 if is_uae else 14
    fbn_fee_ref = "FBN_UAE_Fees" if is_uae else "FBN_KSA_Fees"
    fbp_dropoff_col = 2 if is_uae else 4
    fbp_pickup_col = 3 if is_uae else 5

    # ================================================================
    # COLUMN LAYOUT MAP — all positions defined here, formulas use these
    # ================================================================
    # Zone 1: Input (A-N, 14 cols)
    C_SKU, C_NAME, C_CAT, C_PURCHASE = 1, 2, 3, 4
    C_SEA_HEAD, C_DM_HEAD_RATE, C_DM_OPS_FEE = 5, 6, 7
    C_L, C_W, C_H, C_WT = 8, 9, 10, 11
    C_PRICE, C_ADS_RATE, C_RET_RATE = 12, 13, 14

    # Zone 2: Profit Summary (O-X, 10 cols)
    C_FBN_PROFIT, C_FBN_MARGIN = 15, 16
    C_OW_PROFIT, C_OW_MARGIN = 17, 18
    C_DM_PROFIT, C_DM_MARGIN = 19, 20
    C_GL_PROFIT, C_GL_MARGIN = 21, 22
    C_BEST_MODE, C_BEST_VAL = 23, 24

    # Zone 3: Shared Calc (Y-AG, 9 cols)
    C_DUTY_RATE = 25
    C_VOL_WT, C_FBP_CW = 26, 27
    C_FBN_TIER, C_FBN_PKG, C_FBN_CW = 28, 29, 30
    C_VAT, C_FBN_COMM, C_FBP_COMM = 31, 32, 33

    # Zone 4: FBN Details (AH-AO, 8 cols)
    C_FBN_HEAD, C_FBN_FEE = 34, 35
    C_FBN_CVAT, C_FBN_ADS, C_FBN_RETFEE = 36, 37, 38
    C_FBN_RETADM, C_FBN_LOGLOSS, C_FBN_TOTAL = 39, 40, 41

    # Zone 5: OW-FBP Details (AP-AW, 8 cols)
    C_OW_HEAD, C_OW_DELFEE, C_OW_REIMB = 42, 43, 44
    C_OW_CVAT, C_OW_ADS = 45, 46
    C_OW_RETADM, C_OW_LOGLOSS, C_OW_TOTAL = 47, 48, 49

    # Zone 6: DM-FBP Details (AX-BE, 8 cols)
    C_DM_HEAD, C_DM_DELFEE, C_DM_REIMB = 50, 51, 52
    C_DM_CVAT, C_DM_ADS = 53, 54
    C_DM_RETADM, C_DM_LOGLOSS, C_DM_TOTAL = 55, 56, 57

    # Zone 7: Global Details (BF-BQ, 12 cols)
    C_GL_CIF, C_GL_DUTY, C_GL_CLEAR = 58, 59, 60
    C_GL_IMPVAT, C_GL_KSAPRICE = 61, 62
    C_GL_CROSSFEE, C_GL_COMM, C_GL_CVAT = 63, 64, 65
    C_GL_ADS, C_GL_RETADM, C_GL_LOGLOSS, C_GL_TOTAL = 66, 67, 68, 69

    TOTAL_COLS = 69

    # Helper: get column letter by position
    def cl(c):
        return get_column_letter(c)

    # Zone ranges for styling
    ZONE_RANGES = {
        'input': (1, 14),
        'summary': (15, 24),
        'shared': (25, 33),
        'fbn': (34, 41),
        'ow_fbp': (42, 49),
        'dm_fbp': (50, 57),
        'global': (58, 69),
    }

    ZONE_TITLES = {
        'input': '📝 产品信息输入区',
        'summary': '🟣 利润总览',
        'shared': '📊 共享计算',
        'fbn': '🔵 FBN 官方仓',
        'ow_fbp': '🟠 海外仓FBP',
        'dm_fbp': '🟢 直邮FBP',
        'global': '🔴 Global跨境 (UAE→KSA)',
    }

    # Row 1: Zone title bars
    for zone, (start, end) in ZONE_RANGES.items():
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        style_cell(ws, 1, start, ZONE_TITLES[zone], font=FONT_TITLE,
                   fill=PatternFill(start_color=COLORS[zone], end_color=COLORS[zone], fill_type='solid'))

    # Row 2: Column headers
    input_headers = [
        "SKU编号", "品名", "类目 Category", "采购价(CNY)",
        "海运头程(CNY/m³)", "直邮头程(CNY/kg)", "直邮操作费(CNY/件)",
        "长L(cm)", "宽W(cm)", "高H(cm)", "实重(kg)",
        f"售价({currency})", "广告占比(%)", "退货率(%)"
    ]
    style_header_row(ws, 2, 1, 14, input_headers, FILLS['input'], font=FONT_BOLD)

    summary_headers = [
        "FBN净利(CNY)", "FBN利润率",
        "海外仓FBP净利(CNY)", "海外仓FBP利润率",
        "直邮FBP净利(CNY)", "直邮FBP利润率",
        "Global净利(CNY)", "Global利润率",
        "最优模式", "最优利润(CNY)"
    ]
    style_header_row(ws, 2, 15, 24, summary_headers, FILLS['summary'], font=FONT_BOLD)

    shared_headers = [
        "关税税率(%)",
        "体积重(kg)", "FBP计费重(kg)", "FBN尺寸层级", "FBN包装附加(kg)",
        "FBN计费重(kg)", f"VAT({currency})", f"FBN佣金({currency})", f"FBP佣金({currency})"
    ]
    style_header_row(ws, 2, 25, 33, shared_headers, FILLS['shared'], font=FONT_BOLD)

    fbn_headers = [
        "头程(CNY)", f"出库费({currency})", f"佣金VAT({currency})", f"广告费({currency})",
        f"退货费({currency})", "退货摊销", "物流损耗", "总成本(CNY)"
    ]
    style_header_row(ws, 2, 34, 41, fbn_headers, FILLS['fbn'], font=FONT_BOLD)

    ow_headers = [
        "头程(CNY)", f"FBP配送费({currency})", f"补贴({currency})",
        f"佣金VAT({currency})", f"广告费({currency})",
        "退货摊销", "物流损耗", "总成本(CNY)"
    ]
    style_header_row(ws, 2, 42, 49, ow_headers, FILLS['ow_fbp'], font=FONT_BOLD)

    dm_headers = [
        "头程(CNY)", f"FBP配送费({currency})", f"补贴({currency})",
        f"佣金VAT({currency})", f"广告费({currency})",
        "退货摊销", "物流损耗", "总成本(CNY)"
    ]
    style_header_row(ws, 2, 50, 57, dm_headers, FILLS['dm_fbp'], font=FONT_BOLD)

    global_headers = [
        "CIF税前价(SAR)", "关税(SAR)", "清关服务费(SAR)", "进口VAT(SAR)",
        "KSA自动售价(SAR)", "跨境费(AED)", "佣金(AED)", "佣金VAT(AED)",
        "广告费(AED)", "退货摊销", "物流损耗(SAR)", "总成本(CNY)"
    ]
    style_header_row(ws, 2, 58, 69, global_headers, FILLS['global'], font=FONT_BOLD)

    # Set column widths
    widths = {1: 10, 2: 14, 3: 28, 4: 12, 5: 16, 6: 14, 7: 16,
              8: 8, 9: 8, 10: 8, 11: 8, 12: 12, 13: 12, 14: 12}
    for c in range(15, TOTAL_COLS + 1):
        widths[c] = 14
    widths[C_CAT] = 28
    widths[C_BEST_MODE] = 14
    for c, w in widths.items():
        set_col_width(ws, c, w)

    # --- DATA ROWS (rows 3-22, 20 product rows) ---
    num_rows = 20

    for row in range(3, 3 + num_rows):
        r = row

        # ========== Zone 1: Input defaults ==========
        style_cell(ws, r, C_PURCHASE, 5, fill=FILLS['editable'], num_fmt='0.00')
        style_cell(ws, r, C_SEA_HEAD, 800, fill=FILLS['editable'])
        style_cell(ws, r, C_DM_HEAD_RATE, 45, fill=FILLS['editable'])
        style_cell(ws, r, C_DM_OPS_FEE, 2, fill=FILLS['editable'])
        style_cell(ws, r, C_L, 5, fill=FILLS['editable'])
        style_cell(ws, r, C_W, 5, fill=FILLS['editable'])
        style_cell(ws, r, C_H, 1, fill=FILLS['editable'])
        style_cell(ws, r, C_WT, 0.01, fill=FILLS['editable'])
        style_cell(ws, r, C_PRICE, 15, fill=FILLS['editable'])
        style_cell(ws, r, C_ADS_RATE, 0.05, fill=FILLS['editable'], num_fmt='0%')
        style_cell(ws, r, C_RET_RATE, 0.05, fill=FILLS['editable'], num_fmt='0%')

        # ========== Zone 3: Shared Calculations ==========
        # Duty rate default
        style_cell(ws, r, C_DUTY_RATE, None, num_fmt='0%')
        ws.cell(row=r, column=C_DUTY_RATE).value = '=GCC_DefaultDuty'

        # Volume weight
        ws.cell(row=r, column=C_VOL_WT).value = (
            f'={cl(C_L)}{r}*{cl(C_W)}{r}*{cl(C_H)}{r}/5000'
        )
        # FBP chargeable weight
        ws.cell(row=r, column=C_FBP_CW).value = (
            f'=MAX({cl(C_WT)}{r},{cl(C_VOL_WT)}{r})'
        )

        # FBN Size Tier (nested IF with MEDIAN/MAX/MIN — WPS compatible)
        ws.cell(row=r, column=C_FBN_TIER).value = gen_size_tier_formula(
            r, cl(C_L), cl(C_W), cl(C_H)
        )

        # FBN packaging weight
        ws.cell(row=r, column=C_FBN_PKG).value = (
            f'=IFERROR(VLOOKUP({cl(C_FBN_TIER)}{r},FBN_SizeTiers,5,FALSE),0)'
        )

        # FBN chargeable weight (capped by tier max)
        ws.cell(row=r, column=C_FBN_CW).value = (
            f'=IFERROR(MIN(MAX({cl(C_WT)}{r},{cl(C_VOL_WT)}{r})+{cl(C_FBN_PKG)}{r},'
            f'VLOOKUP({cl(C_FBN_TIER)}{r},FBN_SizeTiers,6,FALSE)),0)'
        )

        # VAT
        ws.cell(row=r, column=C_VAT).value = (
            f'={cl(C_PRICE)}{r}/(1+{vat_ref})*{vat_ref}'
        )

        # FBN Commission (WPS compatible — no LET)
        ws.cell(row=r, column=C_FBN_COMM).value = gen_comm_formula(
            r, cl(C_CAT), cl(C_PRICE),
            fbn_comm_type_col, fbn_comm_type_col+1, fbn_comm_type_col+2, fbn_comm_type_col+3
        )

        # FBP Commission (WPS compatible — no LET)
        ws.cell(row=r, column=C_FBP_COMM).value = gen_comm_formula(
            r, cl(C_CAT), cl(C_PRICE),
            fbp_comm_type_col, fbp_comm_type_col+1, fbp_comm_type_col+2, fbp_comm_type_col+3
        )

        # ========== Zone 4: FBN Details ==========
        # Head cost (volume)
        ws.cell(row=r, column=C_FBN_HEAD).value = (
            f'=({cl(C_L)}{r}*{cl(C_W)}{r}*{cl(C_H)}{r}/1000000)*{cl(C_SEA_HEAD)}{r}'
        )

        # FBN Outbound Fee — nested IF (WPS compatible, generated from data)
        fee_data = FBN_UAE_FEES if is_uae else FBN_KSA_FEES
        ws.cell(row=r, column=C_FBN_FEE).value = gen_fbn_fee_formula(
            fee_data, r, cl(C_FBN_TIER), cl(C_FBN_CW), cl(C_PRICE)
        )

        # Commission VAT
        ws.cell(row=r, column=C_FBN_CVAT).value = f'={cl(C_FBN_COMM)}{r}*{vat_ref}'
        # Ads
        ws.cell(row=r, column=C_FBN_ADS).value = f'={cl(C_PRICE)}{r}*{cl(C_ADS_RATE)}{r}'
        # Unit return fee
        ws.cell(row=r, column=C_FBN_RETFEE).value = f'=MIN(15,{cl(C_FBN_COMM)}{r}*0.2)'
        # Return admin
        ws.cell(row=r, column=C_FBN_RETADM).value = f'={cl(C_FBN_RETFEE)}{r}*{cl(C_RET_RATE)}{r}'
        # Log loss (outbound fee + restock) × return rate
        ws.cell(row=r, column=C_FBN_LOGLOSS).value = (
            f'=({cl(C_FBN_FEE)}{r}+{restock_ref})*{cl(C_RET_RATE)}{r}'
        )
        # Total cost
        ws.cell(row=r, column=C_FBN_TOTAL).value = (
            f'={cl(C_PURCHASE)}{r}+{cl(C_FBN_HEAD)}{r}+'
            f'({cl(C_VAT)}{r}+{cl(C_FBN_COMM)}{r}+{cl(C_FBN_CVAT)}{r}+{cl(C_FBN_FEE)}{r}'
            f'+{cl(C_FBN_ADS)}{r}+{cl(C_FBN_RETADM)}{r}+{cl(C_FBN_LOGLOSS)}{r})*{exch_ref}'
        )

        # ========== Zone 5: OW-FBP Details ==========
        # Head cost (same as FBN, volume based)
        ws.cell(row=r, column=C_OW_HEAD).value = (
            f'=({cl(C_L)}{r}*{cl(C_W)}{r}*{cl(C_H)}{r}/1000000)*{cl(C_SEA_HEAD)}{r}'
        )
        # FBP delivery fee (WPS compatible — no LET)
        ws.cell(row=r, column=C_OW_DELFEE).value = gen_fbp_fee_formula(
            r, cl(C_FBP_CW), fbp_dropoff_col, fbp_pickup_col
        )
        # Reimbursement
        if is_uae:
            ws.cell(row=r, column=C_OW_REIMB).value = f'=IF({cl(C_PRICE)}{r}<100,10,0)'
        else:
            ws.cell(row=r, column=C_OW_REIMB).value = f'=IF({cl(C_PRICE)}{r}<100,12,IF({cl(C_PRICE)}{r}<=500,6,0))'
        # Comm VAT
        ws.cell(row=r, column=C_OW_CVAT).value = f'={cl(C_FBP_COMM)}{r}*{vat_ref}'
        # Ads
        ws.cell(row=r, column=C_OW_ADS).value = f'={cl(C_PRICE)}{r}*{cl(C_ADS_RATE)}{r}'
        # Return admin
        ws.cell(row=r, column=C_OW_RETADM).value = f'=MIN(15,{cl(C_FBP_COMM)}{r}*0.2)*{cl(C_RET_RATE)}{r}'
        # Log loss (head sunk, only delivery lost)
        ws.cell(row=r, column=C_OW_LOGLOSS).value = (
            f'=({cl(C_OW_DELFEE)}{r}+{restock_ref})*{cl(C_RET_RATE)}{r}'
        )
        # Total cost
        ws.cell(row=r, column=C_OW_TOTAL).value = (
            f'={cl(C_PURCHASE)}{r}+{cl(C_OW_HEAD)}{r}+'
            f'({cl(C_VAT)}{r}+{cl(C_FBP_COMM)}{r}+{cl(C_OW_CVAT)}{r}+{cl(C_OW_DELFEE)}{r}'
            f'+{cl(C_OW_ADS)}{r}+{cl(C_OW_RETADM)}{r}+{cl(C_OW_LOGLOSS)}{r})*{exch_ref}'
        )

        # ========== Zone 6: DM-FBP Details ==========
        # Head cost (weight-based, per shipment) — uses input cols F(直邮头程) and G(操作费)
        ws.cell(row=r, column=C_DM_HEAD).value = (
            f'=CEILING({cl(C_FBP_CW)}{r},0.1)*{cl(C_DM_HEAD_RATE)}{r}+{cl(C_DM_OPS_FEE)}{r}'
        )
        # FBP delivery fee (WPS compatible — no LET)
        ws.cell(row=r, column=C_DM_DELFEE).value = gen_fbp_fee_formula(
            r, cl(C_FBP_CW), fbp_dropoff_col, fbp_pickup_col
        )
        # Reimbursement
        if is_uae:
            ws.cell(row=r, column=C_DM_REIMB).value = f'=IF({cl(C_PRICE)}{r}<100,10,0)'
        else:
            ws.cell(row=r, column=C_DM_REIMB).value = f'=IF({cl(C_PRICE)}{r}<100,12,IF({cl(C_PRICE)}{r}<=500,6,0))'
        # Comm VAT
        ws.cell(row=r, column=C_DM_CVAT).value = f'={cl(C_FBP_COMM)}{r}*{vat_ref}'
        # Ads
        ws.cell(row=r, column=C_DM_ADS).value = f'={cl(C_PRICE)}{r}*{cl(C_ADS_RATE)}{r}'
        # Return admin
        ws.cell(row=r, column=C_DM_RETADM).value = f'=MIN(15,{cl(C_FBP_COMM)}{r}*0.2)*{cl(C_RET_RATE)}{r}'
        # Log loss (head cost IS lost per shipment — convert CNY head to local currency)
        ws.cell(row=r, column=C_DM_LOGLOSS).value = (
            f'=({cl(C_DM_HEAD)}{r}/{exch_ref}+{cl(C_DM_DELFEE)}{r}+{restock_ref})*{cl(C_RET_RATE)}{r}'
        )
        # Total cost
        ws.cell(row=r, column=C_DM_TOTAL).value = (
            f'={cl(C_PURCHASE)}{r}+{cl(C_DM_HEAD)}{r}+'
            f'({cl(C_VAT)}{r}+{cl(C_FBP_COMM)}{r}+{cl(C_DM_CVAT)}{r}+{cl(C_DM_DELFEE)}{r}'
            f'+{cl(C_DM_ADS)}{r}+{cl(C_DM_RETADM)}{r}+{cl(C_DM_LOGLOSS)}{r})*{exch_ref}'
        )

        # ========== Zone 7: Global Details (UAE→KSA) ==========
        # CIF pre-tax price in SAR
        ws.cell(row=r, column=C_GL_CIF).value = (
            f'=({cl(C_PRICE)}{r}/1.05/USD_AED)*USD_SAR'
        )
        # Customs duty
        ws.cell(row=r, column=C_GL_DUTY).value = (
            f'={cl(C_GL_CIF)}{r}*{cl(C_DUTY_RATE)}{r}'
        )
        # Clearance service fee
        ws.cell(row=r, column=C_GL_CLEAR).value = (
            f'=MAX(KSA_ClearanceFeeMin,MIN(KSA_ClearanceFeeMax,{cl(C_GL_CIF)}{r}*KSA_ClearanceFeeRate))'
        )
        # Import VAT
        ws.cell(row=r, column=C_GL_IMPVAT).value = (
            f'=({cl(C_GL_CIF)}{r}+{cl(C_GL_DUTY)}{r})*KSA_VAT'
        )
        # KSA auto price
        ws.cell(row=r, column=C_GL_KSAPRICE).value = (
            f'=({cl(C_GL_CIF)}{r}+{cl(C_GL_DUTY)}{r}+{cl(C_GL_CLEAR)}{r})*(1+KSA_VAT)'
        )
        # Cross-border fee (AED)
        ws.cell(row=r, column=C_GL_CROSSFEE).value = f'={cl(C_PRICE)}{r}*CrossBorderFeeRate'
        # Commission (uses FBN commission — Global based on FBN)
        ws.cell(row=r, column=C_GL_COMM).value = f'={cl(C_FBN_COMM)}{r}'
        # Commission VAT
        ws.cell(row=r, column=C_GL_CVAT).value = f'={cl(C_GL_COMM)}{r}*UAE_VAT'
        # Ads
        ws.cell(row=r, column=C_GL_ADS).value = f'={cl(C_PRICE)}{r}*{cl(C_ADS_RATE)}{r}'
        # Return admin
        ws.cell(row=r, column=C_GL_RETADM).value = (
            f'=MIN(15,{cl(C_GL_COMM)}{r}*0.2)*{cl(C_RET_RATE)}{r}'
        )
        # Log loss (catastrophic — import VAT + customs + clearance all unrecoverable)
        ws.cell(row=r, column=C_GL_LOGLOSS).value = (
            f'=({cl(C_GL_IMPVAT)}{r}+{cl(C_GL_DUTY)}{r}+{cl(C_GL_CLEAR)}{r})*{cl(C_RET_RATE)}{r}'
        )
        # Total cost = purchase + FBN head + (cross fee + comm + commVAT + ads + ret + FBN outbound)*UAE_Exch + log_loss*KSA_Exch
        ws.cell(row=r, column=C_GL_TOTAL).value = (
            f'={cl(C_PURCHASE)}{r}+{cl(C_FBN_HEAD)}{r}+'
            f'({cl(C_GL_CROSSFEE)}{r}+{cl(C_GL_COMM)}{r}+{cl(C_GL_CVAT)}{r}'
            f'+{cl(C_GL_ADS)}{r}+{cl(C_GL_RETADM)}{r}+{cl(C_FBN_FEE)}{r})*UAE_Exch'
            f'+({cl(C_GL_LOGLOSS)}{r})*KSA_Exch'
        )

        # ========== Zone 2: Profit Summary (forward-references detail zones) ==========
        # FBN net profit
        ws.cell(row=r, column=C_FBN_PROFIT).value = (
            f'={cl(C_PRICE)}{r}*{exch_ref}-{cl(C_FBN_TOTAL)}{r}'
        )
        ws.cell(row=r, column=C_FBN_PROFIT).number_format = '0.00'
        # FBN margin
        ws.cell(row=r, column=C_FBN_MARGIN).value = (
            f'=IF({cl(C_PRICE)}{r}>0,{cl(C_FBN_PROFIT)}{r}/({cl(C_PRICE)}{r}*{exch_ref}),0)'
        )
        ws.cell(row=r, column=C_FBN_MARGIN).number_format = '0.00%'

        # OW-FBP net profit (with reimbursement as positive revenue)
        ws.cell(row=r, column=C_OW_PROFIT).value = (
            f'=({cl(C_PRICE)}{r}+{cl(C_OW_REIMB)}{r})*{exch_ref}-{cl(C_OW_TOTAL)}{r}'
        )
        ws.cell(row=r, column=C_OW_PROFIT).number_format = '0.00'
        # OW-FBP margin
        ws.cell(row=r, column=C_OW_MARGIN).value = (
            f'=IF({cl(C_PRICE)}{r}>0,{cl(C_OW_PROFIT)}{r}/({cl(C_PRICE)}{r}*{exch_ref}),0)'
        )
        ws.cell(row=r, column=C_OW_MARGIN).number_format = '0.00%'

        # DM-FBP net profit
        ws.cell(row=r, column=C_DM_PROFIT).value = (
            f'=({cl(C_PRICE)}{r}+{cl(C_DM_REIMB)}{r})*{exch_ref}-{cl(C_DM_TOTAL)}{r}'
        )
        ws.cell(row=r, column=C_DM_PROFIT).number_format = '0.00'
        # DM-FBP margin
        ws.cell(row=r, column=C_DM_MARGIN).value = (
            f'=IF({cl(C_PRICE)}{r}>0,{cl(C_DM_PROFIT)}{r}/({cl(C_PRICE)}{r}*{exch_ref}),0)'
        )
        ws.cell(row=r, column=C_DM_MARGIN).number_format = '0.00%'

        # Global net profit (seller receives CIF SAR value)
        ws.cell(row=r, column=C_GL_PROFIT).value = (
            f'={cl(C_GL_CIF)}{r}*KSA_Exch-{cl(C_GL_TOTAL)}{r}'
        )
        ws.cell(row=r, column=C_GL_PROFIT).number_format = '0.00'
        # Global margin
        ws.cell(row=r, column=C_GL_MARGIN).value = (
            f'=IF({cl(C_GL_CIF)}{r}>0,{cl(C_GL_PROFIT)}{r}/({cl(C_GL_CIF)}{r}*KSA_Exch),0)'
        )
        ws.cell(row=r, column=C_GL_MARGIN).number_format = '0.00%'

        # Best mode
        ws.cell(row=r, column=C_BEST_MODE).value = (
            f'=IF(MAX({cl(C_FBN_PROFIT)}{r},{cl(C_OW_PROFIT)}{r},{cl(C_DM_PROFIT)}{r},{cl(C_GL_PROFIT)}{r})'
            f'={cl(C_FBN_PROFIT)}{r},"FBN",'
            f'IF(MAX({cl(C_FBN_PROFIT)}{r},{cl(C_OW_PROFIT)}{r},{cl(C_DM_PROFIT)}{r},{cl(C_GL_PROFIT)}{r})'
            f'={cl(C_OW_PROFIT)}{r},"海外仓FBP",'
            f'IF(MAX({cl(C_FBN_PROFIT)}{r},{cl(C_OW_PROFIT)}{r},{cl(C_DM_PROFIT)}{r},{cl(C_GL_PROFIT)}{r})'
            f'={cl(C_DM_PROFIT)}{r},"直邮FBP","Global")))'
        )
        # Best profit
        ws.cell(row=r, column=C_BEST_VAL).value = (
            f'=MAX({cl(C_FBN_PROFIT)}{r},{cl(C_OW_PROFIT)}{r},{cl(C_DM_PROFIT)}{r},{cl(C_GL_PROFIT)}{r})'
        )
        ws.cell(row=r, column=C_BEST_VAL).number_format = '0.00'

        # ========== Number formats for detail zones ==========
        for col in [C_VAT, C_FBN_COMM, C_FBP_COMM]:
            ws.cell(row=r, column=col).number_format = '0.00'
        for col in range(C_FBN_HEAD, C_FBN_TOTAL + 1):
            ws.cell(row=r, column=col).number_format = '0.00'
        for col in range(C_OW_HEAD, C_OW_TOTAL + 1):
            ws.cell(row=r, column=col).number_format = '0.00'
        for col in range(C_DM_HEAD, C_DM_TOTAL + 1):
            ws.cell(row=r, column=col).number_format = '0.00'
        for col in range(C_GL_CIF, C_GL_TOTAL + 1):
            ws.cell(row=r, column=col).number_format = '0.00'

        # ========== Zone background colors ==========
        for zone, (start, end) in ZONE_RANGES.items():
            if zone == 'input':
                continue
            fill = PatternFill(start_color=COLORS[zone], end_color=COLORS[zone], fill_type='solid')
            for col in range(start, end + 1):
                ws.cell(row=r, column=col).fill = fill

    # Conditional formatting for profit & margin columns
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    red_font = Font(color='9C0006')

    profit_margin_cols = [
        C_FBN_PROFIT, C_FBN_MARGIN, C_OW_PROFIT, C_OW_MARGIN,
        C_DM_PROFIT, C_DM_MARGIN, C_GL_PROFIT, C_GL_MARGIN, C_BEST_VAL
    ]
    for col in profit_margin_cols:
        col_letter = cl(col)
        range_str = f'{col_letter}3:{col_letter}{2 + num_rows}'
        ws.conditional_formatting.add(range_str,
            CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
        ws.conditional_formatting.add(range_str,
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))

    # Category dropdown validation
    if "Comm_Table" in wb.defined_names:
        dn = wb.defined_names["Comm_Table"]
        ref = str(dn.attr_text)
        parts = ref.replace("Config!", "").replace("$", "").split(":")
        start_row_num = ''.join(filter(str.isdigit, parts[0]))
        end_row_num = ''.join(filter(str.isdigit, parts[1]))
        dv_cat = DataValidation(
            type="list",
            formula1=f"Config!$A${start_row_num}:$A${end_row_num}",
            allow_blank=True
        )
        dv_cat.error = "请从佣金表中选择类目"
        ws.add_data_validation(dv_cat)
        for row in range(3, 3 + num_rows):
            dv_cat.add(ws.cell(row=row, column=C_CAT))

    # Make input zone cells editable-colored
    for row in range(3, 3 + num_rows):
        for col in [C_SKU, C_NAME, C_CAT]:
            ws.cell(row=row, column=col).fill = FILLS['editable']

    # Freeze panes (freeze row 2 and first 14 input columns)
    ws.freeze_panes = f'{cl(C_FBN_PROFIT)}3'

    return ws


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()

    # Build Config sheet
    build_config(wb)

    # Build Pricing sheets
    build_pricing_sheet(wb, "UAE")
    build_pricing_sheet(wb, "KSA")

    # Save
    output_path = "/Users/a16/Documents/claude noon/Noon-V6-定价表.xlsx"
    wb.save(output_path)
    print(f"✅ Excel文件已生成: {output_path}")
    print(f"   - Config Sheet: 全局变量 + 佣金表({len(CATEGORIES)}个类目) + FBN尺寸定义 + FBN出库费(UAE:{len(FBN_UAE_FEES)}行/KSA:{len(FBN_KSA_FEES)}行) + FBP配送费({len(FBP_FEES)}行) + GCC关税参考")
    print(f"   - UAE_Pricing: 4模式(FBN/海外仓FBP/直邮FBP/Global) × 20行")
    print(f"   - KSA_Pricing: 4模式(FBN/海外仓FBP/直邮FBP/Global) × 20行")

if __name__ == "__main__":
    main()
